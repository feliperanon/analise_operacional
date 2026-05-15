# -*- coding: utf-8 -*-
"""Rotas de BI de Entregas e Devoluções."""

from calendar import monthrange
from datetime import datetime, timedelta, date
import math
from typing import Any, Callable, Dict, List, Optional
import io
import time
import csv
import statistics
import json
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Request, Depends, Query
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlmodel import Session, select
from sqlalchemy import func, or_, and_

import models
from database import get_session
from route_duration import route_duration_minutes, route_duration_minutes_mobile_only
from devolucao_kpi_canonical import (
    build_mes_fim_projecao_pct_financeiro,
    counts_devolucao_rotas_concluidas,
    devolucao_competencia_in_period,
    devolucao_competencia_iso,
    is_encerramento_tardio_automatico_return,
    pct_devolucao_sobre_rotas_concluidas,
    pct_valor_devolvido_sobre_base_rotas,
    route_base_financeiro_kpi,
    route_competencia_in_period,
    route_competencia_operacional_iso,
)
from devolucao_perda_labels import (
    canonical_responsabilidade_for_macro_loss as _canonical_responsabilidade_for_macro_loss,
    classify_macro_cause as _classify_macro_cause,
    macro_loss_label as _macro_loss_label,
)
from devolucao_evitada_constants import EVITADA_TIPO_LABELS, EVITADA_TIPOS_ORDENADOS, label_tipo_evitada
from devolucoes_consolidado import merge_unique_helper_id_lists
from utils.business_calendar import commercial_competence_period_iso_bounds, competence_date_str

import bi_clientes_intel as bci_clientes

router = APIRouter()
templates = Jinja2Templates(directory="templates")

# Motivos/responsabilidades: tabelas pequenas; cache em memória evita 2× leitura por página BI.
_BI_MOT_RSP_CACHE: dict = {"ts": 0.0, "mot": None, "rsp": None}
_BI_MOT_RSP_TTL_SEC = 300.0


def _json_for_inline_script(obj: Any) -> str:
    """JSON para <script type=\"application/json\">: impede que '</' em texto feche a tag e invalide o parse."""
    try:
        raw = json.dumps(obj, ensure_ascii=False, default=str, allow_nan=False)
    except (ValueError, TypeError, OverflowError):
        raw = json.dumps(obj, ensure_ascii=False, default=str)
    return raw.replace("</", "\\u003c/")


def _competence_period_window(date_i: date, date_f: date, pad_days: int = 10) -> tuple[str, str, str, str]:
    start = date_i.strftime("%Y-%m-%d")
    end = date_f.strftime("%Y-%m-%d")
    window_start = (date_i - timedelta(days=pad_days)).strftime("%Y-%m-%d")
    window_end = (date_f + timedelta(days=pad_days)).strftime("%Y-%m-%d")
    return start, end, window_start, window_end


def _competence_date_or_self(raw_date: Optional[str]) -> str:
    return competence_date_str(raw_date) or str(raw_date or "")[:10]


def _in_competence_period(raw_date: Optional[str], start: str, end: str) -> bool:
    comp = _competence_date_or_self(raw_date)
    return bool(comp) and start <= comp <= end


def _in_operational_date_range_iso(op_raw: Optional[str], start: str, end: str) -> bool:
    """Recorte do BI por calendário da operação (entrega/romaneio), não por competência."""
    s = str(op_raw or "").strip()[:10]
    if len(s) < 10:
        return False
    return start <= s <= end


def _get_cached_mot_rsp_maps(session: Session):
    now = time.monotonic()
    if _BI_MOT_RSP_CACHE["mot"] is not None and (now - _BI_MOT_RSP_CACHE["ts"]) < _BI_MOT_RSP_TTL_SEC:
        return _BI_MOT_RSP_CACHE["mot"], _BI_MOT_RSP_CACHE["rsp"]
    mot_map = {m.id: m for m in session.exec(select(models.DevolucaoMotivo)).all()}
    rsp_map = {r.id: r for r in session.exec(select(models.DevolucaoResponsabilidade)).all()}
    _BI_MOT_RSP_CACHE["mot"] = mot_map
    _BI_MOT_RSP_CACHE["rsp"] = rsp_map
    _BI_MOT_RSP_CACHE["ts"] = now
    return mot_map, rsp_map


def _fmt_br_1(val):
    """Um decimal: 1.234,5"""
    if val is None:
        return "0,0"
    try:
        n = float(val)
        if not math.isfinite(n):
            return "0,0"
        return f"{n:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(val)


def _fmt_br_2(val):
    """Dois decimais: 1.234,56"""
    if val is None:
        return "0,00"
    try:
        n = float(val)
        if not math.isfinite(n):
            return "0,00"
        return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(val)


def _fmt_br_int(val):
    """Inteiro com ponto milhar: 1.234"""
    if val is None:
        return "0"
    try:
        n = float(val)
        if not math.isfinite(n):
            return "0"
        return f"{int(n):,}".replace(",", ".")
    except Exception:
        return str(val)


def _fmt_br_data(s):
    """YYYY-MM-DD -> DD/MM/YYYY"""
    if not s or not str(s).strip():
        return "-"
    try:
        parts = str(s).strip().split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return str(s)


def _fmt_br_date(val):
    """Data em dd/mm/yyyy (str YYYY-MM-DD, date ou datetime) — alinhado ao filtro global em main.py."""
    if val is None:
        return "—"
    try:
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, str) and len(val.strip()) >= 10:
            d = datetime.strptime(val.strip()[:10], "%Y-%m-%d").date()
            return d.strftime("%d/%m/%Y")
        if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day") and not isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
    except Exception:
        pass
    return str(val) if val else "—"


def _fmt_nb_br(nb: Any) -> str:
    """NB numérico com milhar (BR); texto alfanumérico sem alterar."""
    if nb is None:
        return "—"
    raw = str(nb).strip()
    if not raw:
        return "—"
    digits = raw.replace(".", "").replace(",", "")
    if digits.isdigit() and len(digits) <= 12:
        try:
            return _fmt_br_int(int(digits))
        except Exception:
            return raw
    return raw


def _fmt_br_datetime_local(dt: Optional[datetime]) -> str:
    """Data/hora em pt-BR (fuso America/Sao_Paulo)."""
    if not dt:
        return "—"
    try:
        z = ZoneInfo("America/Sao_Paulo")
        if getattr(dt, "tzinfo", None) is None:
            dtl = dt.replace(tzinfo=z)
        else:
            dtl = dt.astimezone(z)
        return dtl.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return "—"


def _fmt_br_moeda(val):
    """R$ 1.234,56"""
    if val is None:
        return "R$ 0,00"
    try:
        return "R$ " + _fmt_br_2(val)
    except Exception:
        return "R$ —"


def _fmt_br_kg(val):
    """Peso em quilogramas (padrão BR: vírgula decimal)."""
    if val is None:
        return "—"
    try:
        return _fmt_br_2(val) + " kg"
    except Exception:
        return "—"


def _fmt_br_duracao(val):
    """Duração em minutos: <60 → 'X min'; ≥60 → 'X h Y min' (padrão BR)."""
    if val is None:
        return "—"
    try:
        m = int(round(float(val)))
        if m < 0:
            m = 0
        if m < 60:
            return f"{m} min"
        h, mn = m // 60, m % 60
        if mn == 0:
            return f"{h} h"
        return f"{h} h {mn} min"
    except Exception:
        return "—"


templates.env.filters["fmt_br_kg"] = _fmt_br_kg

templates.env.filters["fmt_br_1"] = _fmt_br_1
templates.env.filters["fmt_br_2"] = _fmt_br_2
templates.env.filters["fmt_br_int"] = _fmt_br_int
templates.env.filters["fmt_br_data"] = _fmt_br_data
templates.env.filters["fmt_br_date"] = _fmt_br_date
templates.env.filters["fmt_br_moeda"] = _fmt_br_moeda
templates.env.filters["fmt_br_duracao"] = _fmt_br_duracao


def _norm_text(value: Optional[str]) -> str:
    return str(value or "").strip().lower()


def _safe_pct(numerator: float, denominator: float) -> float:
    den = float(denominator or 0.0)
    num = float(numerator or 0.0)
    if not math.isfinite(den) or not math.isfinite(num) or den <= 0:
        return 0.0
    return (num / den) * 100.0


def _bi_client_return_pct_planned(planned_v: float, delivered_v: float, returned_v: float) -> float:
    """% devolvido sobre base comercial (máx. entre planejado, entregue e devolvido).

    Evita percentuais absurdos quando ``planned_value`` ≈ 0 mas há devoluções MANUAIS
    ou rotas sem valor planejado — o antigo ``max(planned, 0.01)`` inflava o índice.
    """
    den = max(float(planned_v or 0.0), float(delivered_v or 0.0), float(returned_v or 0.0), 0.01)
    return round(_safe_pct(float(returned_v or 0.0), den), 2)


# Custo operacional estimado (R$/hora) — parâmetro explicável no tooltip
_BI_EXEC_HOURLY_COST = 75.0

_DURATION_BUCKET_LABELS = ("≤20 min", "21–40 min", "41–60 min", "61–90 min", ">90 min")


def _duration_bucket_idx(minutes: Optional[float]) -> Optional[int]:
    if minutes is None:
        return None
    try:
        m = float(minutes)
    except (TypeError, ValueError):
        return None
    if m <= 20:
        return 0
    if m <= 40:
        return 1
    if m <= 60:
        return 2
    if m <= 90:
        return 3
    return 4


def _heatmap_delay_cause(row: dict) -> str:
    """Causa operacional para heatmap cidade × causa (visitas com atrito)."""
    st = _norm_text(row.get("status"))
    motivo = _norm_text(row.get("motivo") or "")
    if st == "devolucao":
        if any(k in motivo for k in ("fechado", "ausente", "fechado")):
            return "Cliente ausente / ponto fechado"
        if any(k in motivo for k in ("preco", "preço", "pedido", "produto", "comercial", "prazo")):
            return "Erro comercial (refletido na rota)"
        return "Devolução"
    if int(row.get("reopen_count") or 0) > 0:
        return "Reabertura"
    visit = str(row.get("visit_time") or "").strip()
    window = str(row.get("client_window") or "").strip()
    ws, we = _parse_client_window_range(window)
    if visit and ws is not None and we is not None:
        wmatch = _time_is_within_window(visit, ws, we)
        if wmatch is False:
            return "Horário / janela"
    dur = row.get("duration_m")
    if dur is not None and float(dur) > 55:
        if any(k in motivo for k in ("localiz", "nao localizado", "não localizado", "endereco", "endereço")):
            return "Local não localizado"
        if "acesso" in motivo:
            return "Difícil acesso"
    area = _canonical_responsabilidade_for_macro_loss(row.get("responsabilidade"))
    if area is not None:
        return area
    mc = _classify_macro_cause(row.get("motivo"), row.get("responsabilidade"))
    if mc == "Cadastro / planejamento":
        return "Erro cadastro / planejamento"
    if float(dur or 0) > 60:
        return "Demora operacional"
    return "Outros"


def _client_row_key(client_id, client_name: str) -> str:
    if client_id is not None:
        return f"id:{int(client_id)}"
    return f"name:{str(client_name or '').strip().lower()}"


def _exec_accumulate_rows(
    rows: list[dict],
    synth_group_id: Optional[int] = None,
    group_member_ids: Optional[set] = None,
    financial_rows: Optional[list[dict]] = None,
) -> dict:
    """Métricas executivas a partir de linhas ROTA/MANUAL filtradas."""
    unprod_by_key: dict[str, float] = {}
    macro_value_global: dict[str, float] = {}
    macro_time_global: dict[str, float] = {}
    macro_clients: dict[str, set] = {}
    macro_drivers: dict[str, set] = {}
    delivered_total = 0.0
    returned_total = 0.0
    planned_for_rate = 0.0
    manual_returned_sum = 0.0
    visits_rota = 0
    duration_total = 0.0
    unproductive_total = 0.0
    productive_total = 0.0
    bucket_visits = [0, 0, 0, 0, 0]
    bucket_value = [0.0, 0, 0, 0, 0]
    bucket_duration = [0.0, 0, 0, 0, 0]
    bucket_returns = [0, 0, 0, 0, 0]
    bucket_unprod = [0.0, 0, 0, 0, 0]
    city_bucket: dict[str, list] = {}
    city_cause: dict[str, dict[str, float]] = {}
    driver_unprod: dict[str, float] = {}
    cause_unprod_time: dict[str, float] = {}

    def row_key(row):
        cid = row.get("client_id")
        if synth_group_id and group_member_ids and cid is not None and int(cid) in group_member_ids:
            return _client_row_key(-int(synth_group_id), f"Grupo:{synth_group_id}")
        return _client_row_key(cid, str(row.get("client_name") or ""))

    for row in rows:
        src = str(row.get("source") or "ROTA").strip().upper()
        st = _norm_text(row.get("status"))
        cid = row.get("client_id")
        key = row_key(row)
        dur = row.get("duration_m")
        dval = float(dur or 0) if dur is not None else 0.0
        driver_name = str(row.get("driver_name") or "-").strip() or "-"
        city = str(row.get("client_city") or "Sem cidade").strip() or "Sem cidade"
        deliv_val = float(row.get("delivered_value") or 0.0)
        ret_val = float(row.get("returned_value") or 0.0)
        planned_v = float(row.get("planned_value") or 0.0)
        reopen = int(row.get("reopen_count") or 0)

        if src == "ROTA":
            visits_rota += 1
            planned_for_rate += planned_v
            delivered_total += deliv_val if st == "entregue" else (deliv_val if st == "devolucao" else 0.0)
            if financial_rows is None and st == "devolucao":
                returned_total += ret_val if ret_val > 0 else planned_v
            if dur is not None:
                duration_total += dval
                is_unproductive = st == "devolucao" or reopen > 0 or st in ("reaberta", "cancelada")
                if is_unproductive:
                    unproductive_total += dval
                    unprod_by_key[key] = unprod_by_key.get(key, 0.0) + dval
                    driver_unprod[driver_name] = driver_unprod.get(driver_name, 0.0) + dval
                else:
                    productive_total += dval
                bi = _duration_bucket_idx(dur)
                if bi is not None:
                    bucket_visits[bi] += 1
                    bucket_value[bi] += deliv_val if st == "entregue" else (deliv_val if st == "devolucao" else planned_v * 0.01)
                    bucket_duration[bi] += dval
                    if st == "devolucao":
                        bucket_returns[bi] += 1
                        bucket_unprod[bi] += dval
                    elif is_unproductive:
                        bucket_unprod[bi] += dval
                    city_bucket.setdefault(city, [0, 0, 0, 0, 0])
                    city_bucket[city][bi] += 1
                if st == "devolucao" or reopen > 0 or (dur and float(dur) > 60) or (
                    st != "entregue" and st != "pendente"
                ):
                    cause = _heatmap_delay_cause(row)
                    city_cause.setdefault(city, {})
                    city_cause[city][cause] = city_cause[city].get(cause, 0.0) + 1.0
                    if dur and (st == "devolucao" or reopen > 0):
                        cause_unprod_time[cause] = cause_unprod_time.get(cause, 0.0) + dval
        elif financial_rows is None and src == "MANUAL" and st == "devolucao" and ret_val > 0:
            returned_total += ret_val
            manual_returned_sum += ret_val
        if financial_rows is None and st == "devolucao" and ret_val > 0:
            macro = _macro_loss_label(row.get("motivo"), row.get("responsabilidade"))
            macro_value_global[macro] = macro_value_global.get(macro, 0.0) + ret_val
            if dur:
                macro_time_global[macro] = macro_time_global.get(macro, 0.0) + dval
            macro_clients.setdefault(macro, set())
            if cid is not None:
                macro_clients[macro].add(int(cid))
            macro_drivers.setdefault(macro, set())
            macro_drivers[macro].add(driver_name)

    if financial_rows is not None:
        for row in financial_rows:
            ret_val = float(row.get("value") or row.get("returned_value") or 0.0)
            if ret_val <= 0:
                continue
            cid = row.get("client_id")
            driver_name = str(row.get("driver_name") or "-").strip() or "-"
            returned_total += ret_val
            if row.get("standalone"):
                manual_returned_sum += ret_val
            macro = _macro_loss_label(row.get("motivo"), row.get("responsabilidade"))
            macro_value_global[macro] = macro_value_global.get(macro, 0.0) + ret_val
            macro_clients.setdefault(macro, set())
            if cid is not None:
                macro_clients[macro].add(int(cid))
            macro_drivers.setdefault(macro, set())
            macro_drivers[macro].add(driver_name)

    # Mesma regra do BI Entregas: meta financeira sobre (valor entregue + valor devolvido).
    fin_base = max(delivered_total + returned_total, 0.01)
    return {
        "unprod_by_key": unprod_by_key,
        "macro_value_global": macro_value_global,
        "macro_time_global": macro_time_global,
        "macro_clients": macro_clients,
        "macro_drivers": macro_drivers,
        "delivered_total": round(delivered_total, 2),
        "returned_total": round(returned_total, 2),
        "financial_base": round(max(fin_base, 0.01), 2),
        "visits_rota": visits_rota,
        "duration_total": round(duration_total, 1),
        "unproductive_total": round(unproductive_total, 1),
        "productive_total": round(productive_total, 1),
        "planned_total": round(planned_for_rate, 2),
        "bucket_visits": bucket_visits,
        "bucket_value": [round(x, 2) for x in bucket_value],
        "bucket_duration": [round(x, 1) for x in bucket_duration],
        "bucket_returns": bucket_returns,
        "bucket_unprod": [round(x, 1) for x in bucket_unprod],
        "city_bucket": city_bucket,
        "city_cause": city_cause,
        "driver_unprod": driver_unprod,
        "cause_unprod_time": cause_unprod_time,
    }


def _week_bucket(date_str: Optional[str]) -> str:
    try:
        ref = datetime.strptime(str(date_str or ""), "%Y-%m-%d").date()
        iso = ref.isocalendar()
        return f"{iso.year}-S{iso.week:02d}"
    except Exception:
        return "Sem semana"


def _parse_hhmm_minutes(value: Optional[str]) -> Optional[int]:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        hh, mm = raw[:5].split(":")
        return int(hh) * 60 + int(mm)
    except Exception:
        return None


def _parse_client_window_range(window: Optional[str]) -> tuple[Optional[int], Optional[int]]:
    raw = str(window or "").strip()
    if not raw or raw.lower().startswith("sem janela"):
        return (None, None)
    parts = raw.split(" - ", 1) if " - " in raw else raw.split("-", 1)
    if len(parts) != 2:
        return (None, None)
    return (_parse_hhmm_minutes(parts[0]), _parse_hhmm_minutes(parts[1]))


def _time_is_within_window(value: Optional[str], start_m: Optional[int], end_m: Optional[int]) -> Optional[bool]:
    visit_m = _parse_hhmm_minutes(value)
    if visit_m is None or start_m is None or end_m is None:
        return None
    if end_m >= start_m:
        return start_m <= visit_m <= end_m
    return visit_m >= start_m or visit_m <= end_m


def _aggregate_bi_client_rows(rows: list[dict], financial_rows: Optional[list[dict]] = None) -> dict[str, dict]:
    client_agg: dict[str, dict] = {}

    for row in rows:
        row_client_id = row.get("client_id")
        row_client_name = str(row.get("client_name") or "").strip() or "Sem Cliente"
        key = f"id:{row_client_id}" if row_client_id is not None else f"name:{row_client_name.lower()}"
        source = str(row.get("source") or "ROTA").strip().upper()
        status_row = _norm_text(row.get("status"))
        duration_m = row.get("duration_m")
        duration_value = float(duration_m or 0.0) if duration_m is not None else 0.0
        planned_value = float(row.get("planned_value") or 0.0)
        delivered_value = float(row.get("delivered_value") or 0.0)
        returned_value = float(row.get("returned_value") or 0.0)
        returned_kg = float(row.get("returned_kg") or 0.0)
        reopen_count = int(row.get("reopen_count") or 0)
        driver_name = str(row.get("driver_name") or "-").strip() or "-"
        motivo = str(row.get("motivo") or "Não informado").strip() or "Não informado"
        responsabilidade = str(row.get("responsabilidade") or "Não informado").strip() or "Não informado"
        week_key = _week_bucket(row.get("date"))
        window_label = str(row.get("client_window") or "").strip() or "Sem janela"
        visit_time = str(row.get("visit_time") or "").strip()

        bucket = client_agg.setdefault(
            key,
            {
                "client_id": row_client_id,
                "client_key": key,
                "client_name": row_client_name,
                "city": str(row.get("client_city") or "").strip() or "Sem cidade",
                "bairro": str(row.get("client_bairro") or "").strip() or "Sem bairro",
                "segmento": str(row.get("client_segmento") or "").strip() or "Sem segmento",
                "prioridade": str(row.get("client_prioridade") or "").strip().upper() or "Sem prioridade",
                "status_operacional": str(row.get("client_status_operacional") or "").strip() or "Sem status",
                "status_cadastro": str(row.get("client_status_cadastro") or "").strip() or "Sem cadastro",
                "address": str(row.get("client_address") or "").strip() or "Endereço não informado",
                "window": window_label,
                "visits": 0,
                "delivered_visits": 0,
                "open_visits": 0,
                "returned_occurrences": 0,
                "planned_value": 0.0,
                "delivered_value": 0.0,
                "returned_value": 0.0,
                "returned_kg": 0.0,
                "planned_kg": 0.0,
                "delivered_kg": 0.0,
                "manual_returned_value": 0.0,
                "total_duration_m": 0.0,
                "duration_count": 0,
                "max_duration_m": 0.0,
                "min_duration_m": None,
                "reopen_count": 0,
                "window_checks": 0,
                "window_hits": 0,
                "window_misses": 0,
                "weeks": {},
                "drivers": {},
                "motivos": {},
                "responsabilidades": {},
                "latest_date": str(row.get("date") or ""),
            },
        )

        if str(row.get("date") or "") > bucket["latest_date"]:
            bucket["latest_date"] = str(row.get("date") or "")

        if source == "ROTA":
            bucket["visits"] += 1
            bucket["planned_value"] += planned_value
            bucket["delivered_value"] += delivered_value
            bucket["planned_kg"] += float(row.get("planned_kg") or 0.0)
            bucket["delivered_kg"] += float(row.get("delivered_kg") or 0.0)
            bucket["weeks"][week_key] = bucket["weeks"].get(week_key, 0) + 1

        driver_bucket = bucket["drivers"].setdefault(
            driver_name,
            {"visits": 0, "duration_m": 0.0, "returned_value": 0.0, "returns": 0},
        )
        if source == "ROTA":
            driver_bucket["visits"] += 1
        if duration_m is not None:
            bucket["total_duration_m"] += duration_value
            bucket["duration_count"] += 1
            driver_bucket["duration_m"] += duration_value
            if duration_value > float(bucket.get("max_duration_m") or 0.0):
                bucket["max_duration_m"] = duration_value
            prev_min = bucket.get("min_duration_m")
            if prev_min is None or duration_value < float(prev_min):
                bucket["min_duration_m"] = duration_value
        if reopen_count:
            bucket["reopen_count"] += reopen_count

        if source == "ROTA" and visit_time:
            window_start_m, window_end_m = _parse_client_window_range(bucket["window"])
            window_match = _time_is_within_window(visit_time, window_start_m, window_end_m)
            if window_match is not None:
                bucket["window_checks"] += 1
                if window_match:
                    bucket["window_hits"] += 1
                else:
                    bucket["window_misses"] += 1

        if status_row == "entregue":
            bucket["delivered_visits"] += 1
        elif status_row in {"pendente", "iniciada", "reaberta", "cancelada"}:
            bucket["open_visits"] += 1

        if financial_rows is None and status_row == "devolucao":
            bucket["returned_occurrences"] += 1
            bucket["returned_value"] += returned_value
            bucket["returned_kg"] += returned_kg
            driver_bucket["returned_value"] += returned_value
            driver_bucket["returns"] += 1
            motivo_bucket = bucket["motivos"].setdefault(motivo, {"count": 0, "value": 0.0})
            motivo_bucket["count"] += 1
            motivo_bucket["value"] += returned_value
            resp_bucket = bucket["responsabilidades"].setdefault(responsabilidade, {"count": 0, "value": 0.0})
            resp_bucket["count"] += 1
            resp_bucket["value"] += returned_value

    if financial_rows is not None:
        for row in financial_rows:
            row_client_id = row.get("client_id")
            row_client_name = str(row.get("client_name") or "").strip() or "Sem Cliente"
            key = f"id:{row_client_id}" if row_client_id is not None else f"name:{row_client_name.lower()}"
            driver_name = str(row.get("driver_name") or "-").strip() or "-"
            motivo = str(row.get("motivo") or "Não informado").strip() or "Não informado"
            responsabilidade = str(row.get("responsabilidade") or "Não informado").strip() or "Não informado"
            returned_value = float(row.get("value") or row.get("returned_value") or 0.0)
            bucket = client_agg.setdefault(
                key,
                {
                    "client_id": row_client_id,
                    "client_key": key,
                    "client_name": row_client_name,
                    "city": str(row.get("client_city") or "").strip() or "Sem cidade",
                    "bairro": str(row.get("client_bairro") or "").strip() or "Sem bairro",
                    "segmento": str(row.get("client_segmento") or "").strip() or "Sem segmento",
                    "prioridade": str(row.get("client_prioridade") or "").strip().upper() or "Sem prioridade",
                    "status_operacional": str(row.get("client_status_operacional") or "").strip() or "Sem status",
                    "status_cadastro": str(row.get("client_status_cadastro") or "").strip() or "Sem cadastro",
                    "address": str(row.get("client_address") or "").strip() or "Endereço não informado",
                    "window": str(row.get("client_window") or "").strip() or "Sem janela",
                    "visits": 0,
                    "delivered_visits": 0,
                    "open_visits": 0,
                    "returned_occurrences": 0,
                    "planned_value": 0.0,
                    "delivered_value": 0.0,
                    "returned_value": 0.0,
                    "returned_kg": 0.0,
                    "planned_kg": 0.0,
                    "delivered_kg": 0.0,
                    "manual_returned_value": 0.0,
                    "total_duration_m": 0.0,
                    "duration_count": 0,
                    "max_duration_m": 0.0,
                    "min_duration_m": None,
                    "reopen_count": 0,
                    "window_checks": 0,
                    "window_hits": 0,
                    "window_misses": 0,
                    "weeks": {},
                    "drivers": {},
                    "motivos": {},
                    "responsabilidades": {},
                    "latest_date": str(row.get("date") or ""),
                },
            )
            if str(row.get("date") or "") > bucket["latest_date"]:
                bucket["latest_date"] = str(row.get("date") or "")
            if row.get("standalone"):
                bucket["manual_returned_value"] += returned_value
            bucket["returned_occurrences"] += 1
            bucket["returned_value"] += returned_value
            driver_bucket = bucket["drivers"].setdefault(
                driver_name,
                {"visits": 0, "duration_m": 0.0, "returned_value": 0.0, "returns": 0},
            )
            driver_bucket["returned_value"] += returned_value
            driver_bucket["returns"] += 1
            motivo_bucket = bucket["motivos"].setdefault(motivo, {"count": 0, "value": 0.0})
            motivo_bucket["count"] += 1
            motivo_bucket["value"] += returned_value
            resp_bucket = bucket["responsabilidades"].setdefault(responsabilidade, {"count": 0, "value": 0.0})
            resp_bucket["count"] += 1
            resp_bucket["value"] += returned_value

    return client_agg


def _ma(vals: list[float], w: int) -> list[Optional[float]]:
    out: list[Optional[float]] = []
    for i in range(len(vals)):
        chunk = vals[max(0, i - w + 1): i + 1]
        out.append(round(sum(chunk) / len(chunk), 2) if chunk else None)
    return out


def _get_fleet_plates(session: Session) -> set[str]:
    """Retorna placas ativas da frota (Vehicle). Em caso de erro, retorna set vazio."""
    try:
        rows = session.exec(
            select(models.Vehicle.placa).where(models.Vehicle.is_active == True)
        ).all()
        return {str(p).strip().upper() for p in (rows or []) if p and str(p).strip()}
    except Exception:
        return set()


def _load_financial_devolucao_rows(
    session: Session,
    date_i: date,
    date_f: date,
    shift: str,
    driver_id: Optional[int],
    plate: str,
    status: str,
) -> tuple[list[dict], set[int]]:
    """Base financeira do BI.

    Regra de fechamento:
    - até a última data coberta pelo Excel no recorte, Excel é a fonte oficial;
    - depois disso, usa a base consolidada do sistema sem duplicatas.

    Retorna também os route_id já cobertos por linhas Devolucao em ``selected`` (evita duplicar valor com a rota).
    """
    st = (status or "Todos").strip().lower()
    if st not in ("todos", "devolucao"):
        return [], set()

    pl = (plate or "Todos").strip().upper()
    lookback_start = date_i - timedelta(days=7)
    q = (
        select(models.Devolucao)
        .where(models.Devolucao.data_romaneio >= lookback_start.strftime("%Y-%m-%d"))
        .where(models.Devolucao.data_romaneio <= date_f.strftime("%Y-%m-%d"))
    )
    if driver_id:
        q = q.where(models.Devolucao.motorista_id == driver_id)
    devolucoes = session.exec(q.order_by(models.Devolucao.data_romaneio, models.Devolucao.created_at)).all()
    if not devolucoes:
        return [], set()

    route_ids = sorted({int(d.route_id) for d in devolucoes if d.route_id is not None})
    route_map = {
        r.id: r
        for r in (
            session.exec(select(models.Route).where(models.Route.id.in_(route_ids))).all()
            if route_ids else []
        )
    }

    filtered: list[models.Devolucao] = []
    for d in devolucoes:
        comp_date = competence_date_str(str(d.data_entrega or d.data_romaneio or ""))
        if comp_date:
            if not (date_i.strftime("%Y-%m-%d") <= comp_date <= date_f.strftime("%Y-%m-%d")):
                continue
        route = route_map.get(d.route_id) if d.route_id is not None else None
        if shift and shift != "Todos" and route is not None and (route.shift or "").strip() != shift:
            continue
        if pl and pl != "TODOS":
            route_plate = ((route.delivery_vehicle_plate or "").strip().upper() if route else "")
            if not route_plate or route_plate != pl:
                continue
        filtered.append(d)
    if not filtered:
        return [], set()

    def _effective_date(dev: models.Devolucao) -> str:
        return competence_date_str(str(dev.data_entrega or dev.data_romaneio or "").strip()) or str(dev.data_romaneio or "").strip()

    excel_cutoff = max(
        (_effective_date(d) for d in filtered if str(d.source or "").strip().upper() == "EXCEL" and _effective_date(d)),
        default=None,
    )
    if excel_cutoff:
        selected = []
        for d in filtered:
            eff_date = _effective_date(d)
            source = str(d.source or "").strip().upper()
            if eff_date and eff_date <= excel_cutoff:
                if source == "EXCEL":
                    selected.append(d)
            elif d.duplicate_of_id is None:
                selected.append(d)
    else:
        selected = [d for d in filtered if d.duplicate_of_id is None]
    if not selected:
        return [], set()

    emp_ids = sorted({int(d.motorista_id) for d in selected if d.motorista_id})
    cli_ids = sorted({int(d.client_id) for d in selected if d.client_id})
    mot_ids = sorted({int(d.motivo_id) for d in selected if d.motivo_id})
    rsp_ids = sorted({int(d.responsabilidade_id) for d in selected if d.responsabilidade_id})

    emp_map = {
        e.id: e
        for e in (
            session.exec(select(models.Employee).where(models.Employee.id.in_(emp_ids))).all()
            if emp_ids else []
        )
    }
    cli_map = {
        c.id: c
        for c in (
            session.exec(select(models.Client).where(models.Client.id.in_(cli_ids))).all()
            if cli_ids else []
        )
    }
    mot_map = {
        m.id: m
        for m in (
            session.exec(select(models.DevolucaoMotivo).where(models.DevolucaoMotivo.id.in_(mot_ids))).all()
            if mot_ids else []
        )
    }
    rsp_map = {
        r.id: r
        for r in (
            session.exec(select(models.DevolucaoResponsabilidade).where(models.DevolucaoResponsabilidade.id.in_(rsp_ids))).all()
            if rsp_ids else []
        )
    }

    rows: list[dict] = []
    for d in selected:
        route = route_map.get(d.route_id) if d.route_id is not None else None
        cli = cli_map.get(d.client_id)
        emp = emp_map.get(d.motorista_id)
        val = round(float(d.valor or 0.0), 2)
        client_window = ""
        if cli and getattr(cli, "janela_horario_inicio", None) and getattr(cli, "janela_horario_fim", None):
            client_window = f"{cli.janela_horario_inicio} - {cli.janela_horario_fim}"
        rows.append(
            {
                "devolucao_id": d.id,
                "route_id": d.route_id,
                "date": d.data_romaneio,
                "delivery_date": d.data_entrega,
                "driver_id": d.motorista_id,
                "driver_name": (emp.name if emp else f"Motorista #{d.motorista_id}"),
                "client_id": d.client_id,
                "client_name": (cli.name if cli else f"Cliente #{d.client_id}"),
                "client_city": (getattr(cli, "municipio", None) or "").strip() if cli else "",
                "client_bairro": (getattr(cli, "bairro", None) or "").strip() if cli else "",
                "client_segmento": (getattr(cli, "segmento", None) or "").strip() if cli else "",
                "client_prioridade": (getattr(cli, "prioridade_logistica", None) or "").strip() if cli else "",
                "client_status_operacional": (getattr(cli, "status_operacional", None) or "").strip() if cli else "",
                "client_status_cadastro": (getattr(cli, "status_cliente", None) or "").strip() if cli else "",
                "client_address": (cli.get_full_address() or cli.endereco or "").strip() if cli else "",
                "client_window": client_window,
                "shift": (route.shift or "-") if route else "-",
                "plate": ((route.delivery_vehicle_plate or "-").strip().upper() if route else "-"),
                "value": val,
                "returned_value": val,
                "motivo": (mot_map.get(d.motivo_id).nome if mot_map.get(d.motivo_id) else "Nao informado"),
                "responsabilidade": (rsp_map.get(d.responsabilidade_id).nome if rsp_map.get(d.responsabilidade_id) else "Nao informado"),
                "cluster": d.cluster or "Sem Cluster",
                "source": (d.source or "MANUAL").strip().upper(),
                "standalone": d.route_id is None,
                "acima_300": "SIM" if (str(d.acima_300 or "").upper() == "SIM" or val >= 300) else "NAO",
            }
        )
    covered_route_ids = {int(d.route_id) for d in selected if d.route_id is not None}
    return rows, covered_route_ids


def _financial_gap_rows_from_routes(
    routes: list,
    covered_route_ids: set[int],
    emp_map: dict,
    cli_map: dict,
) -> list[dict]:
    """Rotas com devolução operacional sem registro Devolucao na consolidação financeira (completa valor e gráficos)."""
    gap: list[dict] = []
    for r in routes:
        status_raw = (r.delivery_status or "pendente").strip().lower()
        if status_raw == "devolucao" and is_encerramento_tardio_automatico_return(getattr(r, "delivery_return_reason", None)):
            status_raw = "entregue"
        if status_raw != "devolucao":
            continue
        rid = int(r.id) if r.id is not None else None
        if rid is not None and rid in covered_route_ids:
            continue
        emp = emp_map.get(r.employee_id)
        cli = cli_map.get(r.client_id)
        driver_name = emp.name if emp else f"Motorista #{r.employee_id}"
        client_name = cli.name if cli else f"Cliente #{r.client_id}"
        planned_v = float(r.valor_financeiro or 0.0)
        ret_v = float(r.valor_devolucao if r.valor_devolucao is not None else (planned_v if status_raw == "devolucao" else 0.0))
        val = round(ret_v, 2)
        motivo = (r.delivery_return_reason or "Nao informado").strip() or "Nao informado"
        resp = (r.delivery_return_category or "Nao informado").strip() or "Nao informado"
        client_window = ""
        if cli and getattr(cli, "janela_horario_inicio", None) and getattr(cli, "janela_horario_fim", None):
            client_window = f"{cli.janela_horario_inicio} - {cli.janela_horario_fim}"
        gap.append(
            {
                "devolucao_id": None,
                "route_id": r.id,
                "date": r.date,
                "delivery_date": r.date,
                "driver_id": r.employee_id,
                "driver_name": driver_name,
                "client_id": r.client_id,
                "client_name": client_name,
                "client_city": (getattr(cli, "municipio", None) or "").strip() if cli else "",
                "client_bairro": (getattr(cli, "bairro", None) or "").strip() if cli else "",
                "client_segmento": (getattr(cli, "segmento", None) or "").strip() if cli else "",
                "client_prioridade": (getattr(cli, "prioridade_logistica", None) or "").strip() if cli else "",
                "client_status_operacional": (getattr(cli, "status_operacional", None) or "").strip() if cli else "",
                "client_status_cadastro": (getattr(cli, "status_cliente", None) or "").strip() if cli else "",
                "client_address": (cli.get_full_address() or cli.endereco or "").strip() if cli else "",
                "client_window": client_window,
                "shift": (r.shift or "-").strip() if r.shift else "-",
                "plate": ((r.delivery_vehicle_plate or "-").strip().upper() if r.delivery_vehicle_plate else "-"),
                "value": val,
                "returned_value": val,
                "motivo": motivo,
                "responsabilidade": resp,
                "cluster": "Sem Cluster",
                "source": "ROTA",
                "standalone": False,
                "acima_300": "SIM" if val >= 300 else "NAO",
            }
        )
    return gap


def _build_bi_delivery_dataset(
    session: Session,
    date_from: Optional[str],
    date_to: Optional[str],
    shift: str,
    driver_id: Optional[int],
    plate: str,
    status: str,
    detail_driver_id: Optional[int] = None,
    detail_status: str = "Todos",
) -> dict:
    tz = ZoneInfo("America/Sao_Paulo")
    today = datetime.now(tz).date()

    def _d(raw: Optional[str]) -> Optional[date]:
        if not raw:
            return None
        try:
            return datetime.strptime(raw, "%Y-%m-%d").date()
        except Exception:
            return None

    date_i = _d(date_from) or today.replace(day=1)
    date_f = _d(date_to) or today
    if date_i > date_f:
        date_i, date_f = date_f, date_i
    period_start, period_end, window_start, window_end = _competence_period_window(date_i, date_f)

    st = (status or "Todos").strip().lower()
    pl = (plate or "Todos").strip().upper()
    detail_status_norm = (detail_status or "Todos").strip().lower()

    q = (
        select(models.Route)
        .where(models.Route.type == "delivery")
        .where(models.Route.date >= window_start)
        .where(models.Route.date <= window_end)
    )
    if shift and shift != "Todos":
        q = q.where(models.Route.shift == shift)
    if driver_id:
        q = q.where(models.Route.employee_id == driver_id)
    if pl and pl != "TODOS":
        q = q.where(models.Route.delivery_vehicle_plate == pl)
    if st and st != "todos":
        q = q.where(func.lower(models.Route.delivery_status) == st)
    routes = session.exec(q.order_by(models.Route.date, models.Route.created_at)).all()
    routes = [r for r in routes if _in_competence_period(getattr(r, "date", None), period_start, period_end)]

    # % devolução operacional oficial (rotas concluídas): ignora filtro de status — mesmo critério Central / informativo
    q_rotas_kpi = (
        select(models.Route)
        .where(models.Route.type == "delivery")
        .where(models.Route.date >= window_start)
        .where(models.Route.date <= window_end)
    )
    if shift and shift != "Todos":
        q_rotas_kpi = q_rotas_kpi.where(models.Route.shift == shift)
    if driver_id:
        q_rotas_kpi = q_rotas_kpi.where(models.Route.employee_id == driver_id)
    if pl and pl != "TODOS":
        q_rotas_kpi = q_rotas_kpi.where(models.Route.delivery_vehicle_plate == pl)
    routes_for_kpi_rotas = session.exec(q_rotas_kpi.order_by(models.Route.date, models.Route.created_at)).all()
    routes_for_kpi_rotas = [
        r for r in routes_for_kpi_rotas
        if _in_competence_period(getattr(r, "date", None), period_start, period_end)
    ]
    rotas_devolucao_count, rotas_concluidas_count = counts_devolucao_rotas_concluidas(routes_for_kpi_rotas)
    return_rate_rotas = pct_devolucao_sobre_rotas_concluidas(routes_for_kpi_rotas)

    manual = []
    if (st in ("todos", "devolucao")) and pl == "TODOS":
        qm = (
            select(models.Devolucao)
            .where(models.Devolucao.data_romaneio >= window_start)
            .where(models.Devolucao.data_romaneio <= window_end)
            .where(models.Devolucao.route_id.is_(None))
        )
        if driver_id:
            qm = qm.where(models.Devolucao.motorista_id == driver_id)
        manual = session.exec(qm.order_by(models.Devolucao.data_romaneio, models.Devolucao.created_at)).all()
        manual = [
            d for d in manual
            if _in_competence_period(
                getattr(d, "data_entrega", None) or getattr(d, "data_romaneio", None),
                period_start,
                period_end,
            )
        ]

    emp_ids = sorted({r.employee_id for r in routes if r.employee_id} | {d.motorista_id for d in manual if d.motorista_id})
    cli_ids = sorted({r.client_id for r in routes if r.client_id} | {d.client_id for d in manual if d.client_id})
    emp_map = {e.id: e for e in (session.exec(select(models.Employee).where(models.Employee.id.in_(emp_ids))).all() if emp_ids else [])}
    cli_map = {c.id: c for c in (session.exec(select(models.Client).where(models.Client.id.in_(cli_ids))).all() if cli_ids else [])}
    mot_map, rsp_map = _get_cached_mot_rsp_maps(session)

    def _hm(v: Optional[str]) -> Optional[int]:
        if not v:
            return None
        try:
            h, m = str(v).split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return None

    def _dur(a: Optional[str], b: Optional[str]) -> Optional[int]:
        sa, sb = _hm(a), _hm(b)
        if sa is None or sb is None:
            return None
        if sb < sa:
            sb += 24 * 60
        return max(0, sb - sa)

    def _fallback_route_time(value: Optional[str]) -> Optional[str]:
        raw = str(value or "").strip()
        if not raw or raw in {"00:00", "0:00"}:
            return None
        return raw

    def _pct(numerator: float, denominator: float) -> float:
        """Percentual seguro sem divisor artificial."""
        den = float(denominator or 0.0)
        if den <= 0:
            return 0.0
        return (float(numerator or 0.0) / den) * 100.0

    planned_stops = len(routes)
    planned_kg = planned_value = 0.0
    realized_stops = started_stops = 0
    realized_kg = realized_value = 0.0
    returned_stops = 0
    returned_kg = returned_value = 0.0
    returned_value_manual = 0.0
    reopen_routes = 0
    dur_list: list[int] = []

    per_day: dict[str, dict] = {}
    per_driver: dict[str, dict] = {}
    motivo_agg: dict[str, dict] = {}
    resp_agg: dict[str, dict] = {}
    cluster_agg: dict[str, dict] = {}
    ret_count_day: dict[str, int] = {}
    ret_value_day: dict[str, float] = {}
    client_returns: dict[str, dict] = {}
    reopen_heat: dict[str, dict[str, int]] = {}
    route_rows: list[dict] = []
    ex_rows: list[dict] = []

    def _acc_devol(date_key: str, driver_name: str, client_name: str, motivo: str, resp: str, cluster: str, val: float):
        nonlocal returned_stops, returned_value
        returned_stops += 1
        returned_value += val
        ret_count_day[date_key] = ret_count_day.get(date_key, 0) + 1
        ret_value_day[date_key] = round(ret_value_day.get(date_key, 0.0) + val, 2)
        motivo_agg.setdefault(motivo, {"motivo": motivo, "qtd": 0, "valor": 0.0})
        motivo_agg[motivo]["qtd"] += 1
        motivo_agg[motivo]["valor"] += val
        resp_agg.setdefault(resp, {"responsabilidade": resp, "qtd": 0, "valor": 0.0})
        resp_agg[resp]["qtd"] += 1
        resp_agg[resp]["valor"] += val
        cluster_agg.setdefault(cluster, {"cluster": cluster, "qtd": 0, "valor": 0.0})
        cluster_agg[cluster]["qtd"] += 1
        cluster_agg[cluster]["valor"] += val
        cr = client_returns.setdefault(client_name, {"qtd": 0, "valor": 0.0})
        cr["qtd"] += 1
        cr["valor"] += val

    for r in routes:
        comp_date = _competence_date_or_self(getattr(r, "date", None)) or str(getattr(r, "date", "") or "")
        status_raw = (r.delivery_status or "pendente").strip().lower()
        # Legacy: encerramento automático foi marcado como devolução por engano → tratar como entregue
        if status_raw == "devolucao" and is_encerramento_tardio_automatico_return(getattr(r, "delivery_return_reason", None)):
            status_raw = "entregue"
        emp = emp_map.get(r.employee_id)
        cli = cli_map.get(r.client_id)
        driver = emp.name if emp else f"Motorista #{r.employee_id}"
        client = cli.name if cli else f"Cliente #{r.client_id}"
        planned_w = float(r.tonnage or 0.0)
        planned_v = float(r.valor_financeiro or 0.0)
        ret_w = float(r.devolucao_volume if r.devolucao_volume is not None else (planned_w if status_raw == "devolucao" else 0.0))
        ret_v = float(r.valor_devolucao if r.valor_devolucao is not None else (planned_v if status_raw == "devolucao" else 0.0))
        del_w = max(0.0, planned_w - ret_w) if status_raw == "devolucao" else (planned_w if status_raw == "entregue" else 0.0)
        del_v = max(0.0, planned_v - ret_v) if status_raw == "devolucao" else (planned_v if status_raw == "entregue" else 0.0)
        # Só contabiliza tempo com registro mobile (GPS início + fim); web não mede duração.
        dur = route_duration_minutes_mobile_only(r)
        start_ref = (r.delivery_started_at or "").strip() or _fallback_route_time(r.start_time)
        end_ref = (r.delivery_finished_at or "").strip() or _fallback_route_time(r.end_time)

        planned_kg += planned_w
        planned_value += planned_v
        if status_raw in ("iniciada", "devolucao", "entregue"):
            started_stops += 1
        if status_raw in ("devolucao", "entregue"):
            realized_stops += 1
            realized_kg += del_w
            realized_value += del_v
            if dur is not None:
                dur_list.append(dur)
        if status_raw == "devolucao":
            returned_kg += ret_w
            _acc_devol(comp_date, driver, client, (r.delivery_return_reason or "Nao informado"), (r.delivery_return_category or "Nao informado"), "Sem Cluster", ret_v)
        if (r.delivery_reopen_count or 0) > 0:
            reopen_routes += 1
            reopen_heat.setdefault(comp_date, {})
            reopen_heat[comp_date][driver] = reopen_heat[comp_date].get(driver, 0) + int(r.delivery_reopen_count or 0)

        per_day.setdefault(comp_date, {"date": comp_date, "planned_stops": 0, "started_stops": 0, "realized_stops": 0, "returned_stops": 0, "planned_kg": 0.0, "realized_kg": 0.0, "returned_kg": 0.0, "planned_value": 0.0, "returned_value": 0.0})
        d = per_day[comp_date]
        d["planned_stops"] += 1
        d["planned_kg"] += planned_w
        d["planned_value"] += planned_v
        if status_raw in ("iniciada", "devolucao", "entregue"):
            d["started_stops"] += 1
        if status_raw in ("devolucao", "entregue"):
            d["realized_stops"] += 1
            d["realized_kg"] = round(float(d.get("realized_kg") or 0.0) + del_w, 4)
        if status_raw == "devolucao":
            d["returned_stops"] += 1
            d["returned_kg"] += ret_w
            d["returned_value"] += ret_v

        per_driver.setdefault(driver, {"driver_name": driver, "driver_id": r.employee_id, "planned_stops": 0, "realized_stops": 0, "started_stops": 0, "returned_stops": 0, "planned_kg": 0.0, "realized_kg": 0.0, "returned_kg": 0.0, "planned_value": 0.0, "realized_value": 0.0, "returned_value": 0.0, "manual_returned_value": 0.0, "reopen_count": 0, "durations": [], "main_plate": (r.delivery_vehicle_plate or "-").upper()})
        b = per_driver[driver]
        b["planned_stops"] += 1
        b["planned_kg"] += planned_w
        b["planned_value"] += planned_v
        b["reopen_count"] += (r.delivery_reopen_count or 0)
        if status_raw in ("iniciada", "devolucao", "entregue"):
            b["started_stops"] += 1
        if status_raw in ("devolucao", "entregue"):
            b["realized_stops"] += 1
            b["realized_kg"] += del_w
            b["realized_value"] += del_v
        if status_raw == "devolucao":
            b["returned_stops"] += 1
            b["returned_kg"] += ret_w
            b["returned_value"] += ret_v
        if dur is not None:
            b["durations"].append(dur)

        client_city = (getattr(cli, "municipio", None) or r.delivery_city or "").strip() if cli or r.delivery_city else ""
        client_bairro = (getattr(cli, "bairro", None) or r.delivery_neighborhood or "").strip() if cli or r.delivery_neighborhood else ""
        client_segmento = (getattr(cli, "segmento", None) or "").strip() if cli else ""
        client_prioridade = (getattr(cli, "prioridade_logistica", None) or "").strip() if cli else ""
        client_status_operacional = (getattr(cli, "status_operacional", None) or "").strip() if cli else ""
        client_status_cadastro = (getattr(cli, "status_cliente", None) or "").strip() if cli else ""
        client_address = ""
        if cli:
            client_address = (cli.get_full_address() or cli.endereco or "").strip()
        if not client_address:
            client_address = (r.delivery_address or "").strip()
        client_window = ""
        if cli and getattr(cli, "janela_horario_inicio", None) and getattr(cli, "janela_horario_fim", None):
            client_window = f"{cli.janela_horario_inicio} - {cli.janela_horario_fim}"
        row = {"route_id": r.id, "date": comp_date, "shift": r.shift, "driver_id": r.employee_id, "driver_name": driver, "client_id": r.client_id, "client_name": client, "client_city": client_city, "client_bairro": client_bairro, "client_segmento": client_segmento, "client_prioridade": client_prioridade, "client_status_operacional": client_status_operacional, "client_status_cadastro": client_status_cadastro, "client_address": client_address, "client_window": client_window, "visit_time": (start_ref or ""), "status": status_raw, "planned_kg": round(planned_w, 2), "planned_value": round(planned_v, 2), "delivered_kg": round(del_w, 2), "delivered_value": round(del_v, 2), "returned_kg": round(ret_w if status_raw == "devolucao" else 0.0, 2), "returned_value": round(ret_v if status_raw == "devolucao" else 0.0, 2), "reopen_count": r.delivery_reopen_count or 0, "duration_m": dur, "plate": (r.delivery_vehicle_plate or "-").upper(), "order_number": r.delivery_order_number or "-", "motivo": r.delivery_return_reason or "-", "responsabilidade": r.delivery_return_category or "-", "cluster": "Sem Cluster", "acima_300": ("SIM" if ret_v >= 300 and status_raw == "devolucao" else "NAO"), "source": "ROTA", "possible_duplicate": False}
        route_rows.append(row)
        score = (55 if status_raw == "devolucao" else 20 if status_raw == "iniciada" else 25 if status_raw in ("pendente", "reaberta") else 0) + min(20, (r.delivery_reopen_count or 0) * 6) + (10 if (dur or 0) > 120 else 0) + (8 if planned_w >= 500 else 0)
        if score >= 30:
            ex_rows.append({"score": score, "date": comp_date, "shift": r.shift, "driver_name": driver, "driver_id": r.employee_id, "client_name": client, "status": status_raw, "planned_kg": round(planned_w, 2), "planned_value": round(planned_v, 2), "returned_kg": round(ret_w if status_raw == "devolucao" else 0.0, 2), "returned_value": round(ret_v if status_raw == "devolucao" else 0.0, 2), "reopen_count": r.delivery_reopen_count or 0, "duration_m": dur, "source": "ROTA"})

    rota_devol_keys_for_dup: set[tuple] = {
        (r["date"], r["client_id"], r["driver_id"], round(float(r.get("returned_value") or 0), 2))
        for r in route_rows
        if (r.get("status") or "").lower() == "devolucao"
    }
    for d in manual:
        # Ignora Devolucao vinculada a rota: já está representada na linha ROTA (não é manual)
        if d.route_id is not None:
            continue
        driver = (emp_map.get(d.motorista_id).name if emp_map.get(d.motorista_id) else f"Motorista #{d.motorista_id}")
        client = (cli_map.get(d.client_id).name if cli_map.get(d.client_id) else f"Cliente #{d.client_id}")
        motivo = (mot_map.get(d.motivo_id).nome if mot_map.get(d.motivo_id) else "Nao informado")
        resp = (rsp_map.get(d.responsabilidade_id).nome if rsp_map.get(d.responsabilidade_id) else "Nao informado")
        cluster = d.cluster or "Sem Cluster"
        ret_v = float(d.valor or 0.0)
        above = "SIM" if (d.acima_300 or "").upper() == "SIM" or ret_v >= 300 else "NAO"
        dup_key = (d.data_romaneio, d.client_id, d.motorista_id, round(ret_v, 2))
        is_possible_dup = dup_key in rota_devol_keys_for_dup
        if not is_possible_dup:
            _acc_devol(d.data_romaneio, driver, client, motivo, resp, cluster, ret_v)
            returned_value_manual += ret_v
            per_driver.setdefault(driver, {"driver_name": driver, "driver_id": d.motorista_id, "planned_stops": 0, "realized_stops": 0, "started_stops": 0, "returned_stops": 0, "planned_kg": 0.0, "realized_kg": 0.0, "returned_kg": 0.0, "planned_value": 0.0, "realized_value": 0.0, "returned_value": 0.0, "manual_returned_value": 0.0, "reopen_count": 0, "durations": [], "main_plate": "-"})
            per_driver[driver]["returned_stops"] += 1
            per_driver[driver]["returned_value"] += ret_v
            per_driver[driver]["manual_returned_value"] += ret_v
        route_rows.append({"route_id": -d.id, "date": d.data_romaneio, "shift": "-", "driver_id": d.motorista_id, "driver_name": driver, "client_id": d.client_id, "client_name": client, "client_city": (getattr(cli_map.get(d.client_id), "municipio", None) or "").strip() if cli_map.get(d.client_id) else "", "client_bairro": (getattr(cli_map.get(d.client_id), "bairro", None) or "").strip() if cli_map.get(d.client_id) else "", "client_segmento": (getattr(cli_map.get(d.client_id), "segmento", None) or "").strip() if cli_map.get(d.client_id) else "", "client_prioridade": (getattr(cli_map.get(d.client_id), "prioridade_logistica", None) or "").strip() if cli_map.get(d.client_id) else "", "client_status_operacional": (getattr(cli_map.get(d.client_id), "status_operacional", None) or "").strip() if cli_map.get(d.client_id) else "", "client_status_cadastro": (getattr(cli_map.get(d.client_id), "status_cliente", None) or "").strip() if cli_map.get(d.client_id) else "", "client_address": (cli_map.get(d.client_id).get_full_address() or cli_map.get(d.client_id).endereco or "").strip() if cli_map.get(d.client_id) else "", "client_window": (f"{cli_map.get(d.client_id).janela_horario_inicio} - {cli_map.get(d.client_id).janela_horario_fim}" if cli_map.get(d.client_id) and getattr(cli_map.get(d.client_id), "janela_horario_inicio", None) and getattr(cli_map.get(d.client_id), "janela_horario_fim", None) else ""), "visit_time": "", "status": "devolucao", "planned_kg": 0.0, "planned_value": 0.0, "delivered_kg": 0.0, "delivered_value": 0.0, "returned_kg": 0.0, "returned_value": round(ret_v, 2), "reopen_count": 0, "duration_m": None, "plate": "-", "order_number": f"Man. {d.id}", "motivo": motivo, "responsabilidade": resp, "cluster": cluster, "acima_300": above, "source": "MANUAL", "possible_duplicate": is_possible_dup})
        if not is_possible_dup:
            ex_rows.append({"score": 55, "date": d.data_romaneio, "shift": "-", "driver_name": driver, "driver_id": d.motorista_id, "client_name": client, "status": "devolucao", "planned_kg": 0.0, "planned_value": 0.0, "returned_kg": 0.0, "returned_value": round(ret_v, 2), "reopen_count": 0, "duration_m": None, "source": "MANUAL"})

    financial_rows, covered_financial_route_ids = _load_financial_devolucao_rows(
        session=session,
        date_i=date_i,
        date_f=date_f,
        shift=shift,
        driver_id=driver_id,
        plate=plate,
        status=status,
    )
    gap_financial_rows: list[dict] = []
    if financial_rows:
        returned_value = 0.0
        returned_value_manual = 0.0
        ret_count_day = {}
        ret_value_day = {}
        client_returns = {}
        motivo_agg = {}
        resp_agg = {}
        cluster_agg = {}
        driver_return_value: dict[str, float] = {}
        driver_standalone_value: dict[str, float] = {}

        def _accumulate_financial_row(row: dict) -> None:
            nonlocal returned_value, returned_value_manual
            date_key = str(row.get("date") or "")
            driver_name = str(row.get("driver_name") or "-").strip() or "-"
            client_name = str(row.get("client_name") or "Sem cliente").strip() or "Sem cliente"
            motivo = str(row.get("motivo") or "Nao informado").strip() or "Nao informado"
            resp = str(row.get("responsabilidade") or "Nao informado").strip() or "Nao informado"
            cluster = str(row.get("cluster") or "Sem Cluster").strip() or "Sem Cluster"
            val = round(float(row.get("value") or row.get("returned_value") or 0.0), 2)
            returned_value += val
            if row.get("standalone"):
                returned_value_manual += val
                driver_standalone_value[driver_name] = round(driver_standalone_value.get(driver_name, 0.0) + val, 2)
            driver_return_value[driver_name] = round(driver_return_value.get(driver_name, 0.0) + val, 2)
            ret_count_day[date_key] = ret_count_day.get(date_key, 0) + 1
            ret_value_day[date_key] = round(ret_value_day.get(date_key, 0.0) + val, 2)
            motivo_agg.setdefault(motivo, {"motivo": motivo, "qtd": 0, "valor": 0.0})
            motivo_agg[motivo]["qtd"] += 1
            motivo_agg[motivo]["valor"] = round(motivo_agg[motivo]["valor"] + val, 2)
            resp_agg.setdefault(resp, {"responsabilidade": resp, "qtd": 0, "valor": 0.0})
            resp_agg[resp]["qtd"] += 1
            resp_agg[resp]["valor"] = round(resp_agg[resp]["valor"] + val, 2)
            cluster_agg.setdefault(cluster, {"cluster": cluster, "qtd": 0, "valor": 0.0})
            cluster_agg[cluster]["qtd"] += 1
            cluster_agg[cluster]["valor"] = round(cluster_agg[cluster]["valor"] + val, 2)
            cr = client_returns.setdefault(client_name, {"qtd": 0, "valor": 0.0})
            cr["qtd"] += 1
            cr["valor"] = round(cr["valor"] + val, 2)

        for row in financial_rows:
            _accumulate_financial_row(row)

        gap_financial_rows = _financial_gap_rows_from_routes(
            routes, covered_financial_route_ids, emp_map, cli_map
        )
        for row in gap_financial_rows:
            _accumulate_financial_row(row)

        for driver_name, driver_data in per_driver.items():
            driver_data["returned_value"] = round(driver_return_value.get(driver_name, 0.0), 2)
            driver_data["manual_returned_value"] = round(driver_standalone_value.get(driver_name, 0.0), 2)
        for daily in per_day.values():
            daily["returned_value"] = 0.0
        for date_key, val in ret_value_day.items():
            per_day.setdefault(
                date_key,
                {
                    "date": date_key,
                    "planned_stops": 0,
                    "started_stops": 0,
                    "realized_stops": 0,
                    "returned_stops": 0,
                    "planned_kg": 0.0,
                    "realized_kg": 0.0,
                    "returned_kg": 0.0,
                    "planned_value": 0.0,
                    "returned_value": 0.0,
                },
            )
            per_day[date_key]["returned_value"] = round(val, 2)

    financial_rows_all = (financial_rows + gap_financial_rows) if financial_rows else []

    # Valor devolvido oficial = cadastro (Devolucao + lacunas rota sem registro), independente de ajuste de responsabilidade.
    if financial_rows_all:
        canonical_returned_value = round(
            sum(float(r.get("value") or r.get("returned_value") or 0.0) for r in financial_rows_all),
            2,
        )
    else:
        canonical_returned_value = round(float(returned_value), 2)
    returned_value = canonical_returned_value

    avg_duration = statistics.mean(dur_list) if dur_list else 0.0
    global_return_rate = (returned_stops / max(1, planned_stops) * 100.0) if (planned_stops or manual) else 0.0
    tactical = []
    for row in per_driver.values():
        p = max(1, row["planned_stops"])
        value_base = (row["planned_value"] or 0.0) + (row.get("manual_returned_value") or 0.0)
        if value_base <= 0:
            value_base = (row["realized_value"] or 0.0) + (row["returned_value"] or 0.0)
        delivered_stops = max(0, row["realized_stops"] - row["returned_stops"])
        tactical.append({
            **row,
            "delivered_stops": delivered_stops,
            "efficiency": round(delivered_stops / p * 100.0, 2) if row["planned_stops"] else 0.0,
            "return_rate": round(row["returned_stops"] / p * 100.0, 2) if row["planned_stops"] else (100.0 if row["returned_stops"] > 0 else 0.0),
            "started_rate": round(row["started_stops"] / p * 100.0, 2) if row["planned_stops"] else 0.0,
            "avg_duration": round(statistics.mean(row["durations"]), 1) if row["durations"] else 0.0,
            "returned_value_pct": round(_pct(row["returned_value"], value_base), 2)
        })
    tactical.sort(key=lambda x: (x["efficiency"], -x["return_rate"], x["planned_stops"]), reverse=True)

    daily_rows = []
    for day in sorted(per_day):
        row = per_day[day]
        denom = max(1, row["planned_stops"])
        row["started_rate"] = round(row["started_stops"] / denom * 100.0, 2)
        row["return_rate"] = round(row["returned_stops"] / denom * 100.0, 2)
        daily_rows.append(row)
    rec7 = daily_rows[-7:] if len(daily_rows) >= 7 else daily_rows
    forecast_stops = round(statistics.mean([x["planned_stops"] for x in rec7]), 1) if rec7 else 0.0
    forecast_return_qtd = round(statistics.mean([x["return_rate"] for x in rec7]), 2) if rec7 else 0.0
    forecast_return_value = round(
        statistics.mean([
            ((x.get("returned_value", 0.0) or 0.0) / max(0.01, (x.get("planned_value", 0.0) or 0.0))) * 100.0
            if (x.get("planned_value", 0.0) or 0.0) > 0 else 0.0
            for x in rec7
        ]),
        2
    ) if rec7 else 0.0
    risk_label, risk_severity = ("Critico", "danger") if forecast_return_value >= 4 else ("Atencao", "warning") if forecast_return_value >= 2 else ("Controlado", "success")
    anomaly_flags: list[str] = []
    # % valor: denominador = faturamento efetivo (entregue + devolvido), incluindo todo valor devolvido cadastrado.
    financial_base_value = max(realized_value + returned_value, 0.01)
    global_return_rate_value = _pct(returned_value, financial_base_value)

    if global_return_rate_value >= 2:
        anomaly_flags.append(f"Ponto de atencao financeiro: devolucao em valor em {_fmt_br_1(global_return_rate_value)}% (meta <= 2,0%).")
    if client_returns:
        rec_client = [(n, c) for n, c in sorted(client_returns.items(), key=lambda x: x[1]["qtd"], reverse=True) if c["qtd"] >= 2][:3]
        if rec_client:
            parts = []
            for nome, d in rec_client:
                pct_val = _pct(d["valor"], financial_base_value)
                parts.append(f"{nome} ({d['qtd']} devolucoes, {_fmt_br_1(pct_val)}% do valor)")
            anomaly_flags.append(f"Clientes com devolucao recorrente: {'; '.join(parts)}.")
    if cluster_agg:
        cl = max(cluster_agg.values(), key=lambda x: x["valor"])
        anomaly_flags.append(f"Cluster critico por valor: {cl['cluster']} (R$ {_fmt_br_2(cl['valor'])}).")
    if ret_value_day:
        dpk, vpk = max(ret_value_day.items(), key=lambda x: x[1])
        anomaly_flags.append(f"Pico de devolucao em {_fmt_br_data(dpk)}: R$ {_fmt_br_2(vpk)}.")
    high_value_drivers = sorted(
        [x for x in tactical if (x.get("returned_value_pct") or 0) >= 2 and (x.get("planned_stops") or 0) >= 5],
        key=lambda x: (x.get("returned_value_pct") or 0),
        reverse=True
    )[:3]
    if high_value_drivers:
        anomaly_flags.append(
            "Motoristas acima da meta financeira: " +
            ", ".join([f"{x['driver_name']} ({_fmt_br_1(x.get('returned_value_pct') or 0)}%)" for x in high_value_drivers]) +
            "."
        )
    dias_acima_meta = [
        d for d, v in per_day.items()
        if (v.get("planned_value", 0.0) or 0.0) > 0 and _pct((v.get("returned_value", 0.0) or 0.0), (v.get("planned_value", 0.0) or 0.0)) >= 2
    ]
    if len(dias_acima_meta) >= 3:
        anomaly_flags.append(f"Serie de risco: {len(dias_acima_meta)} dia(s) com devolucao em valor acima de 2,0%.")

    # Detectar possíveis duplicatas ROTA + MANUAL (mesma data, cliente, motorista e valor)
    rota_devol_keys: set[tuple] = set()
    for r in route_rows:
        if (r.get("source") or "").upper() == "ROTA" and (r.get("status") or "").lower() == "devolucao":
            k = (r.get("date"), r.get("client_id"), r.get("driver_id"), round(float(r.get("returned_value") or 0), 2))
            rota_devol_keys.add(k)
    dup_count = 0
    for r in route_rows:
        if (r.get("source") or "").upper() != "MANUAL":
            continue
        k = (r.get("date"), r.get("client_id"), r.get("driver_id"), round(float(r.get("returned_value") or 0), 2))
        if k in rota_devol_keys:
            dup_count += 1
    if dup_count > 0:
        anomaly_flags.append(f"Possivel duplicata ROTA+MANUAL: {dup_count} devolucao(oes) com mesma data, cliente, motorista e valor. Revisar cadastro.")

    recommendations: list[str] = []
    if dup_count > 0:
        recommendations.append("Revisar devolucoes manuais que coincidem com rotas ja marcadas como devolucao (evitar contagem dupla no BI).")
    if global_return_rate_value >= 2:
        recommendations.append("Priorizar plano de contencao financeira: devolver no maximo 2,0% do valor real.")
    if tactical:
        wr = max(tactical, key=lambda x: x["returned_value_pct"])
        if wr["returned_value_pct"] >= 2 and wr["planned_stops"] >= 5:
            recommendations.append(f"Rebalancear carteira de {wr['driver_name']} (devolucao em valor: {_fmt_br_1(wr['returned_value_pct'])}%).")
    if started_stops < planned_stops:
        recommendations.append("Atuar na fila de pendentes com prioridade por maior impacto.")
    if avg_duration >= 120:
        recommendations.append("Tempo medio elevado: revisar sequencia e janela de entrega.")
    # Exclui "Não informado" / "Nao informado" dos gráficos (devem vir de dados mal preenchidos)
    _nao_inf = ("nao informado", "não informado")
    motivo_agg_ok = {k: v for k, v in motivo_agg.items() if (k or "").strip().lower() not in _nao_inf}
    resp_agg_ok = {k: v for k, v in resp_agg.items() if (k or "").strip().lower() not in _nao_inf}
    if motivo_agg_ok:
        top_motivo_val = max(motivo_agg_ok.values(), key=lambda x: x["valor"])
        recommendations.append(f"Estudo 80/20: atacar primeiro o motivo '{top_motivo_val['motivo']}' (R$ {_fmt_br_2(top_motivo_val['valor'])}).")
    if client_returns:
        top_cliente, top_d = max(client_returns.items(), key=lambda x: x[1]["qtd"])
        top_qtd = top_d["qtd"]
        if top_qtd >= 2:
            recommendations.append(f"Estudo de causa raiz com cliente {top_cliente}: {top_qtd} devolucoes no periodo.")
    if not recommendations:
        recommendations.append("Operacao estavel no periodo; manter monitoramento diario.")

    total_mot = sum(v["qtd"] for v in motivo_agg_ok.values()) or 1
    motivos_rows = sorted([{"motivo": v["motivo"], "qtd": int(v["qtd"]), "valor": round(v["valor"], 2), "pct": round(v["qtd"] / total_mot * 100.0, 2)} for v in motivo_agg_ok.values()], key=lambda x: (x["qtd"], x["valor"]), reverse=True)

    # Motivo x Motorista x Qtd x Valor x % valor real (para gráfico detalhado)
    financial_detail_rows = financial_rows_all if financial_rows_all else [
        {
            "date": r.get("date"),
            "driver_name": r.get("driver_name"),
            "client_name": r.get("client_name"),
            "returned_value": r.get("returned_value"),
            "motivo": r.get("motivo"),
            "responsabilidade": r.get("responsabilidade"),
        }
        for r in route_rows
        if r.get("status") == "devolucao" and (r.get("returned_value") or 0) > 0
    ]
    motivo_motorista: dict[str, dict[str, dict]] = {}
    for r in financial_detail_rows:
        motivo = (r.get("motivo") or "").strip() or "Nao informado"
        if motivo.upper() == "ENCERRAMENTO TARDIO AUTOMATICO" or motivo.lower() in _nao_inf:
            continue
        driver = r.get("driver_name") or "-"
        val = float(r.get("returned_value") or 0.0)
        motivo_motorista.setdefault(motivo, {})
        motivo_motorista[motivo].setdefault(driver, {"qtd": 0, "valor": 0.0})
        motivo_motorista[motivo][driver]["qtd"] += 1
        motivo_motorista[motivo][driver]["valor"] = round(motivo_motorista[motivo][driver]["valor"] + val, 2)
    motivo_names = sorted(motivo_motorista.keys(), key=lambda m: -sum(d["valor"] for d in motivo_motorista[m].values()))[:15]
    driver_names_mot = sorted({d for m_data in motivo_motorista.values() for d in m_data.keys()})
    motivos_detailed = []
    for motivo in motivo_names:
        row = {"motivo": motivo, "qtd": 0, "valor": 0.0, "pct": 0.0, "por_motorista": {}}
        for driver, data in motivo_motorista[motivo].items():
            row["qtd"] += data["qtd"]
            row["valor"] += data["valor"]
            valor_real_drv = (per_driver.get(driver, {}).get("realized_value") or 0.0) + (per_driver.get(driver, {}).get("returned_value") or 0.0)
            pct_val = round(data["valor"] / valor_real_drv * 100.0, 2) if valor_real_drv > 0 else (100.0 if data["valor"] > 0 else 0.0)
            row["por_motorista"][driver] = {"qtd": data["qtd"], "valor": data["valor"], "pct_valor_real": pct_val}
        row["valor"] = round(row["valor"], 2)
        row["pct"] = round(row["qtd"] / total_mot * 100.0, 2) if total_mot else 0.0
        motivos_detailed.append(row)
    resp_rows = sorted([{"responsabilidade": v["responsabilidade"], "qtd": int(v["qtd"]), "valor": round(v["valor"], 2)} for v in resp_agg_ok.values()], key=lambda x: x["qtd"], reverse=True)
    cluster_rows = sorted([{"cluster": v["cluster"], "qtd": int(v["qtd"]), "valor": round(v["valor"], 2)} for v in cluster_agg.values()], key=lambda x: x["valor"], reverse=True)

    # Motorista x Responsabilidade x Valor: valor devolvido por motorista e responsabilidade, % devolução baseada em valor real
    driver_resp_value: dict[str, dict[str, float]] = {}
    for r in financial_detail_rows:
        motivo_raw = (r.get("motivo") or "").strip().upper()
        if motivo_raw == "ENCERRAMENTO TARDIO AUTOMATICO":
            continue
        drv = r.get("driver_name") or "-"
        resp = (r.get("responsabilidade") or "").strip() or "Nao informado"
        if resp.lower() in _nao_inf:
            continue
        val = float(r.get("returned_value") or 0.0)
        driver_resp_value.setdefault(drv, {})
        driver_resp_value[drv][resp] = round(driver_resp_value[drv].get(resp, 0.0) + val, 2)
    resp_names = sorted({r for drv_data in driver_resp_value.values() for r in drv_data.keys()})
    driver_resp_rows = []
    for drv, data in per_driver.items():
        resp_vals = {r: driver_resp_value.get(drv, {}).get(r, 0.0) for r in resp_names}
        total_ret = sum(resp_vals.values())
        valor_real = (data.get("realized_value") or 0.0) + (data.get("returned_value") or 0.0)
        pct_devolucao_valor = round(total_ret / valor_real * 100.0, 2) if valor_real > 0 else (100.0 if total_ret > 0 else 0.0)
        driver_resp_rows.append({
            "driver": drv,
            "responsabilidades": resp_vals,
            "valor_devolvido": round(total_ret, 2),
            "valor_real": round(valor_real, 2),
            "pct_devolucao_valor": pct_devolucao_valor,
        })
    driver_resp_rows = [r for r in driver_resp_rows if r["valor_devolvido"] > 0]
    driver_resp_rows = sorted(driver_resp_rows, key=lambda x: -x["valor_devolvido"])[:20]

    # Correlacao Motorista x Cliente x Devolucoes (bubble/scatter)
    pair_agg: dict[tuple[str, str], dict] = {}
    for r in financial_detail_rows:
        drv = r.get("driver_name") or "-"
        cli = r.get("client_name") or "Sem cliente"
        val = float(r.get("returned_value") or 0.0)
        key = (drv, cli)
        pair_agg.setdefault(key, {"driver": drv, "client": cli, "qtd": 0, "valor": 0.0})
        pair_agg[key]["qtd"] += 1
        pair_agg[key]["valor"] = round(pair_agg[key]["valor"] + val, 2)

    pair_rows = sorted(pair_agg.values(), key=lambda x: (x["valor"], x["qtd"]), reverse=True)[:120]
    corr_drivers = sorted({x["driver"] for x in pair_rows})
    corr_clients = sorted({x["client"] for x in pair_rows})
    corr_driver_idx = {d: i for i, d in enumerate(corr_drivers)}
    corr_client_idx = {c: i for i, c in enumerate(corr_clients)}
    corr_points = []
    for p in pair_rows:
        d_data = per_driver.get(p["driver"], {})
        valor_real_drv = (d_data.get("realized_value") or 0.0) + (d_data.get("returned_value") or 0.0)
        pct_val = round((p["valor"] / valor_real_drv) * 100.0, 2) if valor_real_drv > 0 else (100.0 if p["valor"] > 0 else 0.0)
        corr_points.append({
            "x": corr_driver_idx.get(p["driver"], 0),
            "y": corr_client_idx.get(p["client"], 0),
            "r": max(4, min(20, 4 + (p["valor"] ** 0.5) * 0.12)),
            "driver": p["driver"],
            "client": p["client"],
            "qtd": p["qtd"],
            "valor": round(p["valor"], 2),
            "pct_valor_real": pct_val,
        })
    def _prev_month_key(date_key: str) -> Optional[str]:
        try:
            dref = datetime.strptime(str(date_key), "%Y-%m-%d").date()
            y = dref.year
            m = dref.month - 1
            if m == 0:
                m = 12
                y -= 1
            # Ajusta para o ultimo dia valido do mes anterior (ex.: 31 -> 30/28/29)
            max_day = [31, 29 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1]
            d = min(dref.day, max_day)
            return date(y, m, d).strftime("%Y-%m-%d")
        except Exception:
            return None

    trend_dates = sorted(set(list(per_day.keys()) + list(ret_count_day.keys())))
    trend_qtd = [ret_count_day.get(k, 0) for k in trend_dates]
    trend_val = [round(ret_value_day.get(k, 0.0), 2) for k in trend_dates]
    trend_meta_2pct = [round((per_day.get(k, {}).get("planned_value", 0.0) or 0.0) * 0.02, 2) for k in trend_dates]
    trend_last_month_val: list[Optional[float]] = []
    for _ in trend_dates:
        trend_last_month_val.append(None)
    heat_rows = [{"date": dt, "driver": drv, "value": v} for dt, d in reopen_heat.items() for drv, v in d.items()]

    filters_payload = {"date_from": date_i.strftime("%Y-%m-%d"), "date_to": date_f.strftime("%Y-%m-%d"), "shift": shift, "driver_id": driver_id, "plate": plate, "status": status, "detail_driver_id": detail_driver_id, "detail_status": detail_status}
    filters_query = urlencode({"date_from": filters_payload["date_from"], "date_to": filters_payload["date_to"], "shift": filters_payload["shift"], "driver_id": filters_payload["driver_id"] or "", "plate": filters_payload["plate"], "status": filters_payload["status"]})
    # Exclui MANUAL duplicado: mesma devolução já registrada em ROTA (evita linha duplicada no drill-through)
    detail_rows = [
        r for r in route_rows
        if not (str(r.get("source") or "").upper() == "MANUAL" and r.get("possible_duplicate"))
    ]
    detail_rows = sorted(detail_rows, key=lambda x: (x["date"], x["shift"], x["driver_name"], x["route_id"]))
    if detail_driver_id:
        detail_rows = [r for r in detail_rows if r["driver_id"] == detail_driver_id]
    if detail_status_norm != "todos":
        detail_rows = [r for r in detail_rows if r["status"] == detail_status_norm]

    delivered_stops = max(0, realized_stops - returned_stops)

    # Devolução mês anterior (query dedicada)
    def _prev_month_range(d: date) -> tuple[date, date]:
        y, m = d.year, d.month - 1
        if m == 0:
            m, y = 12, y - 1
        first = date(y, m, 1)
        last = date(y, m, [31, 29 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1])
        return first, last
    prev_first, prev_last = _prev_month_range(date_i)
    prev_str_first = prev_first.strftime("%Y-%m-%d")
    prev_str_last = prev_last.strftime("%Y-%m-%d")
    q_prev = (
        select(models.Route)
        .where(models.Route.type == "delivery")
        .where(models.Route.date >= prev_str_first)
        .where(models.Route.date <= prev_str_last)
    )
    if shift and shift != "Todos":
        q_prev = q_prev.where(models.Route.shift == shift)
    if driver_id:
        q_prev = q_prev.where(models.Route.employee_id == driver_id)
    if pl and pl != "TODOS":
        q_prev = q_prev.where(models.Route.delivery_vehicle_plate == pl)
    prev_routes = session.exec(q_prev).all()
    prev_month_planned_val = round(sum(float(r.valor_financeiro or 0.0) for r in prev_routes), 2)
    prev_financial_rows, prev_covered_route_ids = _load_financial_devolucao_rows(
        session=session,
        date_i=prev_first,
        date_f=prev_last,
        shift=shift,
        driver_id=driver_id,
        plate=plate,
        status=status,
    )
    prev_emp_ids = sorted({r.employee_id for r in prev_routes if r.employee_id})
    prev_cli_ids = sorted({r.client_id for r in prev_routes if r.client_id})
    prev_emp_map = {
        e.id: e
        for e in (
            session.exec(select(models.Employee).where(models.Employee.id.in_(prev_emp_ids))).all()
            if prev_emp_ids
            else []
        )
    }
    prev_cli_map = {
        c.id: c
        for c in (
            session.exec(select(models.Client).where(models.Client.id.in_(prev_cli_ids))).all()
            if prev_cli_ids
            else []
        )
    }
    prev_gap_rows = _financial_gap_rows_from_routes(
        prev_routes, prev_covered_route_ids, prev_emp_map, prev_cli_map
    )
    prev_financial_rows_all = prev_financial_rows + prev_gap_rows
    prev_month_qtd = len(prev_financial_rows_all)
    prev_month_valor = round(
        sum(float(r.get("value") or r.get("returned_value") or 0.0) for r in prev_financial_rows_all), 2
    )
    prev_month_manual_valor = round(
        sum(float(r.get("value") or r.get("returned_value") or 0.0) for r in prev_financial_rows_all if r.get("standalone")),
        2,
    )
    prev_month_base = prev_month_planned_val + prev_month_manual_valor
    if prev_month_base <= 0:
        prev_month_base = prev_month_valor
    prev_month_pct = round(_pct(prev_month_valor, prev_month_base), 2)
    prev_ret_value_day: dict[str, float] = {}
    for row in prev_financial_rows_all:
        prev_key = str(row.get("date") or "")
        prev_ret_value_day[prev_key] = round(
            prev_ret_value_day.get(prev_key, 0.0) + float(row.get("value") or row.get("returned_value") or 0.0),
            2,
        )
    trend_last_month_val = []
    for k in trend_dates:
        pm = _prev_month_key(k)
        trend_last_month_val.append(round(prev_ret_value_day.get(pm, 0.0), 2) if pm else None)

    devolucoes_acima_300_count = len(
        [
            r for r in (financial_rows_all or [])
            if str(r.get("acima_300") or "").upper() == "SIM"
        ]
    ) if financial_rows_all else len([r for r in route_rows if r.get("acima_300") == "SIM"])

    kpis = {
        "planned_stops": planned_stops,
        "realized_stops": realized_stops,
        "delivered_stops": delivered_stops,
        "started_stops": started_stops,
        "pending_stops": max(0, planned_stops - started_stops),
        "planned_kg": round(planned_kg, 2),
        "realized_kg": round(realized_kg, 2),
        "returned_kg": round(returned_kg, 2),
        "planned_value": round(planned_value, 2),
        "realized_value": round(realized_value, 2),
        "returned_value": round(returned_value, 2),
        "return_rate_qtd": round(returned_stops / max(1, planned_stops) * 100.0, 2) if planned_stops else 0.0,
        "return_rate_rotas": return_rate_rotas,
        "rotas_devolucao_count": rotas_devolucao_count,
        "rotas_concluidas_count": rotas_concluidas_count,
        "return_rate_kg": round(returned_kg / max(1, planned_kg) * 100.0, 2) if planned_kg else 0.0,
        "return_rate_value": round(global_return_rate_value, 2),
        "sla_start": round(started_stops / max(1, planned_stops) * 100.0, 2) if planned_stops else 0.0,
        "sla_finish": round(delivered_stops / max(1, planned_stops) * 100.0, 2) if planned_stops else 0.0,
        "reopen_index": round(reopen_routes / max(1, planned_stops) * 100.0, 2) if planned_stops else 0.0,
        "avg_duration_m": round(avg_duration, 1),
        "forecast_next_stops": forecast_stops,
        "forecast_next_return_rate": forecast_return_value,
        "forecast_next_return_rate_qtd": forecast_return_qtd,
        "forecast_next_return_rate_value": forecast_return_value,
        "total_devolucoes": returned_stops,
        "valor_total_devolvido": round(returned_value, 2),
        "devolucoes_acima_300_count": devolucoes_acima_300_count,
        "devolucoes_acima_300_pct": round((devolucoes_acima_300_count / max(1, len(financial_rows_all) or returned_stops)) * 100.0, 2) if (financial_rows_all or returned_stops) else 0.0,
        "risk_label": risk_label,
        "risk_severity": risk_severity,
        "meta_devolucao_pct": 2.0,
        "devolucao_mes_anterior_qtd": prev_month_qtd,
        "devolucao_mes_anterior_valor": round(prev_month_valor, 2),
        "devolucao_mes_anterior_pct": prev_month_pct,
    }

    delivery_daily_chart = [
        {
            "date": str(r.get("date") or ""),
            "planned_stops": int(r.get("planned_stops") or 0),
            "delivered_stops": max(0, int(r.get("realized_stops") or 0) - int(r.get("returned_stops") or 0)),
            "started_stops": int(r.get("started_stops") or 0),
            "planned_kg": round(float(r.get("planned_kg") or 0.0), 2),
            "realized_kg": round(float(r.get("realized_kg") or 0.0), 2),
        }
        for r in daily_rows
    ]

    chart_payload = {
        "delivery_daily": delivery_daily_chart,
        "trend": {
            "dates": trend_dates,
            "qtd": trend_qtd,
            "valor": trend_val,
            "ma7": _ma(trend_val, 7),
            "ma30": _ma(trend_val, 30),
            "meta_2pct": trend_meta_2pct,
            "last_month_valor": trend_last_month_val,
        },
        "motivos": motivos_rows,
        "motivos_detailed": motivos_detailed,
        "motivos_drivers": driver_names_mot,
        "responsabilidade": resp_rows,
        "cluster": cluster_rows,
        "drivers": [{"driver": r["driver_name"], "eficiencia": r["efficiency"], "devolucao_pct": r["return_rate"], "devolucao_valor_pct": r["returned_value_pct"], "valor_devolvido": r["returned_value"]} for r in tactical][:20],
        "reopen_heatmap": heat_rows,
        "driver_resp_valor": {
            "drivers": [r["driver"] for r in driver_resp_rows],
            "responsabilidades": resp_names,
            "datasets": [{"resp": r, "data": [driver_resp_value.get(drv, {}).get(r, 0.0) for drv in [x["driver"] for x in driver_resp_rows]]} for r in resp_names] if resp_names else [],
            "pct_devolucao_valor": [r["pct_devolucao_valor"] for r in driver_resp_rows],
        },
        "driver_client_corr": {
            "drivers": corr_drivers,
            "clients": corr_clients,
            "points": corr_points,
        },
    }

    return {
        "filters": filters_payload,
        "kpis": kpis,
        "daily_rows": daily_rows,
        "tactical_rows": tactical,
        "exception_rows": sorted(ex_rows, key=lambda x: x["score"], reverse=True)[:25],
        "anomaly_flags": anomaly_flags[:10],
        "recommendations": recommendations[:6],
        "drivers_filter": sorted([{"id": d["driver_id"], "name": d["driver_name"]} for d in tactical], key=lambda x: x["name"]),
        "plates_filter": sorted({x["main_plate"] for x in tactical if x["main_plate"] and x["main_plate"] != "-"}),
        "fleet_plates": sorted(
            _get_fleet_plates(session)
        ),
        "statuses_filter": ["Todos", "Pendente", "Iniciada", "Entregue", "Devolucao", "Reaberta", "Cancelada"],
        "detail_rows": detail_rows[:300],
        "detail_title": "Drill-through operacional",
        "detail_total": len(detail_rows),
        "filters_query": filters_query,
        "all_route_rows": route_rows,
        "all_financial_rows": financial_rows_all,
        "motivos_rows": motivos_rows[:12],
        "responsabilidade_rows": resp_rows,
        "cluster_rows": cluster_rows,
        "chart_payload_json": json.dumps(chart_payload, ensure_ascii=False),
        "detail_rows_json": json.dumps(detail_rows, ensure_ascii=False),
    }


def _build_bi_clientes_dataset(
    session: Session,
    date_from: Optional[str],
    date_to: Optional[str],
    shift: str,
    driver_id: Optional[int],
    plate: str,
    status: str,
    client_id: Optional[int] = None,
    city: str = "Todos",
    priority: str = "Todos",
    client_status: str = "Todos",
    segmentos: Optional[list[str]] = None,
    returns_filter: str = "Todos",
    detail_client_id: Optional[int] = None,
    client_filter_scope: str = "solo",
    vendedor_id: Optional[int] = None,
    search_q: str = "",
    classification_filter: str = "Todos",
    motivo_filter: str = "Todos",
    responsabilidade_filter: str = "Todos",
    purchase_band: str = "Todos",
    return_band: str = "Todos",
    duration_band: str = "Todos",
) -> dict:
    tz = ZoneInfo("America/Sao_Paulo")
    today = datetime.now(tz).date()

    def _parse_d(raw: Optional[str]) -> Optional[date]:
        if not raw:
            return None
        try:
            return datetime.strptime(str(raw).strip(), "%Y-%m-%d").date()
        except Exception:
            return None

    date_i = _parse_d(date_from) or today.replace(day=1)
    date_f = _parse_d(date_to) or today
    if date_i > date_f:
        date_i, date_f = date_f, date_i
    cur_s, cur_e = date_i.strftime("%Y-%m-%d"), date_f.strftime("%Y-%m-%d")
    current_days = max(1, (date_f - date_i).days + 1)
    previous_date_to = date_i - timedelta(days=1)
    previous_date_from = previous_date_to - timedelta(days=current_days - 1)
    prev_s = previous_date_from.strftime("%Y-%m-%d")
    prev_e = previous_date_to.strftime("%Y-%m-%d")

    # Uma única leitura de rotas + manuais (período atual + anterior contíguo) — evita 2ª ida ao Postgres.
    merged_dataset = _build_bi_delivery_dataset(
        session=session,
        date_from=prev_s,
        date_to=cur_e,
        shift=shift,
        driver_id=driver_id,
        plate=plate,
        status=status,
    )
    all_merged = list(merged_dataset.get("all_route_rows", []))
    all_financial = list(merged_dataset.get("all_financial_rows", []))
    base_rows = [r for r in all_merged if cur_s <= r["date"] <= cur_e]
    previous_rows_pool = [r for r in all_merged if prev_s <= r["date"] <= prev_e]
    base_financial_rows = [r for r in all_financial if cur_s <= r["date"] <= cur_e]
    previous_financial_rows_pool = [r for r in all_financial if prev_s <= r["date"] <= prev_e]

    _drivers_map: dict[int, str] = {}
    _plates_set: set = set()
    for r in base_rows:
        did = r.get("driver_id")
        if did is not None:
            _drivers_map[int(did)] = str(r.get("driver_name") or f"Motorista #{did}")
        pl = r.get("plate")
        if pl and str(pl).strip() not in ("", "-"):
            _plates_set.add(str(pl).strip().upper())
    delivery_dataset = {
        "filters": {**merged_dataset.get("filters", {}), "date_from": cur_s, "date_to": cur_e},
        "drivers_filter": sorted([{"id": i, "name": n} for i, n in _drivers_map.items()], key=lambda x: x["name"]),
        "plates_filter": sorted(_plates_set),
        "statuses_filter": merged_dataset.get("statuses_filter", []),
    }

    scope = (client_filter_scope or "solo").strip().lower()
    if scope not in ("solo", "group"):
        scope = "solo"
    client_ids_filter = None
    aggregate_group = None
    group_filter_note = None
    if client_id:
        if scope == "group":
            cobj = session.get(models.Client, int(client_id))
            gid = getattr(cobj, "client_group_id", None) if cobj else None
            if gid:
                mids = session.exec(select(models.Client.id).where(models.Client.client_group_id == gid)).all()
                client_ids_filter = {int(x) for x in mids if x is not None}
                if len(client_ids_filter) > 1:
                    cg = session.get(models.ClientGroup, gid)
                    aggregate_group = (int(gid), (cg.name if cg else f"Grupo #{gid}"))
            if not client_ids_filter:
                client_ids_filter = {int(client_id)}
            elif len(client_ids_filter) == 1:
                group_filter_note = "Apenas uma loja neste grupo; mesmo que filtrar só ela."
        else:
            client_ids_filter = {int(client_id)}

    selected_segmentos = [str(s or "").strip() for s in (segmentos or []) if str(s or "").strip()]
    segmento_norm_set = {_norm_text(s) for s in selected_segmentos if _norm_text(s) and _norm_text(s) != _norm_text("Todos")}
    city_norm = "" if (city or "Todos").strip() == "Todos" else _norm_text(city)
    priority_norm = "" if (priority or "Todos").strip() == "Todos" else str(priority or "").strip().upper()
    client_status_norm = "" if (client_status or "Todos").strip() == "Todos" else _norm_text(client_status)

    clients_filter_map: dict[int, str] = {}
    cities_filter = set()
    priorities_filter = set()
    client_statuses_filter = set()
    segmentos_filter = set()
    for row in base_rows:
        row_client_id = row.get("client_id")
        if row_client_id is not None:
            clients_filter_map[int(row_client_id)] = str(row.get("client_name") or f"Cliente #{row_client_id}")
        row_city = str(row.get("client_city") or "").strip()
        row_priority = str(row.get("client_prioridade") or "").strip().upper()
        row_client_status = str(row.get("client_status_operacional") or "").strip()
        row_segmento = str(row.get("client_segmento") or "").strip()
        if row_city:
            cities_filter.add(row_city)
        if row_priority:
            priorities_filter.add(row_priority)
        if row_client_status:
            client_statuses_filter.add(row_client_status)
        if row_segmento:
            segmentos_filter.add(row_segmento)

    def _apply_client_filters(rows: list[dict]) -> list[dict]:
        filtered = []
        for row in rows:
            row_client_id = row.get("client_id")
            row_city = _norm_text(row.get("client_city"))
            row_priority = str(row.get("client_prioridade") or "").strip().upper()
            row_client_status = _norm_text(row.get("client_status_operacional"))
            row_segmento = _norm_text(row.get("client_segmento"))
            if client_ids_filter is not None:
                if row_client_id is None or int(row_client_id) not in client_ids_filter:
                    continue
            if city_norm and row_city != city_norm:
                continue
            if priority_norm and row_priority != priority_norm:
                continue
            if client_status_norm and row_client_status != client_status_norm:
                continue
            if segmento_norm_set and row_segmento not in segmento_norm_set:
                continue
            filtered.append(row)
        return filtered

    filtered_rows = _apply_client_filters(base_rows)
    filtered_financial_rows = _apply_client_filters(base_financial_rows)

    _seller_scope_ids = sorted({int(r["client_id"]) for r in base_rows if r.get("client_id") is not None})
    sellers_filter: list[dict] = []
    if _seller_scope_ids:
        _cl_seller_rows = session.exec(select(models.Client).where(models.Client.id.in_(_seller_scope_ids))).all()
        _vids_sellers = sorted({int(c.vendedor_id) for c in _cl_seller_rows if c.vendedor_id is not None})
        _emp_sellers = (
            session.exec(select(models.Employee).where(models.Employee.id.in_(_vids_sellers))).all() if _vids_sellers else []
        )
        _emp_seller_map = {e.id: e for e in _emp_sellers}
        _seen_vid: set[int] = set()
        for c in _cl_seller_rows:
            if c.vendedor_id is None or int(c.vendedor_id) in _seen_vid:
                continue
            _seen_vid.add(int(c.vendedor_id))
            ev = _emp_seller_map.get(int(c.vendedor_id))
            sellers_filter.append(
                {
                    "id": int(c.vendedor_id),
                    "name": str(ev.name or f"Vendedor #{c.vendedor_id}") if ev else f"Vendedor #{c.vendedor_id}",
                }
            )
        sellers_filter.sort(key=lambda x: x["name"])

    motivos_filter_options = sorted(
        {
            str(x.get("motivo") or "").strip()
            for x in (filtered_rows + filtered_financial_rows)
            if str(x.get("motivo") or "").strip() and _norm_text(str(x.get("motivo"))) != "nao informado"
        }
    )
    responsabilidades_filter_options = sorted(
        {
            str(x.get("responsabilidade") or "").strip()
            for x in (filtered_rows + filtered_financial_rows)
            if str(x.get("responsabilidade") or "").strip() and _norm_text(str(x.get("responsabilidade"))) != "nao informado"
        }
    )

    detail_client_id = detail_client_id or client_id

    def _rows_for_group_aggregate(rows: list[dict]) -> list[dict]:
        if not aggregate_group or not client_ids_filter or len(client_ids_filter) <= 1:
            return rows
        gid, gname = aggregate_group
        synth_id = -int(gid)
        out = []
        for row in rows:
            cid = row.get("client_id")
            if cid is None or int(cid) not in client_ids_filter:
                continue
            r = dict(row)
            r["client_id"] = synth_id
            r["client_name"] = f"Grupo: {gname}"
            r["client_city"] = "—"
            r["client_bairro"] = "—"
            r["client_segmento"] = "—"
            r["client_prioridade"] = "—"
            r["client_status_operacional"] = "—"
            r["client_status_cadastro"] = "—"
            r["client_address"] = f"{len(client_ids_filter)} lojas"
            r["client_window"] = "—"
            out.append(r)
        return out

    rows_for_agg = _rows_for_group_aggregate(filtered_rows)

    def _risk_label(score: int) -> tuple[str, str]:
        if score >= 70:
            return ("Crítico", "danger")
        if score >= 40:
            return ("Atenção", "warning")
        return ("Controlado", "success")

    client_agg = _aggregate_bi_client_rows(rows_for_agg, _rows_for_group_aggregate(filtered_financial_rows))

    filters_base = delivery_dataset.get("filters", {})
    current_date_from = datetime.strptime(str(filters_base.get("date_from") or date_from), "%Y-%m-%d").date()
    current_date_to = datetime.strptime(str(filters_base.get("date_to") or date_to), "%Y-%m-%d").date()
    current_days = max(1, (current_date_to - current_date_from).days + 1)
    previous_date_to = current_date_from - timedelta(days=1)
    previous_date_from = previous_date_to - timedelta(days=current_days - 1)
    previous_label = f"{_fmt_br_data(previous_date_from.strftime('%Y-%m-%d'))} a {_fmt_br_data(previous_date_to.strftime('%Y-%m-%d'))}"
    current_label = f"{_fmt_br_data(current_date_from.strftime('%Y-%m-%d'))} a {_fmt_br_data(current_date_to.strftime('%Y-%m-%d'))}"

    previous_rows = _apply_client_filters(previous_rows_pool)
    previous_financial_rows = _apply_client_filters(previous_financial_rows_pool)
    previous_client_agg = _aggregate_bi_client_rows(
        _rows_for_group_aggregate(previous_rows),
        _rows_for_group_aggregate(previous_financial_rows),
    )

    _exec_gid = None
    _exec_mem = None
    if aggregate_group and client_ids_filter and len(client_ids_filter) > 1:
        _exec_gid = int(aggregate_group[0])
        _exec_mem = set(int(x) for x in client_ids_filter)
    exec_cur = _exec_accumulate_rows(filtered_rows, _exec_gid, _exec_mem, filtered_financial_rows)
    exec_prev = _exec_accumulate_rows(
        _apply_client_filters(previous_rows_pool),
        _exec_gid,
        _exec_mem,
        previous_financial_rows,
    )

    ranking_rows = []
    for item in client_agg.values():
        previous_item = previous_client_agg.get(item["client_key"], {})
        weekly_peak = max(item["weeks"].values()) if item["weeks"] else 0
        weekly_avg = round(statistics.mean(item["weeks"].values()), 2) if item["weeks"] else 0.0
        avg_duration_m = round(item["total_duration_m"] / item["duration_count"], 1) if item["duration_count"] else 0.0
        financial_base = max(
            float(item["delivered_value"] or 0.0) + float(item["returned_value"] or 0.0),
            0.01,
        )
        qty_den = item["visits"] if item["visits"] > 0 else item["returned_occurrences"]
        return_rate_qtd = round(_safe_pct(item["returned_occurrences"], qty_den), 2)
        return_rate_value = round(_safe_pct(item["returned_value"], financial_base if financial_base > 0 else item["returned_value"]), 2)

        prev_financial_base = max(
            float(previous_item.get("delivered_value", 0.0) or 0.0)
            + float(previous_item.get("returned_value", 0.0) or 0.0),
            0.01,
        )
        prev_visits = int(previous_item.get("visits", 0) or 0)
        prev_returns = int(previous_item.get("returned_occurrences", 0) or 0)
        prev_returned_value = round(float(previous_item.get("returned_value", 0.0) or 0.0), 2)
        has_previous_data = bool(prev_visits or prev_returns or prev_financial_base > 0 or prev_returned_value > 0)
        prev_qty_den = prev_visits if prev_visits > 0 else prev_returns
        prev_return_rate_qtd = round(_safe_pct(prev_returns, prev_qty_den), 2) if has_previous_data else 0.0
        prev_return_rate_value = round(_safe_pct(prev_returned_value, prev_financial_base if prev_financial_base > 0 else prev_returned_value), 2) if has_previous_data else 0.0
        delta_return_rate_value = round(return_rate_value - prev_return_rate_value, 2)
        delta_return_rate_qtd = round(return_rate_qtd - prev_return_rate_qtd, 2)
        delta_returned_value = round(float(item["returned_value"] or 0.0) - prev_returned_value, 2)
        prev_delivered_value = (
            round(float(previous_item.get("delivered_value", 0.0) or 0.0), 2) if has_previous_data else 0.0
        )
        delta_delivered_value = round(float(item.get("delivered_value") or 0.0) - prev_delivered_value, 2)

        top_driver_name = "-"
        top_driver_visits = 0
        top_driver_return_name = "-"
        top_driver_return_value = 0.0
        top_driver_return_share = 0.0
        if item["drivers"]:
            top_driver_name, top_driver_data = max(
                item["drivers"].items(),
                key=lambda kv: (kv[1].get("visits", 0), kv[1].get("returned_value", 0.0), kv[1].get("duration_m", 0.0)),
            )
            top_driver_visits = int(top_driver_data.get("visits", 0) or 0)
            if float(item["returned_value"] or 0.0) > 0:
                top_driver_return_name, top_driver_return_data = max(
                    item["drivers"].items(),
                    key=lambda kv: (kv[1].get("returned_value", 0.0), kv[1].get("returns", 0), kv[1].get("visits", 0)),
                )
                top_driver_return_value = round(float(top_driver_return_data.get("returned_value", 0.0) or 0.0), 2)
                top_driver_return_share = round(_safe_pct(top_driver_return_value, item["returned_value"]), 2)
            else:
                top_driver_return_name = top_driver_name
        top_motivo_name = "-"
        top_motivo_count = 0
        if item["motivos"]:
            top_motivo_name, top_motivo_data = max(
                item["motivos"].items(),
                key=lambda kv: (kv[1].get("value", 0.0), kv[1].get("count", 0)),
            )
            top_motivo_count = int(top_motivo_data.get("count", 0) or 0)
        top_resp_name = "-"
        top_resp_return_value = 0.0
        top_resp_return_share = 0.0
        if item["responsabilidades"]:
            top_resp_name, top_resp_data = max(
                item["responsabilidades"].items(),
                key=lambda kv: (kv[1].get("value", 0.0), kv[1].get("count", 0)),
            )
            if float(item["returned_value"] or 0.0) > 0:
                top_resp_return_value = round(float(top_resp_data.get("value", 0.0) or 0.0), 2)
                top_resp_return_share = round(_safe_pct(top_resp_return_value, item["returned_value"]), 2)

        window_checks = int(item.get("window_checks", 0) or 0)
        window_hits = int(item.get("window_hits", 0) or 0)
        window_misses = int(item.get("window_misses", 0) or 0)
        window_adherence_pct = round(_safe_pct(window_hits, window_checks), 2) if window_checks > 0 else None

        risk_score = min(
            100,
            int(
                round(
                    (return_rate_value * 10.5)
                    + (return_rate_qtd * 1.6)
                    + min(18.0, avg_duration_m / 6.0)
                    + min(16.0, item["reopen_count"] * 4.0)
                    + (10.0 if item["returned_occurrences"] >= 2 else 0.0)
                    + (8.0 if weekly_peak >= 3 else 0.0)
                )
            ),
        )
        risk_label, risk_tone = _risk_label(risk_score)

        ck = item["client_key"]
        unproductive_m = round(float(exec_cur["unprod_by_key"].get(ck, 0.0) or 0.0), 1)
        dom_macro = "—"
        dom_macro_share = 0.0
        rv = float(item.get("returned_value") or 0)
        if rv > 0 and item.get("responsabilidades") and top_resp_name not in ("-", "—", ""):
            dom_macro = top_resp_name
            dom_macro_share = top_resp_return_share
        else:
            macro_vals: dict[str, float] = {}
            for mot, data in (item.get("motivos") or {}).items():
                mac = _macro_loss_label(mot, "")
                macro_vals[mac] = macro_vals.get(mac, 0.0) + float(data.get("value") or 0)
            if macro_vals:
                dom_macro, mv = max(macro_vals.items(), key=lambda x: x[1])
                tv = sum(macro_vals.values())
                dom_macro_share = round(_safe_pct(mv, tv), 1) if tv > 0 else 0.0
            elif rv > 0:
                dom_macro = _macro_loss_label(top_motivo_name, top_resp_name)
                dom_macro_share = 100.0
        dv_client = float(item.get("delivered_value") or 0)
        tdur = float(item.get("total_duration_m") or 0)
        min_per_1000 = round(tdur / max(dv_client / 1000.0, 0.01), 1) if dv_client > 0 else 0.0
        est_op_cost = round((tdur / 60.0) * _BI_EXEC_HOURLY_COST, 2)
        est_balance = round(dv_client - est_op_cost - float(item.get("returned_value") or 0), 2)
        logistic_weight = round(tdur / max(dv_client, 1.0), 4) if dv_client > 0 else 0.0
        return_per_hour = round(float(item.get("returned_value") or 0) / max(tdur / 60.0, 0.01), 2)
        ur = unproductive_m / max(tdur, 0.01)
        wear_score = min(
            100,
            int(
                round(
                    return_rate_value * 10.0
                    + min(28.0, ur * 42.0)
                    + min(20.0, avg_duration_m / 5.5)
                    + min(14.0, float(item.get("reopen_count") or 0) * 3.5)
                    + return_rate_qtd * 2.2
                    + (10.0 if item["returned_occurrences"] >= 2 else 0.0)
                )
            ),
        )
        if wear_score >= 75:
            wear_tier, wear_tone = ("Destrutivo", "danger")
        elif wear_score >= 52:
            wear_tier, wear_tone = ("Crítico", "danger")
        elif wear_score >= 30:
            wear_tier, wear_tone = ("Atenção", "warning")
        else:
            wear_tier, wear_tone = ("Saudável", "success")
        dm_l = _norm_text(dom_macro)
        if dom_macro == "Cadastro / planejamento" or dm_l == "cadastro / planejamento":
            suggested_action = "Auditoria de endereço, janela e acesso no cadastro."
        elif dom_macro == "Cliente / mercado" or dm_l == "cliente / mercado":
            suggested_action = "Renegociar janela e confirmação de recebimento (D-1)."
        elif dom_macro == "Financeiro / pagamento" or dm_l == "financeiro / pagamento":
            suggested_action = "Validar forma de pagamento e limite com financeiro/comercial."
        elif "logist" in dm_l:
            suggested_action = "Foco em conferência de carga, separação e execução de rota."
        elif "mercado" in dm_l and "/" not in str(dom_macro):
            suggested_action = "Verificar qualidade percebida e acordo comercial de troca."
        elif "comercial" in dm_l:
            suggested_action = "Alinhar pedido, preço e prazo com comercial antes da próxima expedição."
        else:
            suggested_action = "Revisão conjunta comercial + operação no ponto."

        return_pct_planned = _bi_client_return_pct_planned(
            float(item["planned_value"] or 0.0),
            float(item["delivered_value"] or 0.0),
            float(item["returned_value"] or 0.0),
        )
        delivery_efficiency_pct = round(
            _safe_pct(int(item["delivered_visits"] or 0), max(int(item["visits"] or 0), 1)),
            2,
        )
        treatable_val = bci_clientes.treatable_return_value(item.get("motivos"), float(item.get("returned_value") or 0.0))

        ranking_rows.append(
            {
                "client_id": item["client_id"],
                "client_name": item["client_name"],
                "city": item["city"],
                "bairro": item["bairro"],
                "segmento": item["segmento"],
                "prioridade": item["prioridade"],
                "status_operacional": item["status_operacional"],
                "status_cadastro": item["status_cadastro"],
                "address": item["address"],
                "window": item["window"],
                "visits": item["visits"],
                "delivered_visits": item["delivered_visits"],
                "open_visits": item["open_visits"],
                "returned_occurrences": item["returned_occurrences"],
                "weekly_peak_visits": weekly_peak,
                "weekly_avg_visits": weekly_avg,
                "planned_value": round(item["planned_value"], 2),
                "delivered_value": round(dv_client, 2),
                "returned_value": round(item["returned_value"], 2),
                "planned_kg": round(float(item.get("planned_kg") or 0.0), 2),
                "delivered_kg": round(float(item.get("delivered_kg") or 0.0), 2),
                "returned_kg": round(item["returned_kg"], 2),
                "max_duration_m": round(float(item.get("max_duration_m") or 0.0), 1),
                "min_duration_m": (round(float(item["min_duration_m"]), 1) if item.get("min_duration_m") is not None else None),
                "return_rate_qtd": return_rate_qtd,
                "return_rate_value": return_rate_value,
                "return_pct_planned": return_pct_planned,
                "delivery_efficiency_pct": delivery_efficiency_pct,
                "treatable_returned_value": treatable_val,
                "_treatable_motivos": bci_clientes.treatable_motivo_breakdown(item.get("motivos")),
                "total_duration_m": round(item["total_duration_m"], 1),
                "avg_duration_m": avg_duration_m,
                "reopen_count": item["reopen_count"],
                "top_driver_name": top_driver_name,
                "top_driver_visits": top_driver_visits,
                "top_driver_return_name": top_driver_return_name,
                "top_motivo_name": top_motivo_name,
                "top_motivo_count": top_motivo_count,
                "top_responsabilidade_name": top_resp_name,
                "latest_date": item["latest_date"],
                "risk_score": risk_score,
                "risk_label": risk_label,
                "risk_tone": risk_tone,
                "has_previous_data": has_previous_data,
                "previous_return_rate_qtd": prev_return_rate_qtd,
                "previous_return_rate_value": prev_return_rate_value,
                "previous_returned_value": prev_returned_value,
                "previous_delivered_value": prev_delivered_value,
                "delta_return_rate_qtd": delta_return_rate_qtd,
                "delta_return_rate_value": delta_return_rate_value,
                "delta_returned_value": delta_returned_value,
                "delta_delivered_value": delta_delivered_value,
                "window_checks": window_checks,
                "window_hits": window_hits,
                "window_misses": window_misses,
                "window_adherence_pct": window_adherence_pct,
                "top_driver_return_value": top_driver_return_value,
                "top_driver_return_share": top_driver_return_share,
                "top_responsabilidade_return_value": top_resp_return_value,
                "top_responsabilidade_return_share": top_resp_return_share,
                "unproductive_m": unproductive_m,
                "dominant_macro": dom_macro,
                "dominant_macro_share": dom_macro_share,
                "min_per_1000_brl": min_per_1000,
                "est_operational_cost": est_op_cost,
                "est_balance": est_balance,
                "logistic_weight": logistic_weight,
                "return_per_hour": return_per_hour,
                "wear_score": wear_score,
                "wear_tier": wear_tier,
                "wear_tone": wear_tone,
                "suggested_action": suggested_action,
            }
        )

    ranking_rows.sort(
        key=lambda row: (
            row.get("wear_score", 0),
            row.get("risk_score", 0),
            row.get("returned_value", 0.0),
            row.get("total_duration_m", 0.0),
        ),
        reverse=True,
    )

    rf = str(returns_filter or "Todos").strip().lower()
    if rf == "com_devolucao":
        ranking_rows = [row for row in ranking_rows if float(row.get("returned_value") or 0) > 0 or int(row.get("returned_occurrences") or 0) > 0]
    elif rf == "acima_meta":
        ranking_rows = [row for row in ranking_rows if float(row.get("return_pct_planned") or row.get("return_rate_value") or 0) >= 2.0]

    dvals = [float(r.get("delivered_value") or 0) for r in ranking_rows]
    sd = sorted(dvals)
    median_delivered = float(statistics.median(sd)) if sd else 0.0
    p75_delivered = float(sd[max(0, int(len(sd) * 0.75) - 1)]) if sd else 0.0
    adurs = [float(r.get("avg_duration_m") or 0) for r in ranking_rows if int(r.get("visits") or 0) > 0]
    avg_duration_global = float(statistics.mean(adurs)) if adurs else 0.0

    pos_cli_ids = sorted(
        {int(r["client_id"]) for r in ranking_rows if r.get("client_id") is not None and int(r["client_id"]) > 0}
    )
    cli_by_id: dict[int, models.Client] = {}
    if pos_cli_ids:
        for cobj in session.exec(select(models.Client).where(models.Client.id.in_(pos_cli_ids))).all():
            if cobj.id is not None:
                cli_by_id[int(cobj.id)] = cobj
    seller_ids = sorted(
        {
            int(getattr(cli_by_id[i], "vendedor_id"))
            for i in pos_cli_ids
            if i in cli_by_id and getattr(cli_by_id[i], "vendedor_id", None) is not None
        }
    )
    seller_name_by_id: dict[int, str] = {}
    if seller_ids:
        for em in session.exec(select(models.Employee).where(models.Employee.id.in_(seller_ids))).all():
            if em.id is not None:
                seller_name_by_id[int(em.id)] = str(em.name or f"Vendedor #{em.id}")

    for r in ranking_rows:
        cid = r.get("client_id")
        cobj = cli_by_id.get(int(cid)) if cid is not None and str(cid).lstrip("-").isdigit() and int(cid) > 0 else None
        nb = (str(cobj.nb).strip() if cobj and cobj.nb else "") or ""
        vid = int(cobj.vendedor_id) if cobj and cobj.vendedor_id is not None else None
        r["client_code"] = nb or "—"
        r["vendedor_id"] = vid
        r["vendedor_name"] = seller_name_by_id.get(vid, "Sem vendedor") if vid is not None else "Sem vendedor"
        r["_search_blob"] = _norm_text(
            " ".join(
                str(x or "")
                for x in (
                    r.get("client_name"),
                    nb,
                    r.get("vendedor_name"),
                    r.get("top_motivo_name"),
                    r.get("top_responsabilidade_name"),
                )
            )
        )
        cls = bci_clientes.classificacao_cliente(
            delivered_value=float(r.get("delivered_value") or 0),
            planned_value=float(r.get("planned_value") or 0),
            returned_value=float(r.get("returned_value") or 0),
            return_pct_planned=float(r.get("return_pct_planned") or 0),
            avg_duration_m=float(r.get("avg_duration_m") or 0),
            avg_duration_global=avg_duration_global,
            reopen_count=int(r.get("reopen_count") or 0),
            returned_occurrences=int(r.get("returned_occurrences") or 0),
            visits=int(r.get("visits") or 0),
            median_delivered=median_delivered,
            p75_delivered=p75_delivered,
            top_motivo_name=str(r.get("top_motivo_name") or ""),
            top_resp_name=str(r.get("top_responsabilidade_name") or ""),
        )
        r["classification_code"] = cls[0]
        r["classification_title"] = cls[1]
        r["classification_message"] = cls[2]
        sc = bci_clientes.score_cliente(
            delivered_value=float(r.get("delivered_value") or 0),
            return_pct_planned=float(r.get("return_pct_planned") or 0),
            avg_duration_m=float(r.get("avg_duration_m") or 0),
            avg_duration_global=avg_duration_global,
            returned_occurrences=int(r.get("returned_occurrences") or 0),
            visits=int(r.get("visits") or 0),
            reopen_count=int(r.get("reopen_count") or 0),
        )
        r["cliente_score"] = sc[0]
        r["cliente_score_band"] = sc[1]
        r["cliente_score_tone"] = sc[2]
        r["cliente_score_parts_json"] = json.dumps(sc[4], ensure_ascii=False)
        r["operational_impact"] = bci_clientes.impacto_operacional(
            float(r.get("return_pct_planned") or 0),
            float(r.get("return_rate_qtd") or 0),
            float(r.get("avg_duration_m") or 0),
            avg_duration_global,
            int(r.get("reopen_count") or 0),
        )
        r["action_recommendation"] = bci_clientes.acao_recomendada_por_classificacao(
            str(r.get("classification_code") or ""),
            str(r.get("top_motivo_name") or ""),
            str(r.get("top_responsabilidade_name") or ""),
        )

    if vendedor_id is not None:
        if int(vendedor_id) == -1:
            ranking_rows = [r for r in ranking_rows if r.get("vendedor_id") is None]
        else:
            ranking_rows = [r for r in ranking_rows if r.get("vendedor_id") == int(vendedor_id)]

    if (classification_filter or "").strip() and (classification_filter or "").strip().lower() != "todos":
        cf = (classification_filter or "").strip().upper()
        ranking_rows = [r for r in ranking_rows if str(r.get("classification_code") or "").upper() == cf]

    if (motivo_filter or "").strip() and (motivo_filter or "").strip().lower() != "todos":
        mf = _norm_text(motivo_filter)
        ranking_rows = [r for r in ranking_rows if mf and mf in _norm_text(r.get("top_motivo_name"))]

    if (responsabilidade_filter or "").strip() and (responsabilidade_filter or "").strip().lower() != "todos":
        rf_txt = _norm_text(responsabilidade_filter)
        ranking_rows = [r for r in ranking_rows if rf_txt and rf_txt in _norm_text(r.get("top_responsabilidade_name"))]

    if (search_q or "").strip():
        sq = _norm_text(search_q.strip())
        if sq:
            ranking_rows = [r for r in ranking_rows if sq in (r.get("_search_blob") or "")]

    def _band_delivered(row, band: str) -> bool:
        v = float(row.get("delivered_value") or 0)
        b = (band or "Todos").strip().lower()
        if b in ("", "todos"):
            return True
        if b == "ate_5k":
            return v <= 5000
        if b == "5k_20k":
            return 5000 < v <= 20000
        if b == "20k_50k":
            return 20000 < v <= 50000
        if b == "acima_50k":
            return v > 50000
        return True

    def _band_return_pct(row, band: str) -> bool:
        v = float(row.get("return_pct_planned") or 0)
        b = (band or "Todos").strip().lower()
        if b in ("", "todos"):
            return True
        if b == "zero":
            return v <= 0.01
        if b == "ate_2":
            return v <= 2.0
        if b == "2_5":
            return 2.0 < v <= 5.0
        if b == "acima_5":
            return v > 5.0
        return True

    def _band_duration(row, band: str) -> bool:
        v = float(row.get("avg_duration_m") or 0)
        b = (band or "Todos").strip().lower()
        if b in ("", "todos"):
            return True
        if b == "ate_30":
            return v <= 30
        if b == "30_60":
            return 30 < v <= 60
        if b == "60_90":
            return 60 < v <= 90
        if b == "acima_90":
            return v > 90
        return True

    ranking_rows = [r for r in ranking_rows if _band_delivered(r, purchase_band)]
    ranking_rows = [r for r in ranking_rows if _band_return_pct(r, return_band)]
    ranking_rows = [r for r in ranking_rows if _band_duration(r, duration_band)]

    ranking_rows.sort(key=lambda row: float(row.get("delivered_value") or 0), reverse=True)

    total_visits = sum(int(row.get("visits", 0) or 0) for row in ranking_rows)
    critical_clients = [
        row
        for row in ranking_rows
        if str(row.get("classification_code") or "") in ("CRITICO", "ALTO_VALOR_RISCO")
        or int(row.get("risk_score", 0) or 0) >= 70
        or float(row.get("return_pct_planned") or 0) >= 3.0
    ]
    clients_with_returns = [row for row in ranking_rows if (row.get("returned_occurrences", 0) or 0) > 0]
    top_time_row = max(ranking_rows, key=lambda row: row.get("total_duration_m", 0.0), default=None)
    top_freq_row = max(ranking_rows, key=lambda row: (row.get("weekly_peak_visits", 0), row.get("visits", 0)), default=None)
    top_return_row = max(ranking_rows, key=lambda row: row.get("returned_value", 0.0), default=None)
    top_pct_row = max(
        ranking_rows,
        key=lambda row: float(row.get("return_pct_planned") or 0) or float(row.get("return_rate_value") or 0),
        default=None,
    )
    top_recurrence_row = max(ranking_rows, key=lambda row: row.get("returned_occurrences", 0), default=None)
    top_risk_row = max(ranking_rows, key=lambda row: int(row.get("risk_score") or 0), default=None)
    worsening_candidates = [row for row in ranking_rows if row.get("has_previous_data")]
    top_worsening_row = max(
        worsening_candidates,
        key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("delta_returned_value", 0.0)),
        default=None,
    )
    window_rows = [row for row in ranking_rows if (row.get("window_checks", 0) or 0) > 0]
    top_window_risk_row = min(
        window_rows,
        key=lambda row: (row.get("window_adherence_pct", 100.0), -int(row.get("window_misses", 0) or 0)),
        default=None,
    )
    driver_concentration_rows = [row for row in ranking_rows if (row.get("top_driver_return_share", 0.0) or 0.0) > 0]
    top_driver_concentration_row = max(
        driver_concentration_rows,
        key=lambda row: (row.get("top_driver_return_share", 0.0), row.get("returned_value", 0.0)),
        default=None,
    )
    resp_concentration_rows = [row for row in ranking_rows if (row.get("top_responsabilidade_return_share", 0.0) or 0.0) > 0]
    top_resp_concentration_row = max(
        resp_concentration_rows,
        key=lambda row: (row.get("top_responsabilidade_return_share", 0.0), row.get("returned_value", 0.0)),
        default=None,
    )

    city_agg: dict[str, dict] = {}
    for row in ranking_rows:
        city_label = str(row.get("city") or "Sem cidade").strip() or "Sem cidade"
        city_bucket = city_agg.setdefault(city_label, {"city": city_label, "clients": 0, "visits": 0, "returned_value": 0.0})
        city_bucket["clients"] += 1
        city_bucket["visits"] += int(row.get("visits", 0) or 0)
        city_bucket["returned_value"] += float(row.get("returned_value", 0.0) or 0.0)
    city_rows = sorted(city_agg.values(), key=lambda row: (row["visits"], row["returned_value"]), reverse=True)
    top_city_row = city_rows[0] if city_rows else None
    top_city_share = round(_safe_pct(top_city_row["visits"], total_visits), 1) if top_city_row else 0.0

    ec, ep = exec_cur, exec_prev

    def _delta_pct_exec(a, b):
        try:
            a_f = float(a or 0)
            b_f = float(b or 0)
            if not math.isfinite(a_f) or not math.isfinite(b_f) or b_f <= 0:
                return None
            return round((a_f - b_f) / b_f * 100.0, 1)
        except Exception:
            return None

    clients_over_60 = sum(1 for r in ranking_rows if float(r.get("avg_duration_m") or 0) > 60)
    clients_over_90 = sum(1 for r in ranking_rows if float(r.get("avg_duration_m") or 0) > 90)

    macro_v = ec["macro_value_global"]
    tot_delivered_grid = sum(float(r.get("delivered_value") or 0) for r in ranking_rows) or 1.0
    top10_delivered_block = sorted(ranking_rows, key=lambda r: float(r.get("delivered_value") or 0), reverse=True)[:10]
    top10_share_delivered = round(
        _safe_pct(sum(float(x.get("delivered_value") or 0) for x in top10_delivered_block), tot_delivered_grid), 1
    )
    n_above_meta = sum(
        1 for r in ranking_rows if float(r.get("return_pct_planned") or 0) > bci_clientes.META_DEVOLUCAO_VALOR_PCT
    )
    dvals_f = [float(r.get("delivered_value") or 0) for r in ranking_rows]
    med_f = float(statistics.median(sorted(dvals_f))) if dvals_f else 0.0
    ad_f = [float(r.get("avg_duration_m") or 0) for r in ranking_rows if int(r.get("visits") or 0) > 0]
    avg_dur_f = float(statistics.mean(ad_f)) if ad_f else avg_duration_global
    small_high_pool_full = [
        r
        for r in ranking_rows
        if float(r.get("delivered_value") or 0) < med_f
        and (
            float(r.get("returned_value") or 0) > 0
            or float(r.get("avg_duration_m") or 0) > avg_dur_f
            or int(r.get("reopen_count") or 0) > 0
        )
    ]
    n_small_high = len(small_high_pool_full)
    treatable_total = round(sum(float(r.get("treatable_returned_value") or 0) for r in ranking_rows), 2)

    treatable_drilldown: list[dict[str, Any]] = []
    for _r in ranking_rows:
        tm = _r.pop("_treatable_motivos", None) or []
        tv = float(_r.get("treatable_returned_value") or 0)
        if tv <= 0:
            continue
        hints = bci_clientes.suggest_treatable_resolutions(
            treatable_motivos=tm,
            top_motivo_name=str(_r.get("top_motivo_name") or ""),
            top_responsabilidade_name=str(_r.get("top_responsabilidade_name") or ""),
            classification_code=str(_r.get("classification_code") or ""),
            action_recommendation=str(_r.get("action_recommendation") or ""),
        )
        treatable_drilldown.append(
            {
                "client_id": _r.get("client_id"),
                "client_name": _r.get("client_name"),
                "client_code": _r.get("client_code"),
                "vendedor_name": _r.get("vendedor_name"),
                "treatable_returned_value": tv,
                "returned_value": float(_r.get("returned_value") or 0),
                "delivered_value": float(_r.get("delivered_value") or 0),
                "top_motivo_name": _r.get("top_motivo_name"),
                "top_responsabilidade_name": _r.get("top_responsabilidade_name"),
                "classification_title": _r.get("classification_title"),
                "treatable_motivos": tm,
                "hints": hints,
            }
        )
    treatable_drilldown.sort(key=lambda x: -float(x["treatable_returned_value"]))
    treatable_drilldown = treatable_drilldown[:120]
    treatable_drilldown_json = _json_for_inline_script(treatable_drilldown)

    main_motivo_period = "—"
    if ranking_rows:
        from collections import Counter

        mc = Counter()
        for r in ranking_rows:
            m = str(r.get("top_motivo_name") or "").strip()
            if m and m not in ("-", "—"):
                mc[m] += int(r.get("returned_occurrences") or 0) + 1
        if mc:
            main_motivo_period = mc.most_common(1)[0][0]
    main_resp_period = "—"
    main_resp_period_val = 0.0
    if macro_v:
        mr, mv = max(macro_v.items(), key=lambda kv: kv[1])
        main_resp_period, main_resp_period_val = str(mr), round(float(mv), 2)

    clients_good_intel = sum(
        1
        for r in ranking_rows
        if str(r.get("classification_code") or "") in ("PREMIUM_OPERACIONAL", "ESTAVEL")
        and int(r.get("cliente_score") or 0) >= bci_clientes.BEST_CLIENT_MIN_SCORE
    )

    # KPIs do topo: mesmo universo da tabela (ranking_rows após todos os filtros, ex.: busca por NB).
    kpi_planned = round(sum(float(r.get("planned_value") or 0) for r in ranking_rows), 2)
    kpi_delivered = round(sum(float(r.get("delivered_value") or 0) for r in ranking_rows), 2)
    kpi_returned = round(sum(float(r.get("returned_value") or 0) for r in ranking_rows), 2)
    kpi_fin_b = max(kpi_delivered + kpi_returned, 0.01)
    kpi_pct_dev = round(_safe_pct(kpi_returned, kpi_fin_b), 2)
    kpi_return_pct_planned_global = _bi_client_return_pct_planned(kpi_planned, kpi_delivered, kpi_returned)
    kpi_visits = sum(int(r.get("visits") or 0) for r in ranking_rows)
    kpi_duration_total = round(sum(float(r.get("total_duration_m") or 0) for r in ranking_rows), 1)
    kpi_unproductive = round(sum(float(r.get("unproductive_m") or 0) for r in ranking_rows), 1)
    kpi_productive = round(max(0.0, kpi_duration_total - kpi_unproductive), 1)
    kpi_waste_pct = round(_safe_pct(kpi_unproductive, kpi_duration_total), 2) if kpi_duration_total > 0 else 0.0

    prev_planned = prev_delivered = prev_returned = 0.0
    prev_duration_total = 0.0
    prev_unproductive_sum = 0.0
    for r in ranking_rows:
        ck = _client_row_key(r.get("client_id"), str(r.get("client_name") or ""))
        piv = previous_client_agg.get(ck, {})
        prev_planned += float(piv.get("planned_value") or 0.0)
        prev_delivered += float(piv.get("delivered_value") or 0.0)
        prev_returned += float(piv.get("returned_value") or 0.0)
        prev_duration_total += float(piv.get("total_duration_m") or 0.0)
        prev_unproductive_sum += float(ep["unprod_by_key"].get(ck, 0.0) or 0.0)
    prev_planned = round(prev_planned, 2)
    prev_delivered = round(prev_delivered, 2)
    prev_returned = round(prev_returned, 2)
    prev_duration_total = round(prev_duration_total, 1)
    prev_unproductive_sum = round(prev_unproductive_sum, 1)
    prev_fin_b = max(prev_delivered + prev_returned, 0.01)
    prev_pct_dev = round(_safe_pct(prev_returned, prev_fin_b), 2)
    prev_waste_pct = (
        round(_safe_pct(prev_unproductive_sum, prev_duration_total), 2) if prev_duration_total > 0 else 0.0
    )

    executive_kpis = {
        "delivered_value": kpi_delivered,
        "returned_value": kpi_returned,
        "return_pct_value": kpi_pct_dev,
        "total_duration_min": kpi_duration_total,
        "unproductive_min": kpi_unproductive,
        "productive_min": kpi_productive,
        "waste_pct": kpi_waste_pct,
        "monitored_clients": len(ranking_rows),
        "deliveries_count": kpi_visits,
        "clients_with_returns": len(clients_with_returns),
        "clients_avg_over_60": clients_over_60,
        "clients_avg_over_90": clients_over_90,
        "delta_delivered_pct": _delta_pct_exec(kpi_delivered, prev_delivered),
        "delta_return_pp": round(kpi_pct_dev - prev_pct_dev, 2),
        "delta_duration_pct": _delta_pct_exec(kpi_duration_total, prev_duration_total),
        "delta_unproductive_pct": _delta_pct_exec(kpi_unproductive, prev_unproductive_sum),
        "delta_waste_pp": round(kpi_waste_pct - prev_waste_pct, 2),
        "period_current": current_label,
        "period_previous": previous_label,
        "planned_value_total": kpi_planned,
        "return_pct_planned_global": kpi_return_pct_planned_global,
        "uniq_clients": len(ranking_rows),
        "clients_critical_intel": len(critical_clients),
        "clients_good_intel": clients_good_intel,
        "treatable_returned_total": treatable_total,
        "avg_duration_clients_m": round(avg_dur_f, 1),
        "return_pct_stops_global": round(
            _safe_pct(
                sum(int(r.get("returned_occurrences") or 0) for r in ranking_rows),
                max(sum(int(r.get("visits") or 0) for r in ranking_rows), 1),
            ),
            2,
        ),
        "return_pct_kpi_warn_points": round(float(bci_clientes.KPI_RETURN_RATE_WARNING) * 100.0, 6),
        "return_pct_kpi_danger_points": round(float(bci_clientes.KPI_RETURN_RATE_DANGER) * 100.0, 6),
    }

    reading_cards = bci_clientes.build_operational_reading_cards(
        returned_total=float(kpi_returned),
        delivered_total=float(kpi_delivered),
        planned_total=float(kpi_planned),
        n_clients=len(ranking_rows),
        n_above_meta=n_above_meta,
        n_small_high_impact=n_small_high,
        top10_delivered_share_pct=top10_share_delivered,
        main_motivo=main_motivo_period,
        main_resp=main_resp_period,
        main_resp_value=main_resp_period_val,
        treatable_value=treatable_total,
    )
    reading_cards_json = _json_for_inline_script(reading_cards)

    top10_returned = sorted(ranking_rows, key=lambda r: float(r.get("returned_value") or 0), reverse=True)[:10]
    small_high_block = sorted(
        [r for r in small_high_pool_full if float(r.get("operational_impact") or 0) > 0],
        key=lambda r: float(r.get("operational_impact") or 0),
        reverse=True,
    )[:10]
    p75_delivered_thr = sorted(dvals_f)[max(0, int(len(dvals_f) * 0.75) - 1)] if dvals_f else 0.0
    min_grande_del = max(float(p75_delivered_thr or 0), float(bci_clientes.LARGE_RISK_MIN_DELIVERED_FLOOR_BRL))
    large_risk_pool = [
        r
        for r in ranking_rows
        if float(r.get("delivered_value") or 0) >= min_grande_del
        and float(r.get("return_pct_planned") or 0) > float(bci_clientes.LARGE_RISK_MIN_RETURN_RATE) * 100.0
    ]
    large_risk_block = sorted(
        large_risk_pool,
        key=lambda r: float(r.get("return_pct_planned") or 0),
        reverse=True,
    )[:10]
    best_ops = sorted(
        [
            r
            for r in ranking_rows
            if float(r.get("delivered_value") or 0) >= med_f
            and float(r.get("return_pct_planned") or 0) <= bci_clientes.META_DEVOLUCAO_VALOR_PCT
            and int(r.get("cliente_score") or 0) >= 75
        ],
        key=lambda r: float(r.get("delivered_value") or 0),
        reverse=True,
    )[:10]
    worst_time = sorted(ranking_rows, key=lambda r: float(r.get("avg_duration_m") or 0), reverse=True)[:10]
    worst_recurrence = sorted(ranking_rows, key=lambda r: int(r.get("returned_occurrences") or 0), reverse=True)[:10]
    motivos_por_cliente = sorted(
        ranking_rows,
        key=lambda r: int(r.get("returned_occurrences") or 0),
        reverse=True,
    )[:12]

    analytic_blocks = {
        "top10_valor_comprado": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "value": r.get("delivered_value")} for r in top10_delivered_block
        ],
        "top10_valor_devolvido": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "value": r.get("returned_value")} for r in top10_returned
        ],
        "pequenos_alto_impacto": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "impact": r.get("operational_impact"), "delivered": r.get("delivered_value")}
            for r in small_high_block
        ],
        "grandes_com_risco": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "pct": r.get("return_pct_planned"), "delivered": r.get("delivered_value")}
            for r in large_risk_block
        ],
        "melhores_operacionais": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "score": r.get("cliente_score"), "delivered": r.get("delivered_value")}
            for r in best_ops
        ],
        "piores_tempos": [{"id": r.get("client_id"), "name": r.get("client_name"), "avg_m": r.get("avg_duration_m")} for r in worst_time],
        "maior_recorrencia": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "n": r.get("returned_occurrences")} for r in worst_recurrence
        ],
        "motivos_cliente": [
            {"id": r.get("client_id"), "name": r.get("client_name"), "motivo": r.get("top_motivo_name"), "n": r.get("returned_occurrences")}
            for r in motivos_por_cliente
        ],
    }

    large_risk_drilldown: list[dict[str, Any]] = []
    for r in sorted(large_risk_pool, key=lambda x: float(x.get("return_pct_planned") or 0), reverse=True)[:80]:
        large_risk_drilldown.append(
            {
                "client_id": r.get("client_id"),
                "client_name": r.get("client_name"),
                "client_code": r.get("client_code"),
                "vendedor_name": r.get("vendedor_name"),
                "return_pct_planned": round(float(r.get("return_pct_planned") or 0), 2),
                "returned_value": round(float(r.get("returned_value") or 0), 2),
                "delivered_value": round(float(r.get("delivered_value") or 0), 2),
                "classification_title": r.get("classification_title"),
                "top_motivo_name": r.get("top_motivo_name"),
                "top_responsabilidade_name": r.get("top_responsabilidade_name"),
                "context": bci_clientes.large_risk_context_lines(r),
                "hints": bci_clientes.large_risk_solution_lines(r),
            }
        )
    large_risk_drilldown_json = _json_for_inline_script(large_risk_drilldown)

    critical_drilldown: list[dict[str, Any]] = []
    for r in sorted(
        critical_clients,
        key=lambda x: (int(x.get("risk_score") or 0), float(x.get("returned_value") or 0)),
        reverse=True,
    )[:200]:
        critical_drilldown.append(
            {
                "client_id": r.get("client_id"),
                "client_name": r.get("client_name"),
                "client_code": r.get("client_code"),
                "vendedor_name": r.get("vendedor_name"),
                "classification_title": r.get("classification_title"),
                "risk_score": int(r.get("risk_score") or 0),
                "risk_label": r.get("risk_label"),
                "return_pct_planned": round(float(r.get("return_pct_planned") or 0), 2),
                "returned_value": round(float(r.get("returned_value") or 0), 2),
                "top_motivo_name": r.get("top_motivo_name"),
                "has_history": bool(r.get("has_previous_data")),
                "delta_return_rate_value": round(float(r.get("delta_return_rate_value") or 0), 2) if r.get("has_previous_data") else None,
                "context": bci_clientes.critical_client_context_lines(r),
                "hints": bci_clientes.critical_client_solution_lines(r),
            }
        )
    critical_drilldown_json = _json_for_inline_script(critical_drilldown)

    good_rows = [
        r
        for r in ranking_rows
        if str(r.get("classification_code") or "") in ("PREMIUM_OPERACIONAL", "ESTAVEL")
        and int(r.get("cliente_score") or 0) >= bci_clientes.BEST_CLIENT_MIN_SCORE
    ]
    good_drilldown: list[dict[str, Any]] = []
    for r in sorted(good_rows, key=lambda x: float(x.get("delivered_value") or 0), reverse=True)[:300]:
        good_drilldown.append(
            {
                "client_id": r.get("client_id"),
                "client_name": r.get("client_name"),
                "client_code": r.get("client_code"),
                "vendedor_name": r.get("vendedor_name"),
                "cliente_score": int(r.get("cliente_score") or 0),
                "classification_title": r.get("classification_title"),
                "delivered_value": round(float(r.get("delivered_value") or 0), 2),
                "return_pct_planned": round(float(r.get("return_pct_planned") or 0), 2),
                "summary": bci_clientes.good_client_summary_line(r),
            }
        )
    good_clients_drilldown_json = _json_for_inline_script(good_drilldown)

    def _first_motivo_from_rows(rows: list[dict]) -> str:
        agg: dict[str, float] = {}
        for rr in rows:
            m = str(rr.get("top_motivo_name") or "").strip()
            if m and m not in ("-", "—"):
                agg[m] = agg.get(m, 0.0) + float(rr.get("returned_value") or 0)
        if not agg:
            return main_motivo_period
        return max(agg.items(), key=lambda x: x[1])[0]

    small_high_drilldown: list[dict[str, Any]] = []
    for r in sorted(small_high_pool_full, key=lambda x: float(x.get("operational_impact") or 0), reverse=True)[:200]:
        small_high_drilldown.append(
            {
                "client_id": r.get("client_id"),
                "client_name": r.get("client_name"),
                "client_code": r.get("client_code"),
                "vendedor_name": r.get("vendedor_name"),
                "delivered_value": round(float(r.get("delivered_value") or 0), 2),
                "returned_value": round(float(r.get("returned_value") or 0), 2),
                "return_pct_planned": round(float(r.get("return_pct_planned") or 0), 2),
                "avg_duration_m": round(float(r.get("avg_duration_m") or 0), 1),
                "classification_title": r.get("classification_title"),
                "top_motivo_name": r.get("top_motivo_name"),
                "operational_impact": round(float(r.get("operational_impact") or 0), 2),
                "context": bci_clientes.critical_client_context_lines(r)[:4],
                "hints": bci_clientes.critical_client_solution_lines(r)[:4],
            }
        )
    small_high_drilldown_json = _json_for_inline_script(small_high_drilldown)

    action_first_summary = {
        "criticos": {
            "count": len(critical_clients),
            "value": round(sum(float(r.get("returned_value") or 0) for r in critical_clients), 2),
            "motivo": _first_motivo_from_rows(critical_clients),
        },
        "alto_valor_risco": {
            "count": len(large_risk_pool),
            "value": round(sum(float(r.get("delivered_value") or 0) for r in large_risk_pool), 2),
            "motivo": _first_motivo_from_rows(large_risk_pool),
        },
        "pequeno_grande_impacto": {
            "count": len(small_high_pool_full),
            "value": round(sum(float(r.get("returned_value") or 0) for r in small_high_pool_full), 2),
            "motivo": _first_motivo_from_rows(small_high_pool_full),
        },
        "oportunidade": {
            "count": len(treatable_drilldown),
            "value": float(treatable_total),
            "motivo": "Motivos tratáveis",
        },
    }

    def _rank_tab_row(r: dict) -> dict:
        return {
            "client_id": r.get("client_id"),
            "client_name": r.get("client_name"),
            "client_code": r.get("client_code"),
            "vendedor_name": r.get("vendedor_name"),
            "delivered_value": round(float(r.get("delivered_value") or 0), 2),
            "returned_value": round(float(r.get("returned_value") or 0), 2),
            "return_pct_planned": round(float(r.get("return_pct_planned") or 0), 2),
            "avg_duration_m": round(float(r.get("avg_duration_m") or 0), 1),
            "max_duration_m": round(float(r.get("max_duration_m") or 0), 1),
            "classification_title": r.get("classification_title"),
            "cliente_score": int(r.get("cliente_score") or 0),
        }

    MIN_VOL_PCT = float(bci_clientes.MIN_VOLUME_ENTREGUE_PAR_RANKING_PCT_BRL)
    pct_rank_pool = [
        r
        for r in ranking_rows
        if float(r.get("delivered_value") or 0) >= MIN_VOL_PCT
        and float(r.get("delivered_value") or 0) + float(r.get("returned_value") or 0) > 0
    ]
    baixo_vol_pct_pool = [
        r
        for r in ranking_rows
        if float(r.get("delivered_value") or 0) < float(bci_clientes.LOW_VOLUME_DISTORTION_MAX_DELIVERED_BRL)
        and float(r.get("return_pct_planned") or 0) > float(bci_clientes.LOW_VOLUME_DISTORTION_MIN_RETURN_RATE) * 100.0
    ]
    tempo_rank_pool = [r for r in ranking_rows if int(r.get("visits") or 0) > 0]
    melhores_pool = [
        r
        for r in ranking_rows
        if float(r.get("delivered_value") or 0) >= med_f
        and float(r.get("return_pct_planned") or 0) <= float(bci_clientes.BEST_CLIENT_MAX_RETURN_RATE) * 100.0
        and int(r.get("cliente_score") or 0) >= bci_clientes.BEST_CLIENT_MIN_SCORE
        and (
            int(r.get("visits") or 0) == 0
            or float(r.get("avg_duration_m") or 0)
            <= max(avg_dur_f * float(bci_clientes.BEST_CLIENT_AVG_TIME_FACTOR), float(bci_clientes.BEST_CLIENT_MIN_TIME_FALLBACK_MINUTES))
        )
    ]
    client_ranking_tabs = {
        "maior_compra": [_rank_tab_row(x) for x in sorted(ranking_rows, key=lambda z: float(z.get("delivered_value") or 0), reverse=True)[:25]],
        "maior_devolucao": [_rank_tab_row(x) for x in sorted(ranking_rows, key=lambda z: float(z.get("returned_value") or 0), reverse=True)[:25]],
        "maior_pct": [
            _rank_tab_row(x)
            for x in sorted(pct_rank_pool, key=lambda z: float(z.get("return_pct_planned") or 0), reverse=True)[:25]
        ],
        "baixo_volume_pct": [
            _rank_tab_row(x)
            for x in sorted(baixo_vol_pct_pool, key=lambda z: float(z.get("return_pct_planned") or 0), reverse=True)[:25]
        ],
        "maior_tempo": [
            _rank_tab_row(x)
            for x in sorted(
                tempo_rank_pool,
                key=lambda z: (float(z.get("avg_duration_m") or 0), float(z.get("max_duration_m") or 0)),
                reverse=True,
            )[:25]
        ],
        "pequeno_alto_impacto": [
            _rank_tab_row(x)
            for x in sorted(small_high_pool_full, key=lambda z: float(z.get("operational_impact") or 0), reverse=True)[:25]
        ],
        "grandes_risco": [_rank_tab_row(x) for x in sorted(large_risk_pool, key=lambda z: float(z.get("return_pct_planned") or 0), reverse=True)[:25]],
        "melhores": [_rank_tab_row(x) for x in sorted(melhores_pool, key=lambda z: float(z.get("delivered_value") or 0), reverse=True)[:25]],
    }
    client_ranking_tabs_json = _json_for_inline_script(client_ranking_tabs)

    main_dom_label, main_dom_detail = bci_clientes.dominante_operacional_por_valor_devolvido(ranking_rows)
    decision_strip = bci_clientes.build_decision_strip_intel(
        pct_gl=float(executive_kpis["return_pct_planned_global"] or 0),
        meta_pct=float(bci_clientes.META_DEVOLUCAO_VALOR_PCT),
        treatable_total=float(executive_kpis["treatable_returned_total"] or 0),
        returned_total=float(executive_kpis["returned_value"] or 0),
        critical_count=len(critical_clients),
        n_clients=len(ranking_rows),
        main_motivo=main_motivo_period,
        main_responsibility=main_dom_label,
        main_responsibility_detail=main_dom_detail,
    )

    insight_cards_ui: list[dict[str, str]] = []
    for c in reading_cards[:9]:
        _hints = c.get("hints") or []
        _ctx = c.get("context") or []
        insight_cards_ui.append(
            {
                "title": c.get("title") or "—",
                "headline": c.get("body") or "—",
                "interpretation": (_ctx[0] if _ctx else "—"),
                "action": (_hints[0] if _hints else "Manter acompanhamento na rotina semanal."),
            }
        )

    executive_headlines = []
    tmacro = sum(macro_v.values()) or 1.0
    logistica_valor_macro = sum(v for k, v in macro_v.items() if "logist" in _norm_text(k))
    if tmacro > 200 and _safe_pct(logistica_valor_macro, tmacro) < 40:
        executive_headlines.append(
            "Parte relevante das perdas em valor não se concentra na área Logística — aprofundar demais áreas e causas."
        )
    dur_sorted = sorted((float(r.get("total_duration_m") or 0) for r in ranking_rows), reverse=True)
    top5_share = (sum(dur_sorted[:5]) / max(float(ec["duration_total"] or 1), 1.0)) if ranking_rows else 0.0
    if top5_share >= 0.30 and len(ranking_rows) >= 5:
        executive_headlines.append(
            "Grande parte do tempo operacional concentra-se em poucos clientes — priorizar planos de ação direcionados."
        )
    vr = max(1, ec["visits_rota"])
    slow_share = _safe_pct(ec["bucket_visits"][3] + ec["bucket_visits"][4], vr)
    if slow_share >= 12 and ec["visits_rota"] > 15:
        executive_headlines.append(
            "Visitas longas (61+ min) pesam no mix — revisar roteirização, janelas e cadastro de acesso."
        )
    destr_n = sum(1 for r in ranking_rows if r.get("wear_tier") == "Destrutivo")
    if destr_n >= 1:
        executive_headlines.append(
            f"{destr_n} cliente(s) em perfil Destrutivo (alto desgaste) exigem intervenção gerencial prioritária."
        )
    if not executive_headlines:
        executive_headlines.append("Recorte sem alertas executivos extremos; manter monitoramento semanal.")

    false_villains = []
    for r in ranking_rows:
        if float(r.get("returned_value") or 0) <= 0:
            continue
        dm = r.get("dominant_macro") or ""
        if dm in ("—",) or "logist" in _norm_text(dm):
            continue
        if float(r.get("return_rate_value") or 0) >= 0.8 or int(r.get("wear_score") or 0) >= 42:
            false_villains.append(
                {
                    "client_id": r.get("client_id"),
                    "client_name": r.get("client_name"),
                    "city": r.get("city"),
                    "returned_value": r.get("returned_value"),
                    "unproductive_m": r.get("unproductive_m"),
                    "dominant_macro": dm,
                    "driver": r.get("top_driver_return_name") or r.get("top_driver_name"),
                    "suggested_action": r.get("suggested_action"),
                }
            )
    false_villains = false_villains[:15]

    cb_map = exec_cur["city_bucket"]
    city_visit_tot = {c: sum(cb_map[c]) for c in cb_map}
    heatmap_cities = sorted(city_visit_tot.keys(), key=lambda x: city_visit_tot[x], reverse=True)[:14]
    if not heatmap_cities and city_rows:
        heatmap_cities = [r["city"] for r in city_rows[:14]]
    heatmap_bucket_matrix = [[int(cb_map.get(c, [0] * 5)[i]) for i in range(5)] for c in heatmap_cities]
    all_causes = set()
    for cco in exec_cur["city_cause"].values():
        all_causes.update(cco.keys())
    cause_order = sorted(all_causes)[:12] if all_causes else ["Devolução", "Outros"]
    heatmap_cause_matrix = [[int(exec_cur["city_cause"].get(c, {}).get(cause, 0)) for cause in cause_order] for c in heatmap_cities]

    city_roll: dict[str, dict] = {}
    for r in ranking_rows:
        ct = str(r.get("city") or "Sem cidade").strip() or "Sem cidade"
        b = city_roll.setdefault(ct, {"td": 0.0, "v": 0, "unp": 0.0, "del": 0.0, "ret": 0.0, "cli": 0, "m60": 0})
        b["td"] += float(r.get("total_duration_m") or 0)
        b["v"] += int(r.get("visits") or 0)
        b["unp"] += float(r.get("unproductive_m") or 0)
        b["del"] += float(r.get("delivered_value") or 0)
        b["ret"] += float(r.get("returned_value") or 0)
        b["cli"] += 1
        if float(r.get("avg_duration_m") or 0) > 60:
            b["m60"] += 1
    macro_by_city: dict[str, dict[str, int]] = {}
    for r in ranking_rows:
        ct = str(r.get("city") or "Sem cidade")
        dm = r.get("dominant_macro") or ""
        if dm and dm != "—":
            macro_by_city.setdefault(ct, {})
            macro_by_city[ct][dm] = macro_by_city[ct].get(dm, 0) + 1

    def _city_max(getter):
        if not city_roll:
            return "—", 0.0
        best = max(city_roll.items(), key=lambda kv: getter(kv[1]))
        return best[0], getter(best[1])

    ct_time, _ = _city_max(lambda x: x["td"] / max(x["v"], 1))
    ct_unp, _ = _city_max(lambda x: x["unp"])
    ct_m60, v_m60 = _city_max(lambda x: x["m60"] / max(x["cli"], 1) * 100.0)
    ct_dev, _ = _city_max(lambda x: _safe_pct(x["ret"], x["del"] + x["ret"]) if (x["del"] + x["ret"]) > 0 else 0)
    ct_com = "—"
    if macro_by_city:

        def _comercial_mercado_client_count(macros: dict[str, int]) -> int:
            tot = 0
            for k, c in macros.items():
                nk = _norm_text(k)
                if nk in ("comercial", "cliente / mercado", "mercado"):
                    tot += c
                elif "comercial" in nk and "cadastro" not in nk and "/" not in str(k):
                    tot += c
            return tot

        ct_com = max(macro_by_city.items(), key=lambda kv: _comercial_mercado_client_count(kv[1]))[0]

    heatmap_city_kpis = [
        {"label": "Maior tempo médio (min/visita)", "city": ct_time, "hint": "Cidade com maior tempo total / visitas"},
        {"label": "Maior tempo improdutivo acumulado", "city": ct_unp},
        {"label": "Maior % clientes com média >60 min", "city": ct_m60, "value": round(v_m60, 1)},
        {"label": "Maior % devolução s/ valor (agreg.)", "city": ct_dev},
        {"label": "Mais clientes (área Comercial / Mercado)", "city": ct_com},
    ]

    anomaly_flags = []
    if top_risk_row and top_risk_row["risk_score"] >= 70:
        anomaly_flags.append(
            f"Cliente mais crítico: {top_risk_row['client_name']} com score {top_risk_row['risk_score']} e devolução financeira em {_fmt_br_1(top_risk_row['return_rate_value'])}%."
        )
    if top_return_row and top_return_row["returned_value"] > 0:
        anomaly_flags.append(
            f"Maior impacto financeiro: {top_return_row['client_name']} acumulou { _fmt_br_moeda(top_return_row['returned_value']) } em devoluções."
        )
    if top_time_row and top_time_row["total_duration_m"] >= 180:
        anomaly_flags.append(
            f"Tempo operacional concentrado em {top_time_row['client_name']}: { _fmt_br_duracao(top_time_row['total_duration_m']) } no período."
        )
    if top_recurrence_row and top_recurrence_row["returned_occurrences"] >= 2:
        anomaly_flags.append(
            f"Recorrência de devolução: {top_recurrence_row['client_name']} teve {top_recurrence_row['returned_occurrences']} ocorrência(s) no período."
        )
    if top_city_row and top_city_share >= 40:
        anomaly_flags.append(
            f"Concentração geográfica: {top_city_row['city']} representa {_fmt_br_1(top_city_share)}% das visitas filtradas."
        )
    if top_worsening_row and top_worsening_row.get("delta_return_rate_value", 0.0) >= 1.0:
        anomaly_flags.append(
            f"Piora vs período anterior: {top_worsening_row['client_name']} subiu {_fmt_br_1(top_worsening_row['delta_return_rate_value'])} p.p. em devolução por valor."
        )
    if top_window_risk_row and (top_window_risk_row.get("window_adherence_pct") or 0.0) < 85:
        anomaly_flags.append(
            f"Janela crítica: {top_window_risk_row['client_name']} aderiu a {_fmt_br_1(top_window_risk_row['window_adherence_pct'])}% da janela ({top_window_risk_row['window_misses']} fora da janela)."
        )
    if top_driver_concentration_row and top_driver_concentration_row.get("top_driver_return_share", 0.0) >= 75:
        anomaly_flags.append(
            f"Concentração em motorista: {top_driver_concentration_row['client_name']} depende de {top_driver_concentration_row['top_driver_return_name']} em {_fmt_br_1(top_driver_concentration_row['top_driver_return_share'])}% do valor devolvido."
        )
    if top_resp_concentration_row and top_resp_concentration_row.get("top_responsabilidade_return_share", 0.0) >= 75:
        anomaly_flags.append(
            f"Concentração em responsabilidade: {top_resp_concentration_row['client_name']} concentra {_fmt_br_1(top_resp_concentration_row['top_responsabilidade_return_share'])}% do valor em {top_resp_concentration_row['top_responsabilidade_name']}."
        )
    if not anomaly_flags:
        anomaly_flags.append("Sem concentração crítica de cliente no recorte selecionado.")

    recommendations = []
    if top_pct_row and top_pct_row["return_rate_value"] >= 2:
        recommendations.append(
            f"Abrir plano de ação com {top_pct_row['client_name']} e {top_pct_row['top_driver_name']} para reduzir a devolução em valor abaixo de 2,0%."
        )
    if top_time_row and top_time_row["avg_duration_m"] >= 90:
        recommendations.append(
            f"Revisar janela, acesso e sequência de atendimento de {top_time_row['client_name']} para reduzir o tempo médio de {_fmt_br_duracao(top_time_row['avg_duration_m'])}."
        )
    if top_freq_row and top_freq_row["weekly_peak_visits"] >= 3:
        recommendations.append(
            f"Mapear rotina semanal de {top_freq_row['client_name']}: pico de {top_freq_row['weekly_peak_visits']} visita(s) por semana."
        )
    if top_recurrence_row and top_recurrence_row["top_motivo_name"] != "-":
        recommendations.append(
            f"Atacar o motivo recorrente '{top_recurrence_row['top_motivo_name']}' em {top_recurrence_row['client_name']}."
        )
    if top_city_row and top_city_share >= 40:
        recommendations.append(
            f"Separar carteira por cidade em {top_city_row['city']} para reduzir concentração de rota e gargalo operacional."
        )
    if top_worsening_row and top_worsening_row.get("delta_return_rate_value", 0.0) >= 1.0:
        recommendations.append(
            f"Revisar o cliente {top_worsening_row['client_name']} contra o período anterior: foco em causa raiz, embalagem e conferência com {top_worsening_row['top_driver_name']}."
        )
    if top_window_risk_row and (top_window_risk_row.get("window_adherence_pct") or 0.0) < 90:
        recommendations.append(
            f"Negociar janela e sequência de atendimento de {top_window_risk_row['client_name']} para reduzir {top_window_risk_row['window_misses']} visita(s) fora da janela."
        )
    if top_driver_concentration_row and top_driver_concentration_row.get("top_driver_return_share", 0.0) >= 70:
        recommendations.append(
            f"Quebrar concentração de execução em {top_driver_concentration_row['client_name']}: hoje {top_driver_concentration_row['top_driver_return_name']} representa {_fmt_br_1(top_driver_concentration_row['top_driver_return_share'])}% do valor devolvido."
        )
    if top_resp_concentration_row and top_resp_concentration_row.get("top_responsabilidade_return_share", 0.0) >= 70:
        recommendations.append(
            f"Atacar a responsabilidade dominante em {top_resp_concentration_row['client_name']}: {top_resp_concentration_row['top_responsabilidade_name']} concentra {_fmt_br_1(top_resp_concentration_row['top_responsabilidade_return_share'])}%."
        )
    if not recommendations:
        recommendations.append("Operação de clientes sem alarme relevante no período; manter acompanhamento semanal.")

    managerial_actions = []
    for rec in recommendations[:10]:
        cat = "conjunta"
        rl = (rec or "").lower()
        if any(x in rl for x in ("comercial", "pedido", "preço", "preco", "prazo", "venda")):
            cat = "comercial"
        elif any(x in rl for x in ("carga", "expedição", "expedicao", "motorista", "rota", "separação", "separacao", "conferência")):
            cat = "logistica"
        elif any(x in rl for x in ("janela", "cadastro", "acesso", "cidade", "carteira")):
            cat = "cadastro"
        elif "financeiro" in rl or "pagamento" in rl:
            cat = "financeira"
        managerial_actions.append({"category": cat, "text": rec})

    detail_title = "Selecione um cliente no ranking para abrir o drill-through."
    detail_client = None
    detail_rows = []
    if detail_client_id:
        if int(detail_client_id) < 0:
            gid = -int(detail_client_id)
            grp = session.get(models.ClientGroup, gid)
            mem = session.exec(select(models.Client.id).where(models.Client.client_group_id == gid)).all()
            mem_set = {int(x) for x in mem if x is not None}
            detail_rows = [
                row for row in sorted(filtered_rows, key=lambda current: (current.get("date") or "", current.get("route_id") or 0), reverse=True)
                if row.get("client_id") in mem_set
            ]
            gnm = grp.name if grp else f"Grupo #{gid}"
            detail_client = {
                "client_id": detail_client_id,
                "client_name": f"Grupo: {gnm}",
                "city": "—",
                "bairro": "—",
                "segmento": "—",
                "prioridade": "—",
                "status_operacional": "—",
                "status_cadastro": "—",
                "address": "—",
                "window": "—",
            }
            detail_title = f"Drill-through · {detail_client['client_name']} ({len(mem_set)} lojas)"
        else:
            detail_client = next((row for row in ranking_rows if row.get("client_id") == detail_client_id), None)
            detail_rows = [
                row for row in sorted(filtered_rows, key=lambda current: (current.get("date") or "", current.get("route_id") or 0), reverse=True)
                if row.get("client_id") == detail_client_id
            ]
            if detail_client:
                detail_title = f"Drill-through do cliente {detail_client['client_name']}"

    macro_items = sorted(macro_v.items(), key=lambda x: -x[1]) if macro_v else [("Sem dados", 0.0)]
    vr_exec = max(1, ec["visits_rota"])
    scatter_pts = []
    sc_cand = [r for r in ranking_rows if int(r.get("visits") or 0) >= 1][:42]
    if len(sc_cand) >= 2:
        xs = [float(r.get("avg_duration_m") or 0) for r in sc_cand]
        ys = [float(r.get("delivered_value") or 0) for r in sc_cand]
        try:
            mx = float(statistics.median(xs))
            my = float(statistics.median(ys)) if ys else 0.0
        except statistics.StatisticsError:
            mx, my = 45.0, 5000.0
        for r in sc_cand:
            x, y = float(r.get("avg_duration_m") or 0), float(r.get("delivered_value") or 0)
            if y >= my and x <= mx:
                quad = "Eficiente"
            elif y >= my:
                quad = "Estratégico (pesado)"
            elif x >= mx:
                quad = "Destrutivo"
            else:
                quad = "Baixo impacto"
            scatter_pts.append(
                {
                    "x": round(x, 1),
                    "y": round(y, 2),
                    "r": int(min(22, max(5, int(r.get("visits") or 1) * 2))),
                    "name": (r.get("client_name") or "")[:28],
                    "cid": r.get("client_id"),
                    "retpct": round(float(r.get("return_pct_planned") or r.get("return_rate_value") or 0), 2),
                    "quad": quad,
                }
            )
    serve_rank = sorted(
        [r for r in ranking_rows if float(r.get("delivered_value") or 0) > 100],
        key=lambda r: float(r.get("min_per_1000_brl") or 0),
        reverse=True,
    )[:12]
    du_sorted = sorted(exec_cur["driver_unprod"].items(), key=lambda x: -x[1])[:10]
    cu_sorted = sorted(exec_cur["cause_unprod_time"].items(), key=lambda x: -x[1])[:8]

    motivo_val_agg: dict[str, float] = {}
    for _r in ranking_rows:
        _m = str(_r.get("top_motivo_name") or "").strip()
        if _m and _m not in ("-", "—"):
            motivo_val_agg[_m] = motivo_val_agg.get(_m, 0.0) + float(_r.get("returned_value") or 0)
    pareto_motivos_chart = sorted(
        ({"name": k, "value": round(v, 2)} for k, v in motivo_val_agg.items()),
        key=lambda x: -x["value"],
    )[:18]

    daily_evolution_map: dict[str, dict[str, float]] = {}
    for rr in filtered_rows:
        if str(rr.get("source") or "").upper() != "ROTA":
            continue
        ds = str(rr.get("date") or "").strip()[:10]
        if len(ds) < 10:
            continue
        b = daily_evolution_map.setdefault(ds, {"delivered": 0.0, "returned": 0.0})
        b["delivered"] += float(rr.get("delivered_value") or 0.0)
        if _norm_text(rr.get("status")) == "devolucao":
            b["returned"] += float(rr.get("returned_value") or 0.0)
    daily_evolution = sorted(
        (
            {"date": k, "delivered": round(v["delivered"], 2), "returned": round(v["returned"], 2)}
            for k, v in daily_evolution_map.items()
        ),
        key=lambda x: x["date"],
    )
    pareto_returns = sorted(
        (
            {
                "name": (r.get("client_name") or "")[:36],
                "value": round(float(r.get("returned_value") or 0), 2),
                "cid": r.get("client_id"),
            }
            for r in ranking_rows
            if float(r.get("returned_value") or 0) > 0
        ),
        key=lambda x: -x["value"],
    )[:20]
    matrix_impact = [
        {
            "x": round(float(r.get("delivered_value") or 0), 2),
            "y": round(float(r.get("return_pct_planned") or 0), 2),
            "r": max(1.0, round(float(r.get("returned_kg") or 0), 2)),
            "name": (r.get("client_name") or "")[:26],
            "cid": r.get("client_id"),
            "cls": r.get("classification_code"),
        }
        for r in sorted(ranking_rows, key=lambda z: float(z.get("returned_value") or 0), reverse=True)[:35]
    ]

    chart_payload = {
        "time_rank": {
            "labels": [row["client_name"] for row in sorted(ranking_rows, key=lambda row: row.get("total_duration_m", 0.0), reverse=True)[:12]],
            "minutes": [round(row["total_duration_m"], 1) for row in sorted(ranking_rows, key=lambda row: row.get("total_duration_m", 0.0), reverse=True)[:12]],
            "avg_minutes": [round(row["avg_duration_m"], 1) for row in sorted(ranking_rows, key=lambda row: row.get("total_duration_m", 0.0), reverse=True)[:12]],
        },
        "frequency_rank": {
            "labels": [row["client_name"] for row in sorted(ranking_rows, key=lambda row: (row.get("weekly_peak_visits", 0), row.get("visits", 0)), reverse=True)[:12]],
            "visits": [int(row["visits"]) for row in sorted(ranking_rows, key=lambda row: (row.get("weekly_peak_visits", 0), row.get("visits", 0)), reverse=True)[:12]],
            "weekly_peak": [int(row["weekly_peak_visits"]) for row in sorted(ranking_rows, key=lambda row: (row.get("weekly_peak_visits", 0), row.get("visits", 0)), reverse=True)[:12]],
        },
        "returns_rank": {
            "labels": [row["client_name"] for row in sorted(ranking_rows, key=lambda row: row.get("returned_value", 0.0), reverse=True)[:12]],
            "returned_value": [round(row["returned_value"], 2) for row in sorted(ranking_rows, key=lambda row: row.get("returned_value", 0.0), reverse=True)[:12]],
            "return_rate_value": [round(row["return_rate_value"], 2) for row in sorted(ranking_rows, key=lambda row: row.get("returned_value", 0.0), reverse=True)[:12]],
        },
        "city_concentration": {
            "labels": [row["city"] for row in city_rows[:10]],
            "visits": [int(row["visits"]) for row in city_rows[:10]],
            "clients": [int(row["clients"]) for row in city_rows[:10]],
            "returned_value": [round(row["returned_value"], 2) for row in city_rows[:10]],
        },
        "comparison_meta": {
            "current_label": current_label,
            "previous_label": previous_label,
        },
        "period_compare": {
            "labels": [row["client_name"] for row in sorted(worsening_candidates, key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "current_pct": [round(row["return_rate_value"], 2) for row in sorted(worsening_candidates, key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "previous_pct": [round(row["previous_return_rate_value"], 2) for row in sorted(worsening_candidates, key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "delta_pct": [round(row["delta_return_rate_value"], 2) for row in sorted(worsening_candidates, key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "current_value": [round(row["returned_value"], 2) for row in sorted(worsening_candidates, key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "previous_value": [round(row["previous_returned_value"], 2) for row in sorted(worsening_candidates, key=lambda row: (row.get("delta_return_rate_value", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
        },
        "driver_concentration": {
            "labels": [row["client_name"] for row in sorted(driver_concentration_rows, key=lambda row: (row.get("top_driver_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "share": [round(row["top_driver_return_share"], 2) for row in sorted(driver_concentration_rows, key=lambda row: (row.get("top_driver_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "driver": [row["top_driver_return_name"] for row in sorted(driver_concentration_rows, key=lambda row: (row.get("top_driver_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "returned_value": [round(row["returned_value"], 2) for row in sorted(driver_concentration_rows, key=lambda row: (row.get("top_driver_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
        },
        "responsibility_concentration": {
            "labels": [row["client_name"] for row in sorted(resp_concentration_rows, key=lambda row: (row.get("top_responsabilidade_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "share": [round(row["top_responsabilidade_return_share"], 2) for row in sorted(resp_concentration_rows, key=lambda row: (row.get("top_responsabilidade_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "responsabilidade": [row["top_responsabilidade_name"] for row in sorted(resp_concentration_rows, key=lambda row: (row.get("top_responsabilidade_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
            "returned_value": [round(row["returned_value"], 2) for row in sorted(resp_concentration_rows, key=lambda row: (row.get("top_responsabilidade_return_share", 0.0), row.get("returned_value", 0.0)), reverse=True)[:12]],
        },
        "window_adherence": {
            "labels": [row["client_name"] for row in sorted(window_rows, key=lambda row: (row.get("window_adherence_pct", 100.0), -(row.get("window_misses", 0) or 0)), reverse=False)[:12]],
            "adherence_pct": [round(row["window_adherence_pct"], 2) for row in sorted(window_rows, key=lambda row: (row.get("window_adherence_pct", 100.0), -(row.get("window_misses", 0) or 0)), reverse=False)[:12]],
            "hits": [int(row["window_hits"]) for row in sorted(window_rows, key=lambda row: (row.get("window_adherence_pct", 100.0), -(row.get("window_misses", 0) or 0)), reverse=False)[:12]],
            "misses": [int(row["window_misses"]) for row in sorted(window_rows, key=lambda row: (row.get("window_adherence_pct", 100.0), -(row.get("window_misses", 0) or 0)), reverse=False)[:12]],
            "checks": [int(row["window_checks"]) for row in sorted(window_rows, key=lambda row: (row.get("window_adherence_pct", 100.0), -(row.get("window_misses", 0) or 0)), reverse=False)[:12]],
        },
        "duration_buckets": {
            "labels": list(_DURATION_BUCKET_LABELS),
            "visits": list(ec["bucket_visits"]),
            "visit_pct": [round(_safe_pct(ec["bucket_visits"][i], vr_exec), 1) for i in range(5)],
            "delivered_value": list(ec["bucket_value"]),
            "duration_min": list(ec["bucket_duration"]),
            "returns": list(ec["bucket_returns"]),
            "unprod_min": list(ec["bucket_unprod"]),
        },
        "macro_loss": {
            "labels": [a[0] for a in macro_items],
            "values": [round(a[1], 2) for a in macro_items],
            "pct": [round(_safe_pct(a[1], tmacro), 1) for a in macro_items],
            "minutes": [round(exec_cur["macro_time_global"].get(a[0], 0.0), 1) for a in macro_items],
        },
        "productive_split": {
            "productive_min": ec["productive_total"],
            "unproductive_min": ec["unproductive_total"],
        },
        "scatter_clients": scatter_pts,
        "daily_delivered_vs_returned": daily_evolution,
        "pareto_returns_top": pareto_returns,
        "matrix_impact_x_compra": matrix_impact,
        "exec_compare_bars": {
            "labels": ["Valor entregue (R$)", "Valor devolvido (R$)", "Tempo total (h)", "Tempo improd. (h)"],
            "current": [
                round(ec["delivered_total"], 2),
                round(ec["returned_total"], 2),
                round(ec["duration_total"] / 60.0, 2),
                round(ec["unproductive_total"] / 60.0, 2),
            ],
            "previous": [
                round(ep["delivered_total"], 2),
                round(ep["returned_total"], 2),
                round(ep["duration_total"] / 60.0, 2),
                round(ep["unproductive_total"] / 60.0, 2),
            ],
        },
        "serve_cost_rank": {
            "labels": [r.get("client_name", "")[:22] for r in serve_rank],
            "min_per_1000": [float(r.get("min_per_1000_brl") or 0) for r in serve_rank],
            "est_balance": [float(r.get("est_balance") or 0) for r in serve_rank],
        },
        "driver_unproductive": {"labels": [a[0][:18] for a in du_sorted], "minutes": [round(a[1], 1) for a in du_sorted]},
        "cause_unproductive": {"labels": [a[0][:20] for a in cu_sorted], "minutes": [round(a[1], 1) for a in cu_sorted]},
        "heatmap_duration": {
            "cities": heatmap_cities,
            "buckets": list(_DURATION_BUCKET_LABELS),
            "matrix": heatmap_bucket_matrix,
        },
        "heatmap_cause": {
            "cities": heatmap_cities,
            "causes": cause_order,
            "matrix": heatmap_cause_matrix,
        },
        "wear_rank": {
            "labels": [r.get("client_name", "")[:20] for r in ranking_rows[:15]],
            "scores": [int(r.get("wear_score") or 0) for r in ranking_rows[:15]],
            "tiers": [r.get("wear_tier") or "" for r in ranking_rows[:15]],
        },
        "pareto_motivos": pareto_motivos_chart,
    }

    filters = {
        **delivery_dataset.get("filters", {}),
        "client_id": client_id,
        "city": city,
        "priority": priority,
        "client_status": client_status,
        "segmentos": selected_segmentos,
        "returns_filter": returns_filter,
        "detail_client_id": detail_client_id,
        "client_scope": scope if client_id else "solo",
        "group_filter_note": group_filter_note,
        "vendedor_id": vendedor_id,
        "search_q": (search_q or "").strip(),
        "classification_filter": classification_filter,
        "motivo_filter": motivo_filter,
        "responsabilidade_filter": responsabilidade_filter,
        "purchase_band": purchase_band,
        "return_band": return_band,
        "duration_band": duration_band,
    }
    _fq = {
        "date_from": filters.get("date_from") or "",
        "date_to": filters.get("date_to") or "",
        "shift": filters.get("shift") or "Todos",
        "driver_id": filters.get("driver_id") or "",
        "plate": filters.get("plate") or "Todos",
        "status": filters.get("status") or "Todos",
        "client_id": filters.get("client_id") or "",
        "city": filters.get("city") or "Todos",
        "priority": filters.get("priority") or "Todos",
        "client_status": filters.get("client_status") or "Todos",
        "returns_filter": filters.get("returns_filter") or "Todos",
        "client_scope": filters.get("client_scope") or "solo",
    }
    _fq["segmentos"] = filters.get("segmentos") or []
    if filters.get("detail_client_id"):
        _fq["detail_client_id"] = str(filters["detail_client_id"])
    if filters.get("vendedor_id") is not None:
        _fq["vendedor_id"] = str(filters["vendedor_id"])
    if filters.get("search_q"):
        _fq["search_q"] = str(filters["search_q"])
    if filters.get("classification_filter") and str(filters.get("classification_filter")).lower() != "todos":
        _fq["classification_filter"] = str(filters["classification_filter"])
    if filters.get("motivo_filter") and str(filters.get("motivo_filter")).lower() != "todos":
        _fq["motivo_filter"] = str(filters["motivo_filter"])
    if filters.get("responsabilidade_filter") and str(filters.get("responsabilidade_filter")).lower() != "todos":
        _fq["responsabilidade_filter"] = str(filters["responsabilidade_filter"])
    if filters.get("purchase_band") and str(filters.get("purchase_band")).lower() != "todos":
        _fq["purchase_band"] = str(filters["purchase_band"])
    if filters.get("return_band") and str(filters.get("return_band")).lower() != "todos":
        _fq["return_band"] = str(filters["return_band"])
    if filters.get("duration_band") and str(filters.get("duration_band")).lower() != "todos":
        _fq["duration_band"] = str(filters["duration_band"])
    filters_query = urlencode(_fq, doseq=True)

    kpis = {
        "monitored_clients": len(ranking_rows),
        "clients_with_returns": len(clients_with_returns),
        "critical_clients": len(critical_clients),
        "total_visits": total_visits,
        "top_time_client": top_time_row["client_name"] if top_time_row else "—",
        "top_time_minutes": round(top_time_row["total_duration_m"], 1) if top_time_row else 0.0,
        "top_freq_client": top_freq_row["client_name"] if top_freq_row else "—",
        "top_freq_peak": int(top_freq_row["weekly_peak_visits"]) if top_freq_row else 0,
        "top_return_client": top_return_row["client_name"] if top_return_row else "—",
        "top_return_value": round(top_return_row["returned_value"], 2) if top_return_row else 0.0,
        "top_pct_client": top_pct_row["client_name"] if top_pct_row else "—",
        "top_pct_value": round(
            float(
                (top_pct_row.get("return_pct_planned") if top_pct_row else None)
                or (top_pct_row.get("return_rate_value") if top_pct_row else 0)
                or 0
            ),
            2,
        )
        if top_pct_row
        else 0.0,
        "top_recurrence_client": top_recurrence_row["client_name"] if top_recurrence_row else "—",
        "top_recurrence_count": int(top_recurrence_row["returned_occurrences"]) if top_recurrence_row else 0,
        "top_city": top_city_row["city"] if top_city_row else "—",
        "top_city_share": top_city_share,
        "comparison_current_label": current_label,
        "comparison_previous_label": previous_label,
        "worsening_client": top_worsening_row["client_name"] if top_worsening_row else "—",
        "worsening_delta_pct": round(top_worsening_row["delta_return_rate_value"], 2) if top_worsening_row else 0.0,
        "window_client": top_window_risk_row["client_name"] if top_window_risk_row else "—",
        "window_adherence_pct": round(top_window_risk_row["window_adherence_pct"], 2) if top_window_risk_row and top_window_risk_row.get("window_adherence_pct") is not None else None,
        "driver_concentration_client": top_driver_concentration_row["client_name"] if top_driver_concentration_row else "—",
        "driver_concentration_pct": round(top_driver_concentration_row["top_driver_return_share"], 2) if top_driver_concentration_row else 0.0,
        "driver_concentration_driver": top_driver_concentration_row["top_driver_return_name"] if top_driver_concentration_row else "—",
        "responsibility_concentration_client": top_resp_concentration_row["client_name"] if top_resp_concentration_row else "—",
        "responsibility_concentration_pct": round(top_resp_concentration_row["top_responsabilidade_return_share"], 2) if top_resp_concentration_row else 0.0,
        "responsibility_concentration_name": top_resp_concentration_row["top_responsabilidade_name"] if top_resp_concentration_row else "—",
    }

    _intel_limit = 1200
    intel_clients_slim: list[dict] = []
    for r in ranking_rows[:_intel_limit]:
        intel_clients_slim.append({k: v for k, v in r.items() if not str(k).startswith("_")})
    intel_clients_json = _json_for_inline_script(intel_clients_slim)
    intel_clients_truncated = len(ranking_rows) > _intel_limit

    routes_intel_slim: list[dict] = []
    for rr in filtered_rows:
        if str(rr.get("source") or "").upper() != "ROTA":
            continue
        routes_intel_slim.append(
            {
                "client_id": rr.get("client_id"),
                "date": rr.get("date"),
                "order_number": rr.get("order_number"),
                "driver_name": rr.get("driver_name"),
                "plate": rr.get("plate"),
                "status": rr.get("status"),
                "planned_value": rr.get("planned_value"),
                "delivered_value": rr.get("delivered_value"),
                "returned_value": rr.get("returned_value"),
                "planned_kg": rr.get("planned_kg"),
                "delivered_kg": rr.get("delivered_kg"),
                "returned_kg": rr.get("returned_kg"),
                "duration_m": rr.get("duration_m"),
                "reopen_count": rr.get("reopen_count"),
                "motivo": rr.get("motivo"),
                "responsabilidade": rr.get("responsabilidade"),
            }
        )
        if len(routes_intel_slim) >= 650:
            break
    routes_intel_json = _json_for_inline_script(routes_intel_slim)

    primeira_acao_texto = bci_clientes.primeira_acao_prioridade_sp(
        treatable_total=float(executive_kpis["treatable_returned_total"] or 0),
        returned_total=float(executive_kpis["returned_value"] or 0),
        critical_count=len(critical_clients),
        large_risk_count=len(large_risk_pool),
        small_high_count=len(small_high_pool_full),
        n_clients=len(ranking_rows),
        fallback=(recommendations[0] if recommendations else "Manter acompanhamento diário da carteira."),
    )

    classification_filter_options = [
        {"id": "Todos", "label": "Todas"},
        {"id": "PREMIUM_OPERACIONAL", "label": "Cliente premium operacional"},
        {"id": "ALTO_VALOR_RISCO", "label": "Alto valor com risco"},
        {"id": "PEQUENO_ALTO_IMPACTO", "label": "Pequeno com alto impacto"},
        {"id": "CRITICO", "label": "Cliente crítico"},
        {"id": "ESTAVEL", "label": "Cliente estável"},
        {"id": "OBSERVACAO", "label": "Em observação"},
    ]

    return {
        "filters": filters,
        "filters_query": filters_query,
        "kpis": kpis,
        "ranking_rows": ranking_rows[:60],
        "ranking_total": len(ranking_rows),
        "detail_client": detail_client,
        "detail_rows": detail_rows[:200],
        "detail_total": len(detail_rows),
        "detail_title": detail_title,
        "anomaly_flags": anomaly_flags[:6],
        "recommendations": recommendations[:6],
        "drivers_filter": delivery_dataset.get("drivers_filter", []),
        "plates_filter": delivery_dataset.get("plates_filter", []),
        "statuses_filter": delivery_dataset.get("statuses_filter", []),
        "clients_filter": [{"id": client_key, "name": clients_filter_map[client_key]} for client_key in sorted(clients_filter_map, key=lambda current: clients_filter_map[current])],
        "cities_filter": sorted(cities_filter),
        "priorities_filter": sorted(priorities_filter),
        "client_statuses_filter": sorted(client_statuses_filter),
        "segmentos_filter": sorted(segmentos_filter),
        "chart_payload_json": _json_for_inline_script(chart_payload),
        "all_client_rows": ranking_rows,
        "executive_kpis": executive_kpis,
        "executive_headlines": executive_headlines,
        "false_villains": false_villains,
        "managerial_actions": managerial_actions,
        "heatmap_city_kpis": heatmap_city_kpis,
        "reading_cards": reading_cards,
        "reading_cards_json": reading_cards_json,
        "analytic_blocks": analytic_blocks,
        "sellers_filter": sellers_filter,
        "motivos_filter_options": motivos_filter_options,
        "responsabilidades_filter_options": responsabilidades_filter_options,
        "detail_rows_json": _json_for_inline_script(detail_rows[:200]),
        "intel_clients_json": intel_clients_json,
        "intel_clients_truncated": intel_clients_truncated,
        "classification_filter_options": classification_filter_options,
        "routes_intel_json": routes_intel_json,
        "treatable_drilldown_json": treatable_drilldown_json,
        "large_risk_drilldown_json": large_risk_drilldown_json,
        "critical_drilldown_json": critical_drilldown_json,
        "good_clients_drilldown_json": good_clients_drilldown_json,
        "decision_strip": decision_strip,
        "insight_cards_ui": insight_cards_ui,
        "action_first_summary": action_first_summary,
        "small_high_drilldown_json": small_high_drilldown_json,
        "client_ranking_tabs_json": client_ranking_tabs_json,
        "primeira_acao_texto": primeira_acao_texto,
    }


def _build_relatorio_avaliacao_motorista(
    session: Session,
    date_str: str,
    driver_ids: Optional[List[int]] = None,
) -> dict:
    """Monta dados do relatório de avaliação diária do motorista."""
    tz = ZoneInfo("America/Sao_Paulo")
    today = datetime.now(tz).date().strftime("%Y-%m-%d")

    def _parse_hhmm(v: Optional[str]) -> Optional[int]:
        if not v:
            return None
        try:
            h, m = str(v).strip().split(":")[:2]
            return int(h) * 60 + int(m)
        except (ValueError, IndexError):
            return None

    def _dur_m(start_v: Optional[str], end_v: Optional[str]) -> Optional[int]:
        s, e = _parse_hhmm(start_v), _parse_hhmm(end_v)
        if s is None or e is None:
            return None
        if e < s:
            e += 24 * 60
        return max(0, e - s)

    def _fmt_hora(v: Optional[str]) -> str:
        return (v or "--:--").strip() or "--:--"

    def _fmt_min(m: Optional[int]) -> str:
        return _fmt_br_duracao(m)

    def _valid_operation_times(values: list[Optional[str]]) -> list[int]:
        parsed = [_parse_hhmm(v) for v in values if v]
        valid = [v for v in parsed if v is not None and v > 0]
        return valid

    def _tipo_priority(tipo: Optional[str]) -> int:
        raw = (tipo or "").strip().lower()
        if "devol" in raw:
            return 0
        if "entrega" in raw:
            return 1
        return 2

    q = (
        select(models.Route)
        .where(models.Route.type == "delivery")
        .where(models.Route.date == date_str)
    )
    if driver_ids:
        q = q.where(models.Route.employee_id.in_(driver_ids))
    routes = session.exec(q.order_by(models.Route.start_time, models.Route.id)).all()

    emp_ids = list({r.employee_id for r in routes if r.employee_id})
    cli_ids = list({r.client_id for r in routes if r.client_id})
    emp_map = {
        e.id: e
        for e in (
            session.exec(select(models.Employee).where(models.Employee.id.in_(emp_ids))).all()
            if emp_ids
            else []
        )
    }
    cli_map = {
        c.id: c
        for c in (
            session.exec(select(models.Client).where(models.Client.id.in_(cli_ids))).all()
            if cli_ids
            else []
        )
    }
    vehicles = {v.placa.upper(): v for v in session.exec(select(models.Vehicle)).all()}

    by_driver: dict[int, dict] = {}
    for r in routes:
        drv_id = r.employee_id
        if drv_id not in by_driver:
            ds = session.exec(
                select(models.DeliverySession)
                .where(models.DeliverySession.employee_id == drv_id)
                .where(models.DeliverySession.date == date_str)
                .order_by(models.DeliverySession.id.desc())
            ).first()
            helpers = []
            if ds and ds.helpers_json:
                try:
                    hl = json.loads(ds.helpers_json) if isinstance(ds.helpers_json, str) else (ds.helpers_json or [])
                    seen = set()
                    for h in (hl if isinstance(hl, list) else []):
                        if h is None:
                            continue
                        if isinstance(h, (int, str)) and str(h).strip().isdigit():
                            he = emp_map.get(int(h))
                            if he and he.name:
                                key = he.name.strip().lower()
                                if key not in seen:
                                    seen.add(key)
                                    helpers.append(he.name)
                        elif isinstance(h, str) and (h or "").strip():
                            # Mobile envia nomes diretamente em helpers_json
                            name = (h or "").strip()
                            key = name.lower()
                            if key not in seen:
                                seen.add(key)
                                helpers.append(name)
                except Exception:
                    pass
            placa = (r.delivery_vehicle_plate or (ds.vehicle_plate if ds else "") or "").strip().upper()
            veic = vehicles.get(placa)
            modelo = f"{veic.marca} {veic.modelo}" if veic else (placa or "-")

            emp = emp_map.get(drv_id)
            by_driver[drv_id] = {
                "driver_id": drv_id,
                "motorista": emp.name if emp else f"Motorista #{drv_id}",
                "ajudantes": helpers,
                "km_inicial": ds.km_departure if ds else None,
                "km_final": ds.km_return if ds else None,
                "km_total": (float(ds.km_return or 0) - float(ds.km_departure or 0)) if ds and ds.km_return and ds.km_departure else None,
                "placa": placa or "-",
                "modelo": modelo,
                "turnos": set(),
                "hora_inicio": None,
                "hora_fim": None,
                "tempo_operando_min": None,
                "paradas": [],
                "saiu_kg": 0.0,
                "saiu_valor": 0.0,
                "entregue_kg": 0.0,
                "entregue_valor": 0.0,
                "devolucao_kg": 0.0,
                "devolucao_valor": 0.0,
            }

        d = by_driver[drv_id]
        d["turnos"].add((r.shift or "").strip() or "-")
        st = (r.delivery_status or "pendente").strip().lower()
        planned_kg = float(r.tonnage or 0.0)
        planned_val = float(r.valor_financeiro or 0.0)
        ret_kg = float(r.devolucao_volume if r.devolucao_volume is not None else (planned_kg if st == "devolucao" else 0.0))
        ret_val = float(r.valor_devolucao if r.valor_devolucao is not None else (planned_val if st == "devolucao" else 0.0))
        del_kg = max(0.0, planned_kg - ret_kg) if st == "devolucao" else (planned_kg if st == "entregue" else 0.0)
        del_val = max(0.0, planned_val - ret_val) if st == "devolucao" else (planned_val if st == "entregue" else 0.0)

        d["saiu_kg"] += planned_kg
        d["saiu_valor"] += planned_val
        d["entregue_kg"] += del_kg
        d["entregue_valor"] += del_val
        d["devolucao_kg"] += ret_kg
        d["devolucao_valor"] += ret_val

        cli = cli_map.get(r.client_id)
        cli_name = cli.name if cli else f"Cliente #{r.client_id}"
        start_t = r.delivery_started_at or r.start_time
        end_t = r.delivery_finished_at or r.end_time or r.delivery_returned_at
        # Duração por ciclos (cada par iniciar + finalizar/devolucao); evita somar reaberturas como tempo contínuo.
        dur = route_duration_minutes_mobile_only(r)
        tipo = "Devolução" if st == "devolucao" else "Entrega" if st == "entregue" else st

        d["paradas"].append({
            "cliente": cli_name,
            "tipo": tipo,
            "hora_inicio": _fmt_hora(start_t),
            "hora_fim": _fmt_hora(end_t),
            "duracao_min": dur,
            "duracao_fmt": _fmt_min(dur),
            "kg": planned_kg,
            "valor": planned_val,
            "entregue_kg": del_kg,
            "entregue_valor": del_val,
            "devolvido_kg": ret_kg,
            "devolvido_valor": ret_val,
        })

    reports = []
    for drv_id, d in by_driver.items():
        paradas = d["paradas"]
        paradas_ord = sorted(
            paradas,
            key=lambda x: (_tipo_priority(x.get("tipo")), x["hora_inicio"], x["cliente"])
        )
        top5_base = sorted(
            [p for p in paradas_ord if p["duracao_min"] is not None],
            key=lambda x: (-(x["duracao_min"] or 0), x["cliente"])
        )[:5]
        top5 = sorted(
            top5_base,
            key=lambda x: (_tipo_priority(x.get("tipo")), -(x["duracao_min"] or 0), x["cliente"])
        )

        clientes_resumo_map: dict[str, dict] = {}
        for parada in paradas_ord:
            cliente_key = (parada.get("cliente") or "Sem Cliente").strip() or "Sem Cliente"
            item = clientes_resumo_map.setdefault(cliente_key, {
                "cliente": cliente_key,
                "tipos": [],
                "paradas": 0,
                "duracao_total_min": 0,
                "hora_primeira_min": None,
                "hora_ultima_min": None,
                "kg_total": 0.0,
                "valor_total": 0.0,
            })
            item["paradas"] += 1
            item["kg_total"] += float(parada.get("kg") or 0.0)
            item["valor_total"] += float(parada.get("valor") or 0.0)
            if parada.get("tipo") and parada["tipo"] not in item["tipos"]:
                item["tipos"].append(parada["tipo"])

            duracao_min = parada.get("duracao_min")
            if duracao_min is not None:
                item["duracao_total_min"] += int(duracao_min)

            hora_inicio_min = _parse_hhmm(parada.get("hora_inicio"))
            if hora_inicio_min is not None and (
                item["hora_primeira_min"] is None or hora_inicio_min < item["hora_primeira_min"]
            ):
                item["hora_primeira_min"] = hora_inicio_min

            hora_fim_min = _parse_hhmm(parada.get("hora_fim"))
            if hora_fim_min is not None and (
                item["hora_ultima_min"] is None or hora_fim_min > item["hora_ultima_min"]
            ):
                item["hora_ultima_min"] = hora_fim_min

        clientes_resumo = []
        for item in clientes_resumo_map.values():
            hora_primeira_min = item["hora_primeira_min"]
            hora_ultima_min = item["hora_ultima_min"]
            tipo_principal = min((_tipo_priority(tipo) for tipo in item["tipos"]), default=2)
            clientes_resumo.append({
                "cliente": item["cliente"],
                "tipos": " / ".join(item["tipos"]) if item["tipos"] else "—",
                "tipo_principal": tipo_principal,
                "paradas": item["paradas"],
                "duracao_total_min": item["duracao_total_min"],
                "duracao_total_fmt": _fmt_min(item["duracao_total_min"]) if item["duracao_total_min"] else "--",
                "hora_primeira": f"{hora_primeira_min // 60:02d}:{hora_primeira_min % 60:02d}" if hora_primeira_min is not None else "--:--",
                "hora_ultima": f"{hora_ultima_min // 60:02d}:{hora_ultima_min % 60:02d}" if hora_ultima_min is not None else "--:--",
                "kg_total": round(item["kg_total"], 1),
                "valor_total": round(item["valor_total"], 2),
            })
        clientes_resumo.sort(
            key=lambda x: (x["tipo_principal"], x["hora_primeira"] == "--:--", x["hora_primeira"], x["cliente"].lower())
        )

        start_times = _valid_operation_times([p.get("hora_inicio") for p in paradas])
        end_times = _valid_operation_times([p.get("hora_fim") for p in paradas])

        hora_inicio = min(start_times) if start_times else None
        hora_fim = max(end_times) if end_times else (max(start_times) if start_times else None)

        d["hora_inicio"] = f"{hora_inicio // 60:02d}:{hora_inicio % 60:02d}" if hora_inicio is not None else "--:--"
        d["hora_fim"] = f"{hora_fim // 60:02d}:{hora_fim % 60:02d}" if hora_fim is not None else "--:--"

        if hora_inicio is not None and hora_fim is not None:
            d["tempo_operando_min"] = hora_fim - hora_inicio if hora_fim >= hora_inicio else ((24 * 60) - hora_inicio + hora_fim)
        else:
            d["tempo_operando_min"] = None

        base_valor = d["saiu_valor"] or 0.0
        devolucao_pct = (d["devolucao_valor"] / base_valor * 100.0) if base_valor > 0 else 0.0
        meta_pct = 2.0
        dentro_meta = devolucao_pct <= meta_pct

        # Verificar se fez checklist do caminhão (employee_id + placa + date)
        placa_raw = (d.get("placa") or "").strip()
        fez_checklist = False
        if placa_raw and placa_raw != "-":
            checklists = session.exec(
                select(models.TranspalletChecklist)
                .where(models.TranspalletChecklist.employee_id == drv_id)
                .where(models.TranspalletChecklist.date == date_str)
            ).all()
            placa_norm = placa_raw.upper().replace(" ", "").replace("-", "")
            for chk in checklists:
                eq = (chk.equipment_code or "").upper().replace(" ", "").replace("-", "")
                if eq == placa_norm:
                    fez_checklist = True
                    break

        reports.append({
            **d,
            "paradas_ordenadas": paradas_ord,
            "clientes_resumo": clientes_resumo,
            "total_clientes": len(clientes_resumo),
            "total_paradas": len(paradas_ord),
            "top5_tempo": top5,
            "top5_client_names": [p["cliente"] for p in top5],
            "tempo_operando_fmt": _fmt_min(d["tempo_operando_min"]),
            "devolucao_pct": round(devolucao_pct, 2),
            "meta_pct": meta_pct,
            "dentro_meta": dentro_meta,
            "fez_checklist_caminhao": fez_checklist,
            "turnos": sorted(d["turnos"]),
        })

    # Motoristas que tiveram rotas na data (não a lista completa)
    motoristas = [emp_map[eid] for eid in sorted(emp_map.keys(), key=lambda x: (emp_map[x].name or "").lower())]
    return {
        "date": date_str,
        "date_fmt": _fmt_br_data(date_str),
        "reports": reports,
        "motoristas": motoristas,
        "driver_ids": driver_ids or [],
    }


_RELATORIO_DRIVER_STATUS_ORDER = {
    "critico": 0,
    "pendente": 1,
    "atencao": 2,
    "concluido": 3,
}


def _coerce_relatorio_date(raw_date: Optional[str]) -> str:
    tz = ZoneInfo("America/Sao_Paulo")
    today_str = datetime.now(tz).date().strftime("%Y-%m-%d")
    value = (raw_date or "").strip() or today_str
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return today_str
    return value


def _parse_relatorio_driver_ids(raw_values) -> Optional[List[int]]:
    if not raw_values:
        return None
    values = raw_values if isinstance(raw_values, list) else [raw_values]
    parsed: list[int] = []
    seen: set[int] = set()
    for raw in values:
        text = str(raw or "").strip()
        if not text.isdigit():
            continue
        number = int(text)
        if number <= 0 or number in seen:
            continue
        seen.add(number)
        parsed.append(number)
    return parsed or None


def _normalize_relatorio_view(raw_view: Optional[str]) -> str:
    view = _norm_text(raw_view)
    if view in {"pendentes", "concluidos", "criticos", "hoje"}:
        return view
    return "todos"


def _normalize_relatorio_status(raw_status: Optional[str]) -> str:
    status_code = _norm_text(raw_status)
    if status_code in {"concluido", "atencao", "pendente", "critico"}:
        return status_code
    return ""


def _normalize_relatorio_checklist(raw_checklist: Optional[str]) -> str:
    checklist = _norm_text(raw_checklist)
    if checklist in {"com", "sem"}:
        return checklist
    return ""


def _coerce_relatorio_page_size(raw_per_page: Optional[int]) -> int:
    try:
        per_page = int(raw_per_page or 10)
    except (TypeError, ValueError):
        return 10
    return max(5, min(per_page, 100))


def _is_relatorio_driver_employee(emp) -> bool:
    role = _norm_text(getattr(emp, "role", None))
    return ("motorista" in role) and ("ajudante" not in role)


def _list_relatorio_driver_options(
    session: Session,
    selected_driver_ids: Optional[List[int]] = None,
) -> list[models.Employee]:
    selected_set = {int(x) for x in (selected_driver_ids or []) if str(x).isdigit()}
    employees = list(session.exec(select(models.Employee)).all())
    options = []
    seen: set[int] = set()
    for emp in sorted(employees, key=lambda item: ((getattr(item, "name", None) or "").casefold(), getattr(item, "id", 0) or 0)):
        emp_id = getattr(emp, "id", None)
        if not emp_id or emp_id in seen:
            continue
        status_code = _norm_text(getattr(emp, "status", None))
        if not _is_relatorio_driver_employee(emp) and emp_id not in selected_set:
            continue
        if status_code in {"fired", "demitido"} and emp_id not in selected_set:
            continue
        seen.add(emp_id)
        options.append(emp)
    return options


def _relatorio_driver_status(report: dict) -> dict:
    devolucao_pct = float(report.get("devolucao_pct") or 0.0)
    meta_pct = float(report.get("meta_pct") or 0.0)
    has_checklist = bool(report.get("fez_checklist_caminhao"))
    has_session_gap = (
        report.get("tempo_operando_min") is None
        or report.get("km_total") is None
        or (report.get("hora_inicio") or "--:--") == "--:--"
        or (report.get("hora_fim") or "--:--") == "--:--"
    )

    reasons: list[str] = []
    code = "concluido"
    label = "Concluido"
    badge = "ok"

    if devolucao_pct > meta_pct:
        code = "critico"
        label = "Critico"
        badge = "critical"
        reasons.append(
            f"Devolucao em {_fmt_br_2(devolucao_pct)}% acima da meta de {_fmt_br_2(meta_pct)}%."
        )

    if not has_checklist:
        reasons.append("Checklist do caminhao nao encontrado para o dia.")
        if code != "critico":
            code = "pendente"
            label = "Pendente"
            badge = "alert"

    if code not in {"critico", "pendente"} and has_session_gap:
        code = "atencao"
        label = "Atencao"
        badge = "alert"
        reasons.append("Sessao com horario ou quilometragem incompletos.")

    if not reasons:
        reasons.append("Meta, checklist e sessao dentro do esperado.")

    return {
        "code": code,
        "label": label,
        "badge": badge,
        "reasons": reasons,
        "has_checklist": has_checklist,
    }


def _relatorio_build_url(path: str, params: dict) -> str:
    payload = {}
    for key, value in (params or {}).items():
        if value is None:
            continue
        if isinstance(value, list):
            cleaned = [str(item).strip() for item in value if str(item or "").strip()]
            if cleaned:
                payload[key] = cleaned
            continue
        text = str(value).strip()
        if text:
            payload[key] = text
    query = urlencode(payload, doseq=True)
    return f"{path}?{query}" if query else path


def _build_relatorio_summary_row(report: dict, date_str: str) -> dict:
    status_data = _relatorio_driver_status(report)
    clients = report.get("clientes_resumo") or []
    client_names = [str(item.get("cliente") or "").strip() for item in clients if str(item.get("cliente") or "").strip()]
    search_parts = [
        report.get("motorista") or "",
        " ".join(report.get("ajudantes") or []),
        report.get("placa") or "",
        report.get("modelo") or "",
        " ".join(client_names[:12]),
    ]
    top_stop = (report.get("top5_tempo") or [{}])[0] or {}
    driver_id = int(report.get("driver_id") or 0)
    km_total = report.get("km_total")
    return {
        "driver_id": driver_id,
        "motorista": report.get("motorista") or f"Motorista #{driver_id}",
        "ajudantes_label": ", ".join(report.get("ajudantes") or []) or "Sem ajudantes",
        "placa": report.get("placa") or "-",
        "modelo": report.get("modelo") or "-",
        "turnos_label": " / ".join(report.get("turnos") or []) or "-",
        "total_clientes": int(report.get("total_clientes") or 0),
        "total_paradas": int(report.get("total_paradas") or 0),
        "top_cliente": client_names[0] if client_names else "-",
        "top_parada_label": top_stop.get("cliente") or "-",
        "hora_inicio": report.get("hora_inicio") or "--:--",
        "hora_fim": report.get("hora_fim") or "--:--",
        "tempo_operando_fmt": report.get("tempo_operando_fmt") or "—",
        "saiu_valor": round(float(report.get("saiu_valor") or 0.0), 2),
        "saiu_valor_fmt": _fmt_br_moeda(report.get("saiu_valor") or 0.0),
        "entregue_valor_fmt": _fmt_br_moeda(report.get("entregue_valor") or 0.0),
        "devolucao_valor": round(float(report.get("devolucao_valor") or 0.0), 2),
        "devolucao_valor_fmt": _fmt_br_moeda(report.get("devolucao_valor") or 0.0),
        "devolucao_pct": round(float(report.get("devolucao_pct") or 0.0), 2),
        "devolucao_pct_fmt": f"{_fmt_br_2(report.get('devolucao_pct') or 0.0)}%",
        "meta_pct_fmt": f"{_fmt_br_2(report.get('meta_pct') or 0.0)}%",
        "meta_pct": round(float(report.get("meta_pct") or 0.0), 2),
        "checklist_label": "Conferido" if status_data["has_checklist"] else "Pendente",
        "checklist_badge": "ok" if status_data["has_checklist"] else "critical",
        "status_code": status_data["code"],
        "status_label": status_data["label"],
        "status_badge": status_data["badge"],
        "status_reason": status_data["reasons"][0],
        "status_reasons": status_data["reasons"],
        "km_total": km_total,
        "km_total_fmt": (f"{_fmt_br_1(km_total)} km" if km_total is not None else "—"),
        "has_checklist": status_data["has_checklist"],
        "date": date_str,
        "print_href": _relatorio_build_url(
            "/relatorio-avaliacao-motorista/impressao",
            {"date": date_str, "driver_id": [driver_id]},
        ),
        "search_blob": " ".join(search_parts).casefold(),
    }


def _make_relatorio_links(
    date_str: str,
    selected_driver_ids: Optional[List[int]],
    q: str,
    status_filter: str,
    checklist_filter: str,
    per_page: int,
    page: int,
    view: str,
) -> dict:
    base_params = {
        "date": date_str,
        "driver_id": selected_driver_ids or [],
        "q": q,
        "status": status_filter,
        "checklist": checklist_filter,
        "per_page": per_page,
    }
    paged_params = {**base_params, "page": page}
    links = {
        "todos": _relatorio_build_url("/relatorio-avaliacao-motorista", base_params),
        "pendentes": _relatorio_build_url("/relatorio-avaliacao-motorista", {**base_params, "view": "pendentes"}),
        "concluidos": _relatorio_build_url("/relatorio-avaliacao-motorista", {**base_params, "view": "concluidos"}),
        "criticos": _relatorio_build_url("/relatorio-avaliacao-motorista", {**base_params, "view": "criticos"}),
        "hoje": _relatorio_build_url(
            "/relatorio-avaliacao-motorista",
            {**base_params, "date": _coerce_relatorio_date(None), "view": "hoje"},
        ),
        "clear": "/relatorio-avaliacao-motorista",
        "current": _relatorio_build_url("/relatorio-avaliacao-motorista", {**paged_params, "view": ("" if view == "todos" else view)}),
        "export": _relatorio_build_url(
            "/relatorio-avaliacao-motorista/export.csv",
            {**base_params, "view": ("" if view == "todos" else view)},
        ),
        "print": _relatorio_build_url(
            "/relatorio-avaliacao-motorista/impressao",
            {"date": date_str, "driver_id": selected_driver_ids or []},
        ),
    }
    return links


def _build_relatorio_page_fallback(
    session: Session,
    date_str: str,
    selected_driver_ids: Optional[List[int]],
    q: str,
    view: str,
    status_filter: str,
    checklist_filter: str,
    page: int,
    per_page: int,
    load_error: Optional[str] = None,
) -> dict:
    driver_options = _list_relatorio_driver_options(session, selected_driver_ids)
    selected_map = {int(emp.id): emp for emp in driver_options if getattr(emp, "id", None)}
    selected_names = [
        selected_map[driver_id].name
        for driver_id in (selected_driver_ids or [])
        if driver_id in selected_map and getattr(selected_map[driver_id], "name", None)
    ]
    return {
        "date": date_str,
        "date_fmt": _fmt_br_data(date_str),
        "today_str": _coerce_relatorio_date(None),
        "filters": {
            "q": q,
            "status": status_filter,
            "checklist": checklist_filter,
            "page": 1,
            "per_page": per_page,
        },
        "view": view,
        "rows": [],
        "rows_all_filtered": [],
        "has_any_rows": False,
        "has_filtered_rows": False,
        "kpi": {
            "total": 0,
            "concluidos": 0,
            "criticos": 0,
            "pendentes": 0,
            "sem_checklist": 0,
        },
        "links": _make_relatorio_links(
            date_str=date_str,
            selected_driver_ids=selected_driver_ids,
            q=q,
            status_filter=status_filter,
            checklist_filter=checklist_filter,
            per_page=per_page,
            page=page,
            view=view,
        ),
        "pagination": {
            "page": 1,
            "per_page": per_page,
            "total_count": 0,
            "total_pages": 1,
            "page_start": 0,
            "page_end": 0,
            "has_prev": False,
            "has_next": False,
            "prev_href": "#",
            "next_href": "#",
        },
        "motoristas": driver_options,
        "selected_driver_ids": selected_driver_ids or [],
        "selected_driver_names": selected_names[:6],
        "selected_driver_names_more": max(0, len(selected_names) - 6),
        "load_error": load_error,
    }


def _build_relatorio_avaliacao_page_data(
    session: Session,
    date_str: str,
    selected_driver_ids: Optional[List[int]],
    q: str,
    view: str,
    status_filter: str,
    checklist_filter: str,
    page: int,
    per_page: int,
) -> dict:
    base_data = _build_relatorio_avaliacao_motorista(session, date_str, selected_driver_ids)
    rows_all = [_build_relatorio_summary_row(report, date_str) for report in base_data.get("reports") or []]
    rows_all.sort(
        key=lambda row: (
            _RELATORIO_DRIVER_STATUS_ORDER.get(row["status_code"], 99),
            -float(row.get("devolucao_pct") or 0.0),
            (row.get("motorista") or "").casefold(),
        )
    )

    q_norm = _norm_text(q)
    rows_scoped = rows_all
    if q_norm:
        rows_scoped = [row for row in rows_all if q_norm in row.get("search_blob", "")]

    kpi = {
        "total": len(rows_scoped),
        "concluidos": sum(1 for row in rows_scoped if row["status_code"] == "concluido"),
        "criticos": sum(1 for row in rows_scoped if row["status_code"] == "critico"),
        "pendentes": sum(1 for row in rows_scoped if row["status_code"] == "pendente"),
        "sem_checklist": sum(1 for row in rows_scoped if not row["has_checklist"]),
    }

    filtered_rows = list(rows_scoped)
    if status_filter:
        filtered_rows = [row for row in filtered_rows if row["status_code"] == status_filter]
    if checklist_filter == "com":
        filtered_rows = [row for row in filtered_rows if row["has_checklist"]]
    elif checklist_filter == "sem":
        filtered_rows = [row for row in filtered_rows if not row["has_checklist"]]

    if view == "pendentes":
        filtered_rows = [row for row in filtered_rows if row["status_code"] == "pendente"]
    elif view == "concluidos":
        filtered_rows = [row for row in filtered_rows if row["status_code"] == "concluido"]
    elif view == "criticos":
        filtered_rows = [row for row in filtered_rows if row["status_code"] == "critico"]

    total_count = len(filtered_rows)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    current_page = max(1, min(page, total_pages))
    start_idx = (current_page - 1) * per_page
    end_idx = start_idx + per_page
    rows_page = filtered_rows[start_idx:end_idx]

    driver_options = _list_relatorio_driver_options(session, selected_driver_ids)
    selected_map = {int(emp.id): emp for emp in driver_options if getattr(emp, "id", None)}
    selected_names = [
        selected_map[driver_id].name
        for driver_id in (selected_driver_ids or [])
        if driver_id in selected_map and getattr(selected_map[driver_id], "name", None)
    ]

    links = _make_relatorio_links(
        date_str=date_str,
        selected_driver_ids=selected_driver_ids,
        q=q,
        status_filter=status_filter,
        checklist_filter=checklist_filter,
        per_page=per_page,
        page=current_page,
        view=view,
    )

    def _page_href(target_page: int) -> str:
        return _relatorio_build_url(
            "/relatorio-avaliacao-motorista",
            {
                "date": date_str,
                "driver_id": selected_driver_ids or [],
                "q": q,
                "status": status_filter,
                "checklist": checklist_filter,
                "per_page": per_page,
                "page": target_page,
                "view": ("" if view == "todos" else view),
            },
        )

    page_start = (start_idx + 1) if total_count else 0
    page_end = min(end_idx, total_count) if total_count else 0

    return {
        "date": date_str,
        "date_fmt": _fmt_br_data(date_str),
        "today_str": _coerce_relatorio_date(None),
        "filters": {
            "q": q,
            "status": status_filter,
            "checklist": checklist_filter,
            "page": current_page,
            "per_page": per_page,
        },
        "view": view,
        "rows": rows_page,
        "rows_all_filtered": filtered_rows,
        "has_any_rows": bool(rows_all),
        "has_filtered_rows": bool(filtered_rows),
        "kpi": kpi,
        "links": links,
        "pagination": {
            "page": current_page,
            "per_page": per_page,
            "total_count": total_count,
            "total_pages": total_pages,
            "page_start": page_start,
            "page_end": page_end,
            "has_prev": current_page > 1,
            "has_next": current_page < total_pages,
            "prev_href": _page_href(current_page - 1) if current_page > 1 else "#",
            "next_href": _page_href(current_page + 1) if current_page < total_pages else "#",
        },
        "motoristas": driver_options,
        "selected_driver_ids": selected_driver_ids or [],
        "selected_driver_names": selected_names[:6],
        "selected_driver_names_more": max(0, len(selected_names) - 6),
        "load_error": None,
    }


def _build_relatorio_detail_payload(report: dict, date_str: str) -> dict:
    status_data = _relatorio_driver_status(report)
    top_time = []
    for item in report.get("top5_tempo") or []:
        top_time.append(
            {
                "cliente": item.get("cliente") or "-",
                "tipo": item.get("tipo") or "-",
                "duracao": item.get("duracao_fmt") or "—",
            }
        )

    clients = []
    for item in report.get("clientes_resumo") or []:
        clients.append(
            {
                "cliente": item.get("cliente") or "-",
                "tipos": item.get("tipos") or "-",
                "paradas": int(item.get("paradas") or 0),
                "duracao": item.get("duracao_total_fmt") or "—",
                "janela": f"{item.get('hora_primeira') or '--:--'} -> {item.get('hora_ultima') or '--:--'}",
                "valor": _fmt_br_moeda(item.get("valor_total") or 0.0),
            }
        )

    stops = []
    for item in report.get("paradas_ordenadas") or []:
        stops.append(
            {
                "cliente": item.get("cliente") or "-",
                "tipo": item.get("tipo") or "-",
                "inicio": item.get("hora_inicio") or "--:--",
                "fim": item.get("hora_fim") or "--:--",
                "duracao": item.get("duracao_fmt") or "—",
                "kg": f"{_fmt_br_1(item.get('kg') or 0.0)} kg",
                "valor": _fmt_br_moeda(item.get("valor") or 0.0),
            }
        )

    driver_id = int(report.get("driver_id") or 0)
    km_total = report.get("km_total")
    return {
        "driver_id": driver_id,
        "date": date_str,
        "date_fmt": _fmt_br_data(date_str),
        "motorista": report.get("motorista") or f"Motorista #{driver_id}",
        "ajudantes": report.get("ajudantes") or [],
        "placa": report.get("placa") or "-",
        "modelo": report.get("modelo") or "-",
        "turnos_label": " / ".join(report.get("turnos") or []) or "-",
        "hora_inicio": report.get("hora_inicio") or "--:--",
        "hora_fim": report.get("hora_fim") or "--:--",
        "tempo_operando": report.get("tempo_operando_fmt") or "—",
        "km_inicial": (_fmt_br_1(report.get("km_inicial")) if report.get("km_inicial") is not None else "—"),
        "km_final": (_fmt_br_1(report.get("km_final")) if report.get("km_final") is not None else "—"),
        "km_total": (f"{_fmt_br_1(km_total)} km" if km_total is not None else "—"),
        "saiu_valor": _fmt_br_moeda(report.get("saiu_valor") or 0.0),
        "entregue_valor": _fmt_br_moeda(report.get("entregue_valor") or 0.0),
        "devolucao_valor": _fmt_br_moeda(report.get("devolucao_valor") or 0.0),
        "devolucao_pct": f"{_fmt_br_2(report.get('devolucao_pct') or 0.0)}%",
        "meta_pct": f"{_fmt_br_2(report.get('meta_pct') or 0.0)}%",
        "checklist_label": ("Checklist enviado" if report.get("fez_checklist_caminhao") else "Checklist pendente"),
        "checklist_badge": ("ok" if report.get("fez_checklist_caminhao") else "critical"),
        "status_label": status_data["label"],
        "status_badge": status_data["badge"],
        "status_reasons": status_data["reasons"],
        "total_clientes": int(report.get("total_clientes") or 0),
        "total_paradas": int(report.get("total_paradas") or 0),
        "top_time": top_time,
        "clients": clients,
        "stops": stops,
        "print_href": _relatorio_build_url(
            "/relatorio-avaliacao-motorista/impressao",
            {"date": date_str, "driver_id": [driver_id]},
        ),
    }


@router.get("/relatorio-avaliacao-motorista", response_class=HTMLResponse)
async def relatorio_avaliacao_motorista_page(
    request: Request,
    date: Optional[str] = None,
    q: str = "",
    view: str = "todos",
    status: str = "",
    checklist: str = "",
    page: int = 1,
    per_page: int = 10,
    driver_id: Optional[List[str]] = Query(None, alias="driver_id"),
    session: Session = Depends(get_session),
):
    """Pagina operacional do relatorio de avaliacao diaria do motorista."""
    view_norm = _normalize_relatorio_view(view)
    date_str = _coerce_relatorio_date(date)
    if view_norm == "hoje":
        date_str = _coerce_relatorio_date(None)
    q_value = (q or "").strip()
    parsed_ids = _parse_relatorio_driver_ids(driver_id)
    status_filter = _normalize_relatorio_status(status)
    checklist_filter = _normalize_relatorio_checklist(checklist)
    current_page = max(1, int(page or 1))
    page_size = _coerce_relatorio_page_size(per_page)

    try:
        data = _build_relatorio_avaliacao_page_data(
            session=session,
            date_str=date_str,
            selected_driver_ids=parsed_ids,
            q=q_value,
            view=view_norm,
            status_filter=status_filter,
            checklist_filter=checklist_filter,
            page=current_page,
            per_page=page_size,
        )
        status_code = 200
    except Exception:
        data = _build_relatorio_page_fallback(
            session=session,
            date_str=date_str,
            selected_driver_ids=parsed_ids,
            q=q_value,
            view=view_norm,
            status_filter=status_filter,
            checklist_filter=checklist_filter,
            page=current_page,
            per_page=page_size,
            load_error="Nao foi possivel carregar a avaliacao agora. Tente novamente com outro filtro ou recarregue a pagina.",
        )
        status_code = 500

    return templates.TemplateResponse(
        "relatorio_avaliacao_motorista_operacional.html",
        {
            "request": request,
            **data,
            "message": request.query_params.get("message"),
            "level": request.query_params.get("level") or "info",
        },
        status_code=status_code,
    )


@router.get("/relatorio-avaliacao-motorista/impressao", response_class=HTMLResponse)
async def relatorio_avaliacao_motorista_print_page(
    request: Request,
    date: Optional[str] = None,
    driver_id: Optional[List[str]] = Query(None, alias="driver_id"),
    session: Session = Depends(get_session),
):
    """Mantem a visao imprimivel do relatorio por motorista."""
    date_str = _coerce_relatorio_date(date)
    parsed_ids = _parse_relatorio_driver_ids(driver_id)
    data = _build_relatorio_avaliacao_motorista(session, date_str, parsed_ids)
    data["motoristas"] = _list_relatorio_driver_options(session, parsed_ids)
    return templates.TemplateResponse("relatorio_avaliacao_motorista.html", {"request": request, **data})


@router.get("/api/relatorio-avaliacao-motorista/detail", response_class=JSONResponse)
async def relatorio_avaliacao_motorista_detail_api(
    date: Optional[str] = None,
    driver_id: Optional[str] = None,
    session: Session = Depends(get_session),
):
    """Detalhe lazy do motorista para modal operacional."""
    date_str = _coerce_relatorio_date(date)
    parsed_ids = _parse_relatorio_driver_ids(driver_id)
    if not parsed_ids:
        return JSONResponse({"error": "Motorista invalido."}, status_code=400)

    data = _build_relatorio_avaliacao_motorista(session, date_str, parsed_ids)
    report = next(
        (item for item in data.get("reports") or [] if int(item.get("driver_id") or 0) == int(parsed_ids[0])),
        None,
    )
    if not report:
        return JSONResponse({"error": "Nenhum relatorio encontrado para este motorista."}, status_code=404)

    return JSONResponse(_build_relatorio_detail_payload(report, date_str))


@router.get("/relatorio-avaliacao-motorista/export.csv")
async def relatorio_avaliacao_motorista_export_csv(
    date: Optional[str] = None,
    q: str = "",
    view: str = "todos",
    status: str = "",
    checklist: str = "",
    per_page: int = 5000,
    driver_id: Optional[List[str]] = Query(None, alias="driver_id"),
    session: Session = Depends(get_session),
):
    """Exporta a visao filtrada da avaliacao operacional."""
    view_norm = _normalize_relatorio_view(view)
    date_str = _coerce_relatorio_date(date)
    if view_norm == "hoje":
        date_str = _coerce_relatorio_date(None)

    data = _build_relatorio_avaliacao_page_data(
        session=session,
        date_str=date_str,
        selected_driver_ids=_parse_relatorio_driver_ids(driver_id),
        q=(q or "").strip(),
        view=view_norm,
        status_filter=_normalize_relatorio_status(status),
        checklist_filter=_normalize_relatorio_checklist(checklist),
        page=1,
        per_page=_coerce_relatorio_page_size(per_page),
    )

    rows = data.get("rows_all_filtered") or []
    stamp = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d_%H%M")
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(
        [
            "data",
            "motorista",
            "turnos",
            "ajudantes",
            "placa",
            "modelo",
            "clientes",
            "paradas",
            "inicio",
            "fim",
            "tempo_operando",
            "km_total",
            "valor_expedido",
            "valor_entregue",
            "valor_devolvido",
            "pct_devolucao",
            "meta_pct",
            "checklist",
            "status",
            "observacao",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                _fmt_br_data(row.get("date")),
                row.get("motorista") or "",
                row.get("turnos_label") or "",
                row.get("ajudantes_label") or "",
                row.get("placa") or "",
                row.get("modelo") or "",
                row.get("total_clientes") or 0,
                row.get("total_paradas") or 0,
                row.get("hora_inicio") or "",
                row.get("hora_fim") or "",
                row.get("tempo_operando_fmt") or "",
                row.get("km_total_fmt") or "",
                row.get("saiu_valor_fmt") or "",
                row.get("entregue_valor_fmt") or "",
                row.get("devolucao_valor_fmt") or "",
                row.get("devolucao_pct_fmt") or "",
                row.get("meta_pct_fmt") or "",
                row.get("checklist_label") or "",
                row.get("status_label") or "",
                row.get("status_reason") or "",
            ]
        )
    buffer = io.BytesIO(out.getvalue().encode("utf-8-sig"))
    return StreamingResponse(
        buffer,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=avaliacao_motoristas_{stamp}.csv"},
    )


def _bi_vendedor_delivery_maps_and_resolver(
    session: Session,
    date_from: Optional[str],
    date_to: Optional[str],
) -> tuple[
    dict,
    list[dict],
    list[dict],
    Callable[[Optional[int]], tuple[Optional[int], str]],
]:
    """Dataset BI entregas, linhas de rota, financeiras de devolução e resolvedor cliente→vendedor."""
    delivery_dataset = _build_bi_delivery_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        shift="Todos",
        driver_id=None,
        plate="Todos",
        status="Todos",
    )
    route_rows = [
        r for r in (delivery_dataset.get("all_route_rows") or [])
        if str(r.get("source") or "").upper() == "ROTA"
    ]
    financial_rows = list(delivery_dataset.get("all_financial_rows") or [])
    client_ids = sorted({
        int(r.get("client_id"))
        for r in (route_rows + financial_rows)
        if r.get("client_id") is not None
    })
    clients = (
        session.exec(select(models.Client).where(models.Client.id.in_(client_ids))).all()
        if client_ids else []
    )
    client_map = {c.id: c for c in clients}
    seller_ids = sorted({
        int(c.vendedor_id)
        for c in clients
        if getattr(c, "vendedor_id", None) is not None
    })
    sellers = (
        session.exec(select(models.Employee).where(models.Employee.id.in_(seller_ids))).all()
        if seller_ids else []
    )
    seller_map = {s.id: s for s in sellers}

    def _seller_for_client(client_id: Optional[int]) -> tuple[Optional[int], str]:
        if client_id is None:
            return None, "Sem vendedor"
        cli = client_map.get(int(client_id))
        sid = getattr(cli, "vendedor_id", None) if cli else None
        if sid is None:
            return None, "Sem vendedor"
        seller = seller_map.get(int(sid))
        if seller:
            return int(sid), str(seller.name or f"Vendedor #{sid}")
        return int(sid), f"Vendedor #{sid}"

    return delivery_dataset, route_rows, financial_rows, _seller_for_client


def _list_bi_vendedor_devolvidos(
    session: Session,
    date_from: Optional[str],
    date_to: Optional[str],
    vendedor_id: int,
) -> dict:
    """Lista ocorrências de devolução (cadastro financeiro) atribuídas ao vendedor. Use vendedor_id=-1 para sem vendedor."""
    _, _, financial_rows, seller_for = _bi_vendedor_delivery_maps_and_resolver(session, date_from, date_to)
    sem_vendedor = vendedor_id == -1
    items: list[dict] = []
    for row in financial_rows:
        sid, _ = seller_for(row.get("client_id"))
        if sem_vendedor:
            if sid is not None:
                continue
        elif sid != vendedor_id:
            continue
        val = float(row.get("value") or row.get("returned_value") or 0.0)
        items.append({
            "client_id": row.get("client_id"),
            "client_name": str(row.get("client_name") or "").strip() or "—",
            "date": str(row.get("date") or "")[:10],
            "valor": round(val, 2),
            "motorista": str(row.get("driver_name") or "").strip() or "—",
            "motivo": str(row.get("motivo") or "").strip() or "—",
            "responsabilidade": str(row.get("responsabilidade") or "").strip() or "—",
        })
    items.sort(key=lambda x: (x["date"], x["client_name"]), reverse=True)
    return {
        "items": items,
        "total_qtd": len(items),
        "total_valor": round(sum(float(i["valor"]) for i in items), 2),
    }


def _build_bi_vendedor_dataset(
    session: Session,
    date_from: Optional[str],
    date_to: Optional[str],
    vendedor_id: Optional[int] = None,
) -> dict:
    """Consolida vendas e devoluções por vendedor para ranking comercial."""
    delivery_dataset, route_rows, financial_rows, _seller_for_client = _bi_vendedor_delivery_maps_and_resolver(
        session, date_from, date_to
    )

    seller_acc: dict[int, dict] = {}
    sem_vendedor_key = -1

    def _ensure_bucket(sid: Optional[int], name: str) -> dict:
        key = int(sid) if sid is not None else sem_vendedor_key
        if key not in seller_acc:
            seller_acc[key] = {
                "vendedor_id": None if sid is None else int(sid),
                "vendedor": name,
                "clientes_set": set(),
                "pedidos_total": 0,
                "pedidos_concluidos": 0,
                "vendas_planejadas": 0.0,
                "vendas_realizadas": 0.0,
                "devolucoes_qtd": 0,
                "devolucoes_valor": 0.0,
            }
        return seller_acc[key]

    for row in route_rows:
        sid, sname = _seller_for_client(row.get("client_id"))
        bucket = _ensure_bucket(sid, sname)
        cid = row.get("client_id")
        if cid is not None:
            bucket["clientes_set"].add(int(cid))
        bucket["pedidos_total"] += 1
        status = str(row.get("status") or "").lower()
        if status in ("entregue", "devolucao"):
            bucket["pedidos_concluidos"] += 1
        bucket["vendas_planejadas"] += float(row.get("planned_value") or 0.0)
        bucket["vendas_realizadas"] += float(row.get("delivered_value") or 0.0)

    for row in financial_rows:
        sid, sname = _seller_for_client(row.get("client_id"))
        bucket = _ensure_bucket(sid, sname)
        cid = row.get("client_id")
        if cid is not None:
            bucket["clientes_set"].add(int(cid))
        val = float(row.get("value") or row.get("returned_value") or 0.0)
        bucket["devolucoes_qtd"] += 1
        bucket["devolucoes_valor"] += val

    rows = []
    for bucket in seller_acc.values():
        base_valor = float(bucket["vendas_realizadas"] or 0.0) + float(bucket["devolucoes_valor"] or 0.0)
        taxa_devol = _safe_pct(bucket["devolucoes_valor"], base_valor)
        ticket_medio = (
            float(bucket["vendas_realizadas"]) / float(bucket["pedidos_concluidos"])
            if bucket["pedidos_concluidos"] > 0 else 0.0
        )
        rows.append({
            "vendedor_id": bucket["vendedor_id"],
            "vendedor": bucket["vendedor"],
            "clientes_ativos": len(bucket["clientes_set"]),
            "pedidos_total": int(bucket["pedidos_total"]),
            "pedidos_concluidos": int(bucket["pedidos_concluidos"]),
            "vendas_planejadas": round(float(bucket["vendas_planejadas"]), 2),
            "vendas_realizadas": round(float(bucket["vendas_realizadas"]), 2),
            "devolucoes_qtd": int(bucket["devolucoes_qtd"]),
            "devolucoes_valor": round(float(bucket["devolucoes_valor"]), 2),
            "taxa_devolucao_valor": round(float(taxa_devol), 2),
            "ticket_medio": round(float(ticket_medio), 2),
        })

    all_rows = sorted(rows, key=lambda x: (-x["vendas_realizadas"], -x["devolucoes_valor"], x["vendedor"]))
    filtered_rows = [
        r for r in all_rows
        if vendedor_id is None or r.get("vendedor_id") == vendedor_id
    ]

    base_rows = filtered_rows if filtered_rows else all_rows
    top_vendas = sorted(base_rows, key=lambda x: x["vendas_realizadas"], reverse=True)[:10]
    top_devolucoes = sorted(base_rows, key=lambda x: x["devolucoes_valor"], reverse=True)[:10]
    top_taxa = [
        r for r in sorted(base_rows, key=lambda x: x["taxa_devolucao_valor"])
        if (r.get("vendas_realizadas") or 0) > 0
    ][:10]

    maior_vendas = top_vendas[0] if top_vendas else None
    maior_devolucao = top_devolucoes[0] if top_devolucoes else None
    menor_taxa = top_taxa[0] if top_taxa else None
    soma_vendas = round(sum(float(r["vendas_realizadas"]) for r in base_rows), 2)
    soma_devolucoes_rows = round(sum(float(r["devolucoes_valor"]) for r in base_rows), 2)
    # Total em R$ alinhado ao BI Entregas / cadastro (evita divergência vs soma das linhas por vendedor).
    soma_devolucoes = round(
        float(delivery_dataset.get("kpis", {}).get("valor_total_devolvido") or soma_devolucoes_rows),
        2,
    )
    taxa_geral = round(_safe_pct(soma_devolucoes, soma_vendas + soma_devolucoes), 2)

    filters_payload = {
        "date_from": delivery_dataset.get("filters", {}).get("date_from"),
        "date_to": delivery_dataset.get("filters", {}).get("date_to"),
        "vendedor_id": vendedor_id,
    }
    filters_query = urlencode({
        "date_from": filters_payload["date_from"] or "",
        "date_to": filters_payload["date_to"] or "",
        "vendedor_id": vendedor_id if vendedor_id is not None else "",
    })

    chart_payload = {
        "top_vendas": top_vendas,
        "top_devolucoes": top_devolucoes,
        "top_taxa": top_taxa,
    }

    sellers_filter = sorted(
        [{"id": r.get("vendedor_id"), "name": r.get("vendedor")} for r in all_rows if r.get("vendedor_id") is not None],
        key=lambda x: str(x["name"] or ""),
    )

    return {
        "filters": filters_payload,
        "filters_query": filters_query,
        "sellers_filter": sellers_filter,
        "rows": base_rows,
        "kpis": {
            "vendedores_ativos": len(base_rows),
            "vendas_realizadas": soma_vendas,
            "devolucoes_valor": soma_devolucoes,
            "taxa_devolucao_geral": taxa_geral,
        },
        "maior_vendas": maior_vendas,
        "maior_devolucao": maior_devolucao,
        "menor_taxa": menor_taxa,
        "chart_payload_json": json.dumps(chart_payload, ensure_ascii=False),
    }


@router.get("/bi/vendedor", response_class=HTMLResponse)
async def bi_vendedor_page(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendedor_id: Optional[str] = None,
    session: Session = Depends(get_session),
):
    parsed_vendedor_id: Optional[int] = int(vendedor_id) if (vendedor_id or "").strip().isdigit() else None
    dataset = _build_bi_vendedor_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        vendedor_id=parsed_vendedor_id,
    )
    return templates.TemplateResponse("bi_vendedor.html", {"request": request, **dataset})


@router.get("/bi/vendedor/devolvidos", response_class=JSONResponse)
async def bi_vendedor_devolvidos_json(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendedor_id: str = Query(..., description="ID do vendedor no cadastro ou -1 para Sem vendedor"),
    session: Session = Depends(get_session),
):
    raw = (vendedor_id or "").strip()
    if raw == "-1":
        parsed_vid = -1
    elif raw.isdigit():
        parsed_vid = int(raw)
    else:
        return JSONResponse({"error": "Parametro vendedor_id invalido."}, status_code=400)
    payload = _list_bi_vendedor_devolvidos(
        session=session,
        date_from=date_from,
        date_to=date_to,
        vendedor_id=parsed_vid,
    )
    return JSONResponse(payload)


@router.get("/bi/delivery", response_class=HTMLResponse)
async def bi_delivery_page(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    shift: str = "Todos",
    driver_id: Optional[str] = None,
    plate: str = "Todos",
    status: str = "Todos",
    detail_driver_id: Optional[str] = None,
    detail_status: str = "Todos",
    session: Session = Depends(get_session),
):
    parsed_driver_id: Optional[int] = int(driver_id) if (driver_id or "").strip().isdigit() else None
    parsed_detail_driver_id: Optional[int] = int(detail_driver_id) if (detail_driver_id or "").strip().isdigit() else None
    dataset = _build_bi_delivery_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        shift=shift,
        driver_id=parsed_driver_id,
        plate=plate,
        status=status,
        detail_driver_id=parsed_detail_driver_id,
        detail_status=detail_status,
    )
    return templates.TemplateResponse("bi_delivery.html", {"request": request, **dataset})


@router.get("/bi/clientes", response_class=HTMLResponse)
async def bi_clientes_page(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    shift: str = "Todos",
    driver_id: Optional[str] = None,
    plate: str = "Todos",
    status: str = "Todos",
    client_id: Optional[str] = None,
    city: str = "Todos",
    priority: str = "Todos",
    client_status: str = "Todos",
    segmentos: Optional[list[str]] = Query(default=None),
    returns_filter: str = "Todos",
    detail_client_id: Optional[str] = None,
    client_scope: Optional[str] = None,
    vendedor_id: Optional[str] = None,
    search_q: str = "",
    classification_filter: str = "Todos",
    motivo_filter: str = "Todos",
    responsabilidade_filter: str = "Todos",
    purchase_band: str = "Todos",
    return_band: str = "Todos",
    duration_band: str = "Todos",
    session: Session = Depends(get_session),
):
    parsed_driver_id: Optional[int] = int(driver_id) if (driver_id or "").strip().isdigit() else None
    parsed_client_id: Optional[int] = int(client_id) if (client_id or "").strip().isdigit() else None
    parsed_vendedor_id: Optional[int] = None
    if (vendedor_id or "").strip() == "-1":
        parsed_vendedor_id = -1
    elif (vendedor_id or "").strip().isdigit():
        parsed_vendedor_id = int(vendedor_id)
    parsed_detail = (detail_client_id or "").strip()
    if parsed_detail.startswith("-") and parsed_detail[1:].isdigit():
        parsed_detail_client_id = int(parsed_detail)
    elif parsed_detail.isdigit():
        parsed_detail_client_id = int(parsed_detail)
    else:
        parsed_detail_client_id = None
    dataset = _build_bi_clientes_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        shift=shift,
        driver_id=parsed_driver_id,
        plate=plate,
        status=status,
        client_id=parsed_client_id,
        city=city,
        priority=priority,
        client_status=client_status,
        segmentos=segmentos,
        returns_filter=returns_filter,
        detail_client_id=parsed_detail_client_id,
        client_filter_scope=(client_scope or "solo").strip().lower(),
        vendedor_id=parsed_vendedor_id,
        search_q=search_q,
        classification_filter=classification_filter,
        motivo_filter=motivo_filter,
        responsabilidade_filter=responsabilidade_filter,
        purchase_band=purchase_band,
        return_band=return_band,
        duration_band=duration_band,
    )
    return templates.TemplateResponse("bi_clientes.html", {"request": request, **dataset})


@router.get("/bi/delivery/export")
async def bi_delivery_export(
    format: str = "csv",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    shift: str = "Todos",
    driver_id: Optional[str] = None,
    plate: str = "Todos",
    status: str = "Todos",
    session: Session = Depends(get_session),
):
    parsed_driver_id: Optional[int] = int(driver_id) if (driver_id or "").strip().isdigit() else None
    dataset = _build_bi_delivery_dataset(session, date_from, date_to, shift, parsed_driver_id, plate, status)
    rows = dataset["all_route_rows"]
    stamp = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d_%H%M")
    fmt = (format or "csv").strip().lower()

    def _row_data_br(r):
        data_br = _fmt_br_data(r["date"])
        kg_p = _fmt_br_2(r.get("planned_kg", 0))
        kg_d = _fmt_br_2(r.get("delivered_kg", 0))
        kg_ret = _fmt_br_2(r.get("returned_kg", 0))
        v_p = _fmt_br_2(r.get("planned_value", 0))
        v_d = _fmt_br_2(r.get("delivered_value", 0))
        v_ret = _fmt_br_2(r.get("returned_value", 0))
        return [r["route_id"], data_br, r["shift"], r["driver_name"], r["client_name"], r["status"], kg_p, kg_d, kg_ret, v_p, v_d, v_ret, r["reopen_count"], r["duration_m"] or "", r["plate"], r["order_number"], r.get("motivo", ""), r.get("responsabilidade", ""), r.get("cluster", ""), r.get("acima_300", ""), r.get("source", "")]

    if fmt == "csv":
        out = io.StringIO()
        w = csv.writer(out, delimiter=";")
        w.writerow(["rota_id", "data", "turno", "motorista", "cliente", "status", "kg_planejado", "kg_entregue", "kg_devolvido", "valor_planejado", "valor_entregue", "valor_devolvido", "reaberturas", "duracao_min", "placa", "pedido", "motivo", "responsabilidade", "cluster", "acima_300", "origem"])
        for r in rows:
            w.writerow(_row_data_br(r))
        buf = io.BytesIO(out.getvalue().encode("utf-8-sig"))
        return StreamingResponse(buf, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename=bi_entregas_{stamp}.csv"})

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "BI Entregas"
        ws.append(["Rota ID", "Data", "Turno", "Motorista", "Cliente", "Status", "Kg Planejado", "Kg Entregue", "Kg Devolvido", "Valor Planejado", "Valor Entregue", "Valor Devolvido", "Reaberturas", "Duracao (min)", "Placa", "Pedido", "Motivo", "Responsabilidade", "Cluster", "Acima 300", "Origem"])
        for r in rows:
            ws.append(_row_data_br(r))
        xbuf = io.BytesIO()
        wb.save(xbuf)
        xbuf.seek(0)
        return StreamingResponse(xbuf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=bi_entregas_{stamp}.xlsx"})

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        pbuf = io.BytesIO()
        c = canvas.Canvas(pbuf, pagesize=A4)
        _, h = A4
        y = h - 40
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "BI Entregas - Relatorio Executivo")
        y -= 16
        c.setFont("Helvetica", 9)
        period_from = _fmt_br_data(dataset["filters"]["date_from"])
        period_to = _fmt_br_data(dataset["filters"]["date_to"])
        c.drawString(30, y, f"Periodo: {period_from} ate {period_to}")
        y -= 14
        c.drawString(
            30,
            y,
            f"Planejadas: {dataset['kpis']['planned_stops']} | Entregues: {dataset['kpis']['delivered_stops']} | "
            f"SLA conclusao: {_fmt_br_1(dataset['kpis']['sla_finish'])}% | Tempo medio: {dataset['kpis']['avg_duration_m']} min",
        )
        y -= 20
        c.setFont("Helvetica", 8)
        for r in rows[:180]:
            if y <= 30:
                c.showPage()
                y = h - 30
                c.setFont("Helvetica", 8)
            c.drawString(30, y, _fmt_br_data(r["date"]))
            c.drawString(95, y, str(r["driver_name"])[:28])
            c.drawString(265, y, str(r["client_name"])[:32])
            c.drawString(450, y, str(r["status"])[:10])
            c.drawRightString(590, y, _fmt_br_2(r.get("planned_value", 0)))
            y -= 10
        c.save()
        pbuf.seek(0)
        return StreamingResponse(pbuf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=bi_entregas_{stamp}.pdf"})

    return JSONResponse({"error": "Formato invalido. Use csv, xlsx ou pdf."}, status_code=400)


@router.get("/bi/clientes/export")
async def bi_clientes_export(
    format: str = "csv",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    shift: str = "Todos",
    driver_id: Optional[str] = None,
    plate: str = "Todos",
    status: str = "Todos",
    client_id: Optional[str] = None,
    city: str = "Todos",
    priority: str = "Todos",
    client_status: str = "Todos",
    segmentos: Optional[list[str]] = Query(default=None),
    returns_filter: str = "Todos",
    client_scope: Optional[str] = None,
    vendedor_id: Optional[str] = None,
    search_q: str = "",
    classification_filter: str = "Todos",
    motivo_filter: str = "Todos",
    responsabilidade_filter: str = "Todos",
    purchase_band: str = "Todos",
    return_band: str = "Todos",
    duration_band: str = "Todos",
    session: Session = Depends(get_session),
):
    parsed_driver_id: Optional[int] = int(driver_id) if (driver_id or "").strip().isdigit() else None
    parsed_client_id: Optional[int] = int(client_id) if (client_id or "").strip().isdigit() else None
    parsed_vendedor_id: Optional[int] = None
    if (vendedor_id or "").strip() == "-1":
        parsed_vendedor_id = -1
    elif (vendedor_id or "").strip().isdigit():
        parsed_vendedor_id = int(vendedor_id)
    dataset = _build_bi_clientes_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        shift=shift,
        driver_id=parsed_driver_id,
        plate=plate,
        status=status,
        client_id=parsed_client_id,
        city=city,
        priority=priority,
        client_status=client_status,
        segmentos=segmentos,
        returns_filter=returns_filter,
        client_filter_scope=(client_scope or "solo").strip().lower(),
        vendedor_id=parsed_vendedor_id,
        search_q=search_q,
        classification_filter=classification_filter,
        motivo_filter=motivo_filter,
        responsabilidade_filter=responsabilidade_filter,
        purchase_band=purchase_band,
        return_band=return_band,
        duration_band=duration_band,
    )
    rows = dataset["all_client_rows"]
    stamp = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d_%H%M")
    fmt = (format or "csv").strip().lower()

    def _client_row_data_br(row: dict) -> list:
        return [
            row.get("client_id") or "",
            row.get("client_code") or "",
            row.get("client_name") or "",
            row.get("vendedor_name") or "",
            row.get("city") or "",
            row.get("bairro") or "",
            row.get("segmento") or "",
            row.get("prioridade") or "",
            row.get("status_operacional") or "",
            row.get("visits") or 0,
            row.get("weekly_peak_visits") or 0,
            _fmt_br_1(row.get("total_duration_m") or 0),
            _fmt_br_1(row.get("avg_duration_m") or 0),
            row.get("returned_occurrences") or 0,
            _fmt_br_2(row.get("planned_value") or 0),
            _fmt_br_2(row.get("delivered_value") or 0),
            _fmt_br_2(row.get("returned_value") or 0),
            _fmt_br_1(row.get("return_rate_qtd") or 0),
            _fmt_br_1(row.get("return_pct_planned") or row.get("return_rate_value") or 0),
            row.get("reopen_count") or 0,
            row.get("top_driver_name") or "-",
            row.get("top_motivo_name") or "-",
            row.get("classification_title") or "",
            row.get("cliente_score") if row.get("cliente_score") is not None else "",
            row.get("risk_label") or "-",
            row.get("risk_score") or 0,
        ]

    if fmt == "csv":
        out = io.StringIO()
        writer = csv.writer(out, delimiter=";")
        writer.writerow(
            [
                "cliente_id",
                "codigo_nb",
                "cliente",
                "vendedor",
                "cidade",
                "bairro",
                "segmento",
                "prioridade",
                "status_operacional",
                "visitas",
                "pico_semanal",
                "tempo_total_min",
                "tempo_medio_min",
                "devolucoes",
                "valor_planejado",
                "valor_entregue",
                "valor_devolvido",
                "devolucao_pct_qtd",
                "devolucao_pct_sobre_planejado",
                "reaberturas",
                "motorista_principal",
                "motivo_principal",
                "classificacao",
                "score_cliente",
                "risco",
                "score_risco",
            ]
        )
        for row in rows:
            writer.writerow(_client_row_data_br(row))
        buf = io.BytesIO(out.getvalue().encode("utf-8-sig"))
        return StreamingResponse(
            buf,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=bi_clientes_{stamp}.csv"},
        )

    if fmt == "xlsx":
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BI Clientes"
        sheet.append(
            [
                "Cliente ID",
                "Codigo NB",
                "Cliente",
                "Vendedor",
                "Cidade",
                "Bairro",
                "Segmento",
                "Prioridade",
                "Status Operacional",
                "Visitas",
                "Pico Semanal",
                "Tempo Total (min)",
                "Tempo Medio (min)",
                "Devolucoes",
                "Valor Planejado",
                "Valor Entregue",
                "Valor Devolvido",
                "Devolucao % Qtd",
                "Devolucao % sobre planejado",
                "Reaberturas",
                "Motorista Principal",
                "Motivo Principal",
                "Classificacao",
                "Score Cliente",
                "Risco",
                "Score de Risco",
            ]
        )
        for row in rows:
            sheet.append(_client_row_data_br(row))
        xbuf = io.BytesIO()
        workbook.save(xbuf)
        xbuf.seek(0)
        return StreamingResponse(
            xbuf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=bi_clientes_{stamp}.xlsx"},
        )

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        pbuf = io.BytesIO()
        c = canvas.Canvas(pbuf, pagesize=A4)
        _, h = A4
        y = h - 40
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y, "BI Clientes - Exportacao")
        y -= 16
        c.setFont("Helvetica", 9)
        period_from = _fmt_br_data(dataset["filters"].get("date_from"))
        period_to = _fmt_br_data(dataset["filters"].get("date_to"))
        c.drawString(30, y, f"Periodo: {period_from} ate {period_to}")
        y -= 18
        c.setFont("Helvetica", 8)
        for row in rows[:100]:
            if y <= 40:
                c.showPage()
                y = h - 40
                c.setFont("Helvetica", 8)
            nm = str(row.get("client_name") or "")[:40]
            c.drawString(30, y, nm)
            c.drawRightString(560, y, _fmt_br_2(row.get("delivered_value") or 0))
            y -= 10
        c.save()
        pbuf.seek(0)
        return StreamingResponse(
            pbuf,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=bi_clientes_{stamp}.pdf"},
        )

    return JSONResponse({"error": "Formato invalido. Use csv, xlsx ou pdf."}, status_code=400)


# ---------------------------------------------------------------------------
# BI DEVOLUÇÕES — Mega BI Page
# ---------------------------------------------------------------------------

def _build_bi_devolucoes_dataset(
    session: Session,
    date_from: Optional[str],
    date_to: Optional[str],
    responsabilidade_id: Optional[int] = None,
    motivo_id: Optional[int] = None,
    motorista_id: Optional[int] = None,
    vendedor_id: Optional[int] = None,
    client_id: Optional[int] = None,
    client_filter_scope: str = "solo",
    valor_faixa: Optional[str] = None,
    somente_criticas: bool = False,
    somente_acima_meta: bool = False,
) -> dict:
    tz = ZoneInfo("America/Sao_Paulo")
    today = datetime.now(tz).date()

    def _d(raw: Optional[str]) -> Optional[date]:
        if not raw:
            return None
        try:
            return datetime.strptime(raw, "%Y-%m-%d").date()
        except Exception:
            return None

    date_i = _d(date_from) or today.replace(day=1)
    date_f = _d(date_to) or today
    if date_i > date_f:
        date_i, date_f = date_f, date_i

    def _is_full_calendar_month(di: date, df: date) -> bool:
        if di.day != 1 or di > df:
            return False
        last = (
            (date(di.year + 1, 1, 1) - timedelta(days=1))
            if di.month == 12
            else (date(di.year, di.month + 1, 1) - timedelta(days=1))
        )
        return df == last

    # Mês civil completo no filtro → alinhar ao mês comercial (ex.: maio/2026 começa em 05/05).
    if _is_full_calendar_month(date_i, date_f):
        ps, pe = commercial_competence_period_iso_bounds(date_i.year, date_i.month)
        date_i = datetime.strptime(ps, "%Y-%m-%d").date()
        date_f = datetime.strptime(pe, "%Y-%m-%d").date()

    period_start, period_end, window_start, window_end = _competence_period_window(date_i, date_f)
    date_str_i = window_start
    date_str_f = window_end

    # --- carregar cadastros ---
    motivos_all = session.exec(select(models.DevolucaoMotivo)).all()
    resps_all = session.exec(select(models.DevolucaoResponsabilidade)).all()
    mot_map = {m.id: m for m in motivos_all}
    rsp_map = {r.id: r for r in resps_all}

    # --- query principal (janela alargada; corte real = competência em [period_start, period_end]) ---
    q = (
        select(models.Devolucao)
        .where(
            or_(
                and_(
                    models.Devolucao.data_romaneio >= window_start,
                    models.Devolucao.data_romaneio <= window_end,
                ),
                and_(
                    models.Devolucao.data_entrega >= window_start,
                    models.Devolucao.data_entrega <= window_end,
                ),
            )
        )
    )
    if responsabilidade_id:
        q = q.where(models.Devolucao.responsabilidade_id == responsabilidade_id)
    if motivo_id:
        q = q.where(models.Devolucao.motivo_id == motivo_id)
    if motorista_id:
        q = q.where(models.Devolucao.motorista_id == motorista_id)
    if vendedor_id:
        q = q.where(models.Devolucao.vendedor_id == vendedor_id)
    cfs = (client_filter_scope or "solo").strip().lower()
    if cfs not in ("solo", "group"):
        cfs = "solo"
    client_ids_dev: Optional[list] = None
    group_filter_note_dev: Optional[str] = None
    if client_id:
        if cfs == "group":
            c = session.get(models.Client, int(client_id))
            if c and getattr(c, "client_group_id", None):
                client_ids_dev = [int(x) for x in session.exec(select(models.Client.id).where(models.Client.client_group_id == c.client_group_id)).all() if x is not None]
            if not client_ids_dev:
                client_ids_dev = [int(client_id)]
            elif len(client_ids_dev) == 1:
                group_filter_note_dev = "Só há uma loja neste grupo."
        else:
            client_ids_dev = [int(client_id)]
    if client_ids_dev:
        if len(client_ids_dev) == 1:
            q = q.where(models.Devolucao.client_id == client_ids_dev[0])
        else:
            q = q.where(models.Devolucao.client_id.in_(client_ids_dev))

    devs_raw = session.exec(q.order_by(models.Devolucao.data_romaneio.desc())).all()
    # Alinhar ao consolidado / KPI mensal (main._kpi_devolucao_mes_romaneio_calendario): não somar duplicatas.
    devs_raw = [d for d in devs_raw if not getattr(d, "duplicate_of_id", None)]

    def _dev_chart_day_iso(d: models.Devolucao) -> str:
        """
        Dia civil no eixo do gráfico e agregações diárias.

        Prioriza datas que caiam em [period_start, period_end] (eixo do BI), para não perder
        ocorrências com romaneio antes do 1º dia civil do mês comercial ou fora do recorte.
        Se romaneio e entrega estão na janela e diferem, usa romaneio (logística).
        """
        raw_rom = str(getattr(d, "data_romaneio", None) or "").strip()
        raw_ent = str(getattr(d, "data_entrega", None) or "").strip()
        rom = raw_rom[:10] if len(raw_rom) >= 10 else ""
        ent = raw_ent[:10] if len(raw_ent) >= 10 else ""
        comp = devolucao_competencia_iso(d) or ""

        def _inwin(p: str) -> bool:
            return len(p) == 10 and period_start <= p <= period_end

        if _inwin(rom) and _inwin(ent) and rom != ent:
            return rom
        if _inwin(rom):
            return rom
        if _inwin(ent):
            return ent
        if len(comp) >= 10 and _inwin(comp[:10]):
            return comp[:10]
        if len(rom) == 10:
            return rom
        if len(ent) == 10:
            return ent
        return comp[:10] if len(comp) >= 10 else ""

    devs_c = [d for d in devs_raw if devolucao_competencia_in_period(d, period_start, period_end)]
    devs_c.sort(
        key=lambda x: (_dev_chart_day_iso(x), str(getattr(x, "data_romaneio", None) or "")),
        reverse=True,
    )

    def _valor_faixa_ok(d: models.Devolucao, faixa: str) -> bool:
        v = float(d.valor or 0)
        f = (faixa or "all").strip().lower()
        if f in ("", "all", "todos"):
            return True
        if f == "ate100":
            return v <= 100
        if f in ("100_300", "v100_300"):
            return 100 < v <= 300
        if f in ("300_800", "v300_800"):
            return 300 < v <= 800
        if f in ("acima800", "v800"):
            return v > 800
        return True

    devs_work: List[models.Devolucao] = list(devs_c)
    devs_work = [d for d in devs_work if _valor_faixa_ok(d, valor_faixa or "all")]
    if somente_criticas:
        devs_work = [
            d
            for d in devs_work
            if float(d.valor or 0) >= 800.0 or (str(d.acima_300 or "").upper() == "SIM")
        ]

    devs: List[models.Devolucao] = list(devs_work)

    def _parse_route_helper_ids(helpers_json: Optional[str]) -> List[int]:
        """Parse delivery_helpers_json da rota para lista de employee_id."""
        if not helpers_json:
            return []
        try:
            data = json.loads(helpers_json) if isinstance(helpers_json, str) else helpers_json
            if not isinstance(data, list):
                return []
            return [int(x) for x in data if x is not None and str(x).strip().isdigit()]
        except Exception:
            return []

    def _parse_helpers_to_ids(helpers_json: Optional[str], emp_by_name: dict) -> List[int]:
        """Parse JSON (ids ou nomes) para lista de employee_id. emp_by_name: nome_lower -> id."""
        if not helpers_json or not emp_by_name:
            return []
        try:
            data = json.loads(helpers_json) if isinstance(helpers_json, str) else helpers_json
            if not isinstance(data, list):
                return []
            ids = []
            for h in data:
                if h is None:
                    continue
                if isinstance(h, int) and h > 0:
                    ids.append(h)
                elif isinstance(h, str) and str(h).strip().isdigit():
                    ids.append(int(h.strip()))
                elif isinstance(h, str) and (h or "").strip():
                    eid = emp_by_name.get((h or "").strip().lower())
                    if eid and eid not in ids:
                        ids.append(eid)
            return ids
        except Exception:
            return []

    # Mapa nome -> id para resolver ajudantes por nome (Route/Session podem enviar nomes)
    all_employees = list(session.exec(select(models.Employee)).all())
    emp_by_name: dict = {e.name.strip().lower(): e.id for e in all_employees if e and getattr(e, "name", None) and getattr(e, "id", None)}

    # Ajudantes das rotas vinculadas (para devoluções sem ajudante_id preenchido)
    route_ids = sorted({d.route_id for d in devs if getattr(d, "route_id", None)})
    route_helpers: dict = {}  # route_id -> [emp_id, ...]
    route_by_id: dict[int, models.Route] = {}
    if route_ids:
        try:
            routes_linked = session.exec(
                select(models.Route).where(models.Route.id.in_(route_ids))
            ).all()
            for r in routes_linked:
                if getattr(r, "id", None) is not None:
                    route_by_id[int(r.id)] = r
                raw = getattr(r, "delivery_helpers_json", None)
                ids = _parse_route_helper_ids(raw) or _parse_helpers_to_ids(raw, emp_by_name)
                if ids:
                    route_helpers[r.id] = ids
        except Exception:
            pass

    # Fallback: ajudantes por (client_id, motorista_id, data) quando devolução não tem route_id
    # Busca rotas de entrega no período com helpers e monta lookup para casar com devoluções
    route_by_client_driver_date: dict = {}  # (client_id, employee_id, date_str) -> [helper_ids]
    try:
        routes_in_range = session.exec(
            select(models.Route)
            .where(models.Route.date >= date_str_i)
            .where(models.Route.date <= date_str_f)
            .where(models.Route.client_id.is_not(None))
            .where(models.Route.employee_id.is_not(None))
        ).all()
        for r in routes_in_range:
            raw = getattr(r, "delivery_helpers_json", None)
            ids = _parse_route_helper_ids(raw) or _parse_helpers_to_ids(raw, emp_by_name)
            if ids and r.client_id and r.employee_id:
                key = (r.client_id, r.employee_id, str(r.date)[:10])
                prev = route_by_client_driver_date.get(key)
                route_by_client_driver_date[key] = merge_unique_helper_id_lists(prev, ids)
    except Exception:
        pass

    # Fallback 2: ajudantes da sessão (date, motorista_id) quando rota não tem delivery_helpers_json
    session_helpers_by_driver_date: dict = {}  # (date_str, employee_id) -> [helper_ids]
    try:
        sessions_in_range = session.exec(
            select(models.DeliverySession)
            .where(models.DeliverySession.date >= date_str_i)
            .where(models.DeliverySession.date <= date_str_f)
        ).all()
        for ds in sessions_in_range:
            raw = getattr(ds, "helpers_json", None)
            ids = _parse_route_helper_ids(raw) or _parse_helpers_to_ids(raw, emp_by_name)
            if ids and ds.employee_id:
                key = (str(getattr(ds, "date", "") or "")[:10], ds.employee_id)
                prev = session_helpers_by_driver_date.get(key)
                session_helpers_by_driver_date[key] = merge_unique_helper_id_lists(prev, ids)
    except Exception:
        pass

    # --- mapas de lookup ---
    helper_ids_from_routes = set()
    for ids in route_helpers.values():
        helper_ids_from_routes.update(ids)
    for ids in route_by_client_driver_date.values():
        helper_ids_from_routes.update(ids)
    for ids in session_helpers_by_driver_date.values():
        helper_ids_from_routes.update(ids)
    emp_ids = sorted({d.motorista_id for d in devs if d.motorista_id} |
                     {d.ajudante_id for d in devs if d.ajudante_id} |
                     {d.vendedor_id for d in devs if d.vendedor_id} |
                     helper_ids_from_routes)
    cli_ids = sorted({d.client_id for d in devs if d.client_id})
    emp_map = {e.id: e for e in (session.exec(select(models.Employee).where(models.Employee.id.in_(emp_ids))).all() if emp_ids else [])}
    cli_map = {c.id: c for c in (session.exec(select(models.Client).where(models.Client.id.in_(cli_ids))).all() if cli_ids else [])}

    # --- filtros para a UI ---
    drivers_filter = sorted(
        [e for e in session.exec(select(models.Employee)).all()
         if any(d.motorista_id == e.id for d in devs_c)],
        key=lambda e: e.name,
    )
    clients_filter = sorted(
        [c for c in session.exec(select(models.Client)).all()
         if any(d.client_id == c.id for d in devs_c)],
        key=lambda c: c.name,
    )
    vendedores_filter = sorted(
        [e for e in session.exec(select(models.Employee)).all()
         if any(d.vendedor_id == e.id for d in devs_c)],
        key=lambda e: e.name,
    )

    # Mesmo critério do painel TV / `main.devolucao_mes`: rotas cuja competência da Route.date cai no período;
    # % financeiro = valor devolvido (cadastro, competência) ÷ base R$ (`pct_valor_devolvido_sobre_base_rotas`).
    rq_base = (
        select(models.Route)
        .where(models.Route.type == "delivery")
        .where(models.Route.date >= window_start)
        .where(models.Route.date <= window_end)
    )
    if motorista_id:
        rq_base = rq_base.where(models.Route.employee_id == motorista_id)
    if client_ids_dev:
        if len(client_ids_dev) == 1:
            rq_base = rq_base.where(models.Route.client_id == client_ids_dev[0])
        else:
            rq_base = rq_base.where(models.Route.client_id.in_(client_ids_dev))
    routes_delivery_period = session.exec(rq_base).all()
    routes_delivery_period = [
        r for r in routes_delivery_period
        if route_competencia_in_period(r, period_start, period_end)
    ]
    pct_devolucao_rotas = pct_devolucao_sobre_rotas_concluidas(routes_delivery_period)

    # Base financeira do KPI por dia de competência (projeção fim de mês / MTD por competência).
    receita_por_dia_comp: dict[str, float] = {}
    # Mesma base R$ agregada pelo dia civil da rota (Route.date), alinhada ao eixo do gráfico de evolução.
    receita_por_dia_operacional: dict[str, float] = {}
    for _r_fin in routes_delivery_period:
        _op_r = route_competencia_operacional_iso(_r_fin)
        if len(_op_r) < 10:
            continue
        _base_d = route_base_financeiro_kpi(_r_fin)
        if _base_d is None:
            continue
        receita_por_dia_comp[_op_r] = receita_por_dia_comp.get(_op_r, 0.0) + float(_base_d)
        dcal = str(getattr(_r_fin, "date", None) or "").strip()[:10]
        if len(dcal) == 10:
            receita_por_dia_operacional[dcal] = receita_por_dia_operacional.get(dcal, 0.0) + float(_base_d)

    devs_pre_acima: List[models.Devolucao] = list(devs)
    dev_por_dia_mtd: dict[str, float] = {}
    for _d_k in devs_pre_acima:
        _op_k = devolucao_competencia_iso(_d_k)
        if not (len(_op_k) == 10 and period_start <= _op_k <= period_end):
            continue
        dev_por_dia_mtd[_op_k] = dev_por_dia_mtd.get(_op_k, 0.0) + float(_d_k.valor or 0)
    total_valor_kpi = sum(float(d.valor or 0) for d in devs_pre_acima)
    total_qtd_kpi = len(devs_pre_acima)
    standalone_base_kpi = sum(
        float(d.valor or 0)
        for d in devs_pre_acima
        if getattr(d, "route_id", None) is None
    )
    pct_devolucao_financeiro, valor_base_rotas = pct_valor_devolvido_sobre_base_rotas(
        float(total_valor_kpi),
        routes_delivery_period,
        suplemento_base_financeira=standalone_base_kpi,
    )
    meta_pp = 2.0
    valor_meta_permitido: Optional[float] = (
        round(valor_base_rotas * (meta_pp / 100.0), 2) if valor_base_rotas > 0 else None
    )
    excedente_sobre_meta: Optional[float] = (
        round(max(0.0, total_valor_kpi - (valor_meta_permitido or 0.0)), 2)
        if valor_meta_permitido is not None
        else None
    )
    desvio_pp: Optional[float] = (
        round(float(pct_devolucao_financeiro) - meta_pp, 2) if pct_devolucao_financeiro is not None else None
    )
    # Meta de 2%: compara % valor (financeiro) quando há base de rotas; senão % paradas concluídas.
    desvio_rotas_pp: float = round(float(pct_devolucao_rotas or 0.0) - float(meta_pp), 1)
    n_ret_r, n_done_r = counts_devolucao_rotas_concluidas(routes_delivery_period)
    situacao_meta = "desconhecido"
    if pct_devolucao_financeiro is not None and valor_base_rotas > 0:
        pr_fin = float(pct_devolucao_financeiro)
        situacao_meta = "dentro" if pr_fin <= meta_pp else "acima"
    elif n_done_r > 0:
        pr_rot = float(pct_devolucao_rotas or 0.0)
        situacao_meta = "dentro" if pr_rot <= meta_pp else "acima"
    faixa_alerta_meta = "neutral"
    if pct_devolucao_financeiro is not None and valor_base_rotas > 0:
        prf = float(pct_devolucao_financeiro)
        if prf <= meta_pp:
            faixa_alerta_meta = "ok"
        elif prf <= 2.5:
            faixa_alerta_meta = "warn"
        else:
            faixa_alerta_meta = "danger"
    elif n_done_r > 0:
        pr_rot = float(pct_devolucao_rotas or 0.0)
        if pr_rot <= meta_pp:
            faixa_alerta_meta = "ok"
        elif pr_rot <= 2.5:
            faixa_alerta_meta = "warn"
        else:
            faixa_alerta_meta = "danger"

    # Projeção % financeiro ao fim do mês (sazonalidade por dia da semana no histórico pré-mês)
    projecao_mes_fim: Optional[dict] = None
    if (date_i.year, date_i.month) == (date_f.year, date_f.month):
        m_last_d = monthrange(date_f.year, date_f.month)[1]
        month_last = date(date_f.year, date_f.month, m_last_d)
        month_first = date(date_f.year, date_f.month, 1)
        if date_f < month_last:
            bs_start = month_first - timedelta(days=140)
            bs_end = month_first - timedelta(days=1)
            if bs_end >= bs_start:
                hist_ws = (bs_start - timedelta(days=10)).strftime("%Y-%m-%d")
                hist_we = (bs_end + timedelta(days=10)).strftime("%Y-%m-%d")
                rq_hist = (
                    select(models.Route)
                    .where(models.Route.type == "delivery")
                    .where(models.Route.date >= hist_ws)
                    .where(models.Route.date <= hist_we)
                )
                if motorista_id:
                    rq_hist = rq_hist.where(models.Route.employee_id == motorista_id)
                if client_ids_dev:
                    if len(client_ids_dev) == 1:
                        rq_hist = rq_hist.where(models.Route.client_id == client_ids_dev[0])
                    else:
                        rq_hist = rq_hist.where(models.Route.client_id.in_(client_ids_dev))
                routes_hist_raw = session.exec(rq_hist).all()
                bs_period_start, bs_period_end, bs_window_start, bs_window_end = _competence_period_window(
                    bs_start, bs_end
                )
                routes_baseline = [
                    r for r in routes_hist_raw if route_competencia_in_period(r, bs_period_start, bs_period_end)
                ]
                base_por_dia_baseline: dict[str, float] = {}
                for _r_bl in routes_baseline:
                    _op_bl = route_competencia_operacional_iso(_r_bl)
                    if len(_op_bl) < 10:
                        continue
                    _base_bl = route_base_financeiro_kpi(_r_bl)
                    if _base_bl is None:
                        continue
                    base_por_dia_baseline[_op_bl] = base_por_dia_baseline.get(_op_bl, 0.0) + float(_base_bl)
                q_hist_dev = (
                    select(models.Devolucao)
                    .where(
                        or_(
                            and_(
                                models.Devolucao.data_romaneio >= bs_window_start,
                                models.Devolucao.data_romaneio <= bs_window_end,
                            ),
                            and_(
                                models.Devolucao.data_entrega >= bs_window_start,
                                models.Devolucao.data_entrega <= bs_window_end,
                            ),
                        )
                    )
                )
                if responsabilidade_id:
                    q_hist_dev = q_hist_dev.where(models.Devolucao.responsabilidade_id == responsabilidade_id)
                if motivo_id:
                    q_hist_dev = q_hist_dev.where(models.Devolucao.motivo_id == motivo_id)
                if motorista_id:
                    q_hist_dev = q_hist_dev.where(models.Devolucao.motorista_id == motorista_id)
                if vendedor_id:
                    q_hist_dev = q_hist_dev.where(models.Devolucao.vendedor_id == vendedor_id)
                if client_ids_dev:
                    if len(client_ids_dev) == 1:
                        q_hist_dev = q_hist_dev.where(models.Devolucao.client_id == client_ids_dev[0])
                    else:
                        q_hist_dev = q_hist_dev.where(models.Devolucao.client_id.in_(client_ids_dev))
                devs_hist_raw = session.exec(q_hist_dev.order_by(models.Devolucao.data_romaneio.desc())).all()
                devs_hist_raw = [d for d in devs_hist_raw if not getattr(d, "duplicate_of_id", None)]
                devs_hist_c = [
                    d for d in devs_hist_raw if devolucao_competencia_in_period(d, bs_period_start, bs_period_end)
                ]
                devs_hist_work = [d for d in devs_hist_c if _valor_faixa_ok(d, valor_faixa or "all")]
                if somente_criticas:
                    devs_hist_work = [
                        d
                        for d in devs_hist_work
                        if float(d.valor or 0) >= 800.0 or (str(d.acima_300 or "").upper() == "SIM")
                    ]
                dev_por_dia_baseline: dict[str, float] = {}
                for _d_bl in devs_hist_work:
                    _op_db = devolucao_competencia_iso(_d_bl)
                    if not (len(_op_db) == 10 and bs_period_start <= _op_db <= bs_period_end):
                        continue
                    dev_por_dia_baseline[_op_db] = dev_por_dia_baseline.get(_op_db, 0.0) + float(_d_bl.valor or 0)

                projecao_mes_fim = build_mes_fim_projecao_pct_financeiro(
                    date_i,
                    date_f,
                    receita_por_dia_comp,
                    dev_por_dia_mtd,
                    base_por_dia_baseline,
                    dev_por_dia_baseline,
                    meta_pp=meta_pp,
                )

    acima_meta_filter_note: Optional[str] = None
    ids_listagem: set[int] = {d.id for d in devs_pre_acima if getattr(d, "id", None)}
    if somente_acima_meta:
        if pct_devolucao_financeiro is None or pct_devolucao_financeiro <= meta_pp:
            ids_listagem = set()
            acima_meta_filter_note = (
                "Filtro «somente acima da meta» ativo: o período está dentro ou na meta financeira de 2% "
                "sobre o valor base das rotas (ou não há valor base). A lista foi esvaziada."
            )
        else:
            devs_alto = [d for d in devs_pre_acima if float(d.valor or 0) >= 300.0]
            if devs_alto:
                ids_listagem = {d.id for d in devs_alto if getattr(d, "id", None)}
            else:
                acima_meta_filter_note = "Nenhuma ocorrência ≥ R$ 300; exibindo todas as devoluções do período."
    devs = list(devs_pre_acima)
    # --- agregações (gráficos e rankings usam o período completo após faixa/críticas; lista respeita ids_listagem) ---
    dev_count_by_client: dict[int, int] = {}
    for _dx in devs_pre_acima:
        if getattr(_dx, "client_id", None):
            _cid = int(_dx.client_id)
            dev_count_by_client[_cid] = dev_count_by_client.get(_cid, 0) + 1

    hist_route_visits: dict[int, int] = {}
    try:
        _rq_hist = (
            select(models.Route.client_id, func.count(models.Route.id))
            .where(models.Route.type == "delivery")
            .where(models.Route.date >= period_start)
            .where(models.Route.date <= period_end)
        )
        if motorista_id:
            _rq_hist = _rq_hist.where(models.Route.employee_id == motorista_id)
        if client_ids_dev:
            if len(client_ids_dev) == 1:
                _rq_hist = _rq_hist.where(models.Route.client_id == client_ids_dev[0])
            else:
                _rq_hist = _rq_hist.where(models.Route.client_id.in_(client_ids_dev))
        _rq_hist = _rq_hist.group_by(models.Route.client_id)
        for _row in session.exec(_rq_hist).all():
            _cid_r, _cnt_r = _row[0], _row[1]
            if _cid_r is not None:
                hist_route_visits[int(_cid_r)] = int(_cnt_r or 0)
    except Exception:
        pass

    total_qtd = int(total_qtd_kpi)
    total_valor = float(total_valor_kpi)

    # Cores em hexadecimal (layout claro GA; sem depender de --color-* do tema escuro)
    _RESP_HEX_DETAIL = {"MERCADO": "#dc2626", "COMERCIAL": "#d97706"}
    _resp_hex_default = "#2563eb"

    def _resp_hex(nome: str) -> str:
        n = (nome or "").upper()
        return next((v for k, v in _RESP_HEX_DETAIL.items() if k in n), _resp_hex_default)

    def _resp_tom(nome: str) -> str:
        """Classe CSS semântica para selo de responsabilidade."""
        n = (nome or "").upper()
        if "MERCADO" in n:
            return "mercado"
        if "COMERCIAL" in n:
            return "comercial"
        if "LOG" in n:
            return "logistica"
        return "outro"

    # por dia
    per_day: dict[str, dict] = {}
    # por semana (ISO)
    per_week: dict[str, dict] = {}
    # por motivo
    per_motivo: dict[str, dict] = {}
    # por responsabilidade
    per_resp: dict[str, dict] = {}
    # drill responsabilidade -> motivos (para modal ao clicar no card)
    per_resp_motivo: dict[str, dict[str, dict]] = {}
    # por cluster
    per_cluster: dict[str, dict] = {}
    # por motorista
    per_motorista: dict[str, dict] = {}
    # por vendedor
    per_vendedor: dict[str, dict] = {}
    # drill vendedor: responsabilidade e motivos por vendedor
    per_vendedor_resp: dict[str, dict[str, dict]] = {}  # vendedor_nome -> resp_nome -> { responsabilidade, qtd, valor }
    per_vendedor_motivo: dict[str, dict[str, dict]] = {}  # vendedor_nome -> motivo_nome -> { motivo, qtd, valor }
    # por cliente
    per_cliente: dict[str, dict] = {}
    # heatmap dia-da-semana (0=Seg..6=Dom) x semana ISO
    heatmap_week: dict[str, dict[int, int]] = {}  # week_label -> {weekday: count}
    # heatmap dia-da-semana x hora_do_dia (não temos hora, então dia-da-semana x dia-do-mês para enriquecer)
    heatmap_dow_dom: dict[int, dict[int, int]] = {i: {} for i in range(7)}  # dow -> {dom: count}

    DOW_LABELS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

    rows_detail: list[dict] = []

    def _classificacao_impacto_valor(v: float) -> str:
        if v <= 100:
            return "Baixo impacto"
        if v <= 300:
            return "Médio impacto"
        if v <= 800:
            return "Alto impacto"
        return "Crítico"

    def _acao_corretiva_sugerida(resp_nome: str) -> str:
        n = (resp_nome or "").upper()
        if "COMERCIAL" in n:
            return "Revisar pedido, cadastro e alinhamento com o cliente antes da expedição."
        if "LOG" in n:
            return "Auditar separação, carga e conferência na entrega."
        if "MERCADO" in n:
            return "Verificar qualidade percebida e acordo comercial de troca."
        return "Registrar causa raiz e acionar o responsável pela área."

    for d in devs_pre_acima:
        motivo = mot_map.get(d.motivo_id)
        resp = rsp_map.get(d.responsabilidade_id)
        cli = cli_map.get(d.client_id)
        motorista = emp_map.get(d.motorista_id)
        ajudante = emp_map.get(d.ajudante_id) if d.ajudante_id else None
        vendedor = emp_map.get(d.vendedor_id) if d.vendedor_id else None

        motivo_nome = motivo.nome if motivo else "Não informado"
        resp_nome = resp.nome if resp else "Não informado"
        cli_nome = cli.name if cli else f"Cliente #{d.client_id}"
        motorista_nome = motorista.name if motorista else f"Motorista #{d.motorista_id}"
        # Ajudante: Devolucao.ajudante_id, ou rota vinculada (route_id), ou rota casada por (cliente, motorista, data)
        if ajudante:
            ajudante_nome = ajudante.name
        else:
            helper_ids = None
            if getattr(d, "route_id", None) and route_helpers.get(d.route_id):
                helper_ids = route_helpers[d.route_id]
            if not helper_ids and d.client_id and d.motorista_id and d.data_romaneio:
                dt_key = str(d.data_romaneio)[:10]
                key = (d.client_id, d.motorista_id, dt_key)
                helper_ids = route_by_client_driver_date.get(key)
            if not helper_ids and d.motorista_id and d.data_romaneio:
                dt_key = str(d.data_romaneio)[:10]
                session_key = (dt_key, d.motorista_id)
                helper_ids = session_helpers_by_driver_date.get(session_key)
            ajudante_id_from_route = helper_ids[0] if helper_ids else None
            if ajudante_id_from_route and ajudante_id_from_route == d.motorista_id and len(helper_ids) > 1:
                ajudante_id_from_route = helper_ids[1]
            elif ajudante_id_from_route == d.motorista_id:
                ajudante_id_from_route = None
            emp_ajud = emp_map.get(ajudante_id_from_route) if ajudante_id_from_route else None
            ajudante_nome = emp_ajud.name if emp_ajud else "—"
        vendedor_nome = vendedor.name if vendedor else "—"

        val = float(d.valor or 0)
        op_raw = _dev_chart_day_iso(d)
        dt_str = devolucao_competencia_iso(d) or ""
        dt_operacional = op_raw if len(op_raw) >= 10 else (dt_str if len(dt_str) >= 10 else "")

        # per_day (eixo = dia calendário da operação)
        slot = per_day.setdefault(dt_operacional, {"data": dt_operacional, "qtd": 0, "valor": 0.0})
        slot["qtd"] += 1
        slot["valor"] = round(slot["valor"] + val, 2)

        # per_week / heatmaps: a partir da data operacional
        try:
            _dt = datetime.strptime(dt_operacional, "%Y-%m-%d").date()
            iso = _dt.isocalendar()
            wk = f"{iso.year}-S{iso.week:02d}"
            dow = _dt.weekday()
            dom = _dt.day
        except Exception:
            wk = "Sem semana"
            dow = 0
            dom = 1
            _dt = date_i

        ws = per_week.setdefault(wk, {"semana": wk, "qtd": 0, "valor": 0.0})
        ws["qtd"] += 1
        ws["valor"] = round(ws["valor"] + val, 2)

        # per_motivo
        ms = per_motivo.setdefault(motivo_nome, {"motivo": motivo_nome, "qtd": 0, "valor": 0.0})
        ms["qtd"] += 1
        ms["valor"] = round(ms["valor"] + val, 2)

        # per_resp
        rs = per_resp.setdefault(resp_nome, {"responsabilidade": resp_nome, "qtd": 0, "valor": 0.0})
        rs["qtd"] += 1
        rs["valor"] = round(rs["valor"] + val, 2)
        # per_resp_motivo (drill: motivos por responsabilidade)
        prm = per_resp_motivo.setdefault(resp_nome, {})
        smr = prm.setdefault(motivo_nome, {"motivo": motivo_nome, "qtd": 0, "valor": 0.0})
        smr["qtd"] += 1
        smr["valor"] = round(smr["valor"] + val, 2)

        # per_cluster
        cluster = str(d.cluster or "Sem cluster")
        cs = per_cluster.setdefault(cluster, {"cluster": cluster, "qtd": 0, "valor": 0.0})
        cs["qtd"] += 1
        cs["valor"] = round(cs["valor"] + val, 2)

        # per_motorista
        mts = per_motorista.setdefault(motorista_nome, {
            "motorista": motorista_nome, "qtd": 0, "valor": 0.0,
            "acima300": 0,
        })
        mts["qtd"] += 1
        mts["valor"] = round(mts["valor"] + val, 2)
        if (d.acima_300 or "NAO") == "SIM":
            mts["acima300"] += 1

        # per_vendedor
        vs = per_vendedor.setdefault(vendedor_nome, {"vendedor": vendedor_nome, "qtd": 0, "valor": 0.0})
        vs["qtd"] += 1
        vs["valor"] = round(vs["valor"] + val, 2)

        # per_vendedor_resp (drill: responsabilidade por vendedor)
        pr = per_vendedor_resp.setdefault(vendedor_nome, {})
        sr = pr.setdefault(resp_nome, {"responsabilidade": resp_nome, "qtd": 0, "valor": 0.0})
        sr["qtd"] += 1
        sr["valor"] = round(sr["valor"] + val, 2)
        # per_vendedor_motivo (drill: motivos por vendedor)
        pm = per_vendedor_motivo.setdefault(vendedor_nome, {})
        sm = pm.setdefault(motivo_nome, {"motivo": motivo_nome, "qtd": 0, "valor": 0.0})
        sm["qtd"] += 1
        sm["valor"] = round(sm["valor"] + val, 2)

        # per_cliente
        cls_ = per_cliente.setdefault(
            cli_nome,
            {"cliente": cli_nome, "qtd": 0, "valor": 0.0, "motivos": {}, "ultima_data": dt_operacional},
        )
        cls_["qtd"] += 1
        cls_["valor"] = round(cls_["valor"] + val, 2)
        cls_["motivos"][motivo_nome] = cls_["motivos"].get(motivo_nome, 0) + 1
        if dt_operacional > cls_.get("ultima_data", ""):
            cls_["ultima_data"] = dt_operacional

        # heatmap semanal
        hw = heatmap_week.setdefault(wk, {i: 0 for i in range(7)})
        hw[dow] = hw.get(dow, 0) + 1

        # heatmap dia-da-semana × dia-do-mês
        heatmap_dow_dom[dow][dom] = heatmap_dow_dom[dow].get(dom, 0) + 1

        did = getattr(d, "id", None)
        if did is not None and did in ids_listagem:
            obs_parts = [
                (getattr(d, "observacao", None) or "").strip(),
                (getattr(d, "observacao_gestor", None) or "").strip(),
            ]
            obs_txt = " — ".join([p for p in obs_parts if p]) or "—"
            pct_imp = round(100.0 * val / total_valor_kpi, 2) if total_valor_kpi > 0 else 0.0
            nb_raw = (str(cli.nb).strip() if cli and getattr(cli, "nb", None) else "") or ""
            rt = route_by_id.get(int(d.route_id)) if getattr(d, "route_id", None) else None
            peso_kg: Optional[float] = None
            if rt is not None:
                try:
                    wv = getattr(rt, "devolucao_volume", None)
                    if wv is None:
                        wv = getattr(rt, "tonnage", None)
                    if wv is not None:
                        peso_kg = round(float(wv), 3)
                except Exception:
                    peso_kg = None
            created_at = getattr(d, "created_at", None)
            rows_detail.append({
                "id": did,
                "data": dt_operacional,
                "data_competencia": dt_str,
                "client_id": int(d.client_id) if getattr(d, "client_id", None) else None,
                "client_nb": nb_raw or None,
                "client_nb_fmt": _fmt_nb_br(nb_raw) if nb_raw else "—",
                "hist_rotas_entrega_periodo": (
                    hist_route_visits.get(int(d.client_id), 0) if getattr(d, "client_id", None) else 0
                ),
                "hist_devolucoes_cliente_periodo": (
                    dev_count_by_client.get(int(d.client_id), 0) if getattr(d, "client_id", None) else 0
                ),
                "cliente": cli_nome,
                "peso_kg": peso_kg,
                "peso_kg_fmt": _fmt_br_kg(peso_kg) if peso_kg is not None else "—",
                "registrado_em_br": _fmt_br_datetime_local(created_at) if created_at else "—",
                "vendedor": vendedor_nome,
                "motorista": motorista_nome,
                "ajudante": ajudante_nome,
                "motivo": motivo_nome,
                "responsabilidade": resp_nome,
                "responsabilidade_hex": _resp_hex(resp_nome),
                "responsabilidade_tom": _resp_tom(resp_nome),
                "cluster": cluster,
                "valor": val,
                "acima_300": d.acima_300 or "NAO",
                "semana": d.semana or 0,
                "dia": d.dia or 0,
                "source": d.source or "—",
                "observacao": obs_txt,
                "pct_impacto": pct_imp,
                "impacto_classificacao": _classificacao_impacto_valor(val),
                "acao_corretiva": _acao_corretiva_sugerida(resp_nome),
                "status_operacional": (
                    "Crítica"
                    if (d.acima_300 or "").upper() == "SIM"
                    else ("Pendente" if (str(d.source or "").upper() == "EXCEL") else "Resolvida")
                ),
            })

    # --- top N (prioriza valor financeiro) ---
    top_clientes = sorted(per_cliente.values(), key=lambda x: x["valor"], reverse=True)[:20]
    top_motoristas = sorted(per_motorista.values(), key=lambda x: x["valor"], reverse=True)[:20]
    top_motivos = sorted(per_motivo.values(), key=lambda x: x["valor"], reverse=True)[:15]
    top_vendedores = sorted(per_vendedor.values(), key=lambda x: x["valor"], reverse=True)[:15]
    top_clusters = sorted(per_cluster.values(), key=lambda x: x["valor"], reverse=True)[:15]

    # --- devoluções evitadas (registro manual no período filtrado) ---
    ev_q = (
        select(models.DevolucaoEvitada)
        .where(models.DevolucaoEvitada.event_date >= period_start)
        .where(models.DevolucaoEvitada.event_date <= period_end)
    )
    if client_ids_dev:
        if len(client_ids_dev) == 1:
            ev_q = ev_q.where(models.DevolucaoEvitada.client_id == client_ids_dev[0])
        else:
            ev_q = ev_q.where(models.DevolucaoEvitada.client_id.in_(client_ids_dev))
    evitadas_db = list(
        session.exec(ev_q.order_by(models.DevolucaoEvitada.event_date.desc(), models.DevolucaoEvitada.id.desc())).all()
    )
    ev_cli_ids = sorted({int(e.client_id) for e in evitadas_db if e.client_id})
    for cid in ev_cli_ids:
        if cid not in cli_map:
            oc = session.get(models.Client, cid)
            if oc:
                cli_map[cid] = oc
    evitadas_by_day: dict[str, dict] = {}
    per_tipo_ev: dict[str, dict] = {}
    for ev in evitadas_db:
        ds_ev = str(getattr(ev, "event_date", None) or "").strip()[:10]
        if len(ds_ev) < 10:
            continue
        slot_e = evitadas_by_day.setdefault(ds_ev, {"qtd": 0, "valor": 0.0})
        slot_e["qtd"] += 1
        slot_e["valor"] = round(slot_e["valor"] + float(getattr(ev, "valor_estimado", None) or 0.0), 2)
        tk = str(getattr(ev, "tipo", None) or "outros").strip().lower() or "outros"
        te = per_tipo_ev.setdefault(tk, {"tipo": tk, "label": label_tipo_evitada(tk), "qtd": 0})
        te["qtd"] += 1
    total_evitadas_qtd = len(evitadas_db)
    total_evitadas_valor_est = round(sum(float(getattr(ev, "valor_estimado", None) or 0.0) for ev in evitadas_db), 2)
    evitadas_clientes_distintos = len({int(e.client_id) for e in evitadas_db if e.client_id})
    evitadas_lancamentos: list[dict] = []
    for ev in evitadas_db[:120]:
        cli_ev = cli_map.get(int(ev.client_id))
        nome_cli = cli_ev.name if cli_ev else f"Cliente #{ev.client_id}"
        evitadas_lancamentos.append(
            {
                "id": int(ev.id) if ev.id is not None else None,
                "event_date": str(ev.event_date)[:10],
                "client_id": int(ev.client_id),
                "cliente": nome_cli,
                "tipo": str(ev.tipo or ""),
                "tipo_label": label_tipo_evitada(str(ev.tipo or "")),
                "observacao": (getattr(ev, "observacao", None) or "").strip(),
                "valor_estimado": round(float(getattr(ev, "valor_estimado", None) or 0.0), 2)
                if getattr(ev, "valor_estimado", None) is not None
                else None,
                "created_by": (getattr(ev, "created_by", None) or "").strip() or "—",
                "created_at": getattr(ev, "created_at", None).isoformat(timespec="seconds")
                if getattr(ev, "created_at", None)
                else "",
            }
        )
    top_tipos_evitadas = sorted(per_tipo_ev.values(), key=lambda x: int(x.get("qtd") or 0), reverse=True)

    # Drill-down por vendedor: responsabilidade e motivos para detalhe no clique
    vendedor_drill: dict[str, dict] = {}
    for vendedor_nome in per_vendedor.keys():
        resp_list = sorted(
            per_vendedor_resp.get(vendedor_nome, {}).values(),
            key=lambda x: x["qtd"],
            reverse=True,
        )
        motivos_list = sorted(
            per_vendedor_motivo.get(vendedor_nome, {}).values(),
            key=lambda x: x["qtd"],
            reverse=True,
        )[:15]
        vendedor_drill[vendedor_nome] = {
            "responsabilidade": resp_list,
            "motivos": motivos_list,
        }

    # Evolução diária: todos os dias do filtro (inclui dias sem devolução), meta = 2% da base financeira do KPI naquele dia
    def _date_str_range_inclusive(s: str, e: str) -> List[str]:
        out: List[str] = []
        try:
            s10 = (s or "").strip()[:10]
            e10 = (e or "").strip()[:10]
            if len(s10) < 10 or len(e10) < 10:
                return []
            di = datetime.strptime(s10, "%Y-%m-%d").date()
            df = datetime.strptime(e10, "%Y-%m-%d").date()
            if di > df:
                di, df = df, di
            cur = di
            while cur <= df:
                out.append(cur.strftime("%Y-%m-%d"))
                cur += timedelta(days=1)
        except Exception:
            return []
        return out

    days_sorted = _date_str_range_inclusive(period_start, period_end)
    if not days_sorted and date_i and date_f:
        cur = date_i
        while cur <= date_f:
            days_sorted.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
    evolucao_diaria: list[dict] = []
    for ds in days_sorted:
        slot_dev = per_day.get(ds, {"data": ds, "qtd": 0, "valor": 0.0})
        rec = round(float(receita_por_dia_operacional.get(ds, 0.0)), 2)
        meta_2 = round(0.02 * rec, 2) if rec > 0 else None
        val_fl = float(slot_dev.get("valor") or 0)
        pct_dia = round(100.0 * val_fl / rec, 2) if rec > 0 else None
        ev_slot = evitadas_by_day.get(ds, {"qtd": 0, "valor": 0.0})
        evolucao_diaria.append(
            {
                "data": ds,
                "qtd": int(slot_dev.get("qtd") or 0),
                "valor": round(val_fl, 2),
                "receita_base": rec,
                "meta_2pct_valor": meta_2,
                "pct_devolucao_dia": pct_dia,
                "evitadas_qtd": int(ev_slot.get("qtd") or 0),
                "evitadas_valor_est": round(float(ev_slot.get("valor") or 0.0), 2),
            }
        )
    # Linha constante (meta ÷ dias corridos) substituída por meta proporcional à receita de cada dia — manter null no JSON
    meta_valor_dia_ref: Optional[float] = None

    # Evolução semanal
    weeks_sorted = sorted(per_week.keys())
    evolucao_semanal = [per_week[k] for k in weeks_sorted]

    # responsabilidade breakdown — cor em hex para gráficos e cartões
    for _rk, _rv in per_resp.items():
        _rv["color"] = _resp_hex(_rk)
        _rv["tom"] = _resp_tom(_rk)
    resp_breakdown = sorted(per_resp.values(), key=lambda x: x["valor"], reverse=True)

    # Drill-down por responsabilidade: motivos com qtd, valor e % (para modal ao clicar no card)
    resp_drill: dict[str, dict] = {}
    for resp_nome, resp_totals in per_resp.items():
        motivos_raw = list(per_resp_motivo.get(resp_nome, {}).values())
        total_qtd_resp = resp_totals["qtd"]
        total_valor_resp = resp_totals["valor"]
        motivos_list = []
        for m in sorted(motivos_raw, key=lambda x: x["qtd"], reverse=True):
            pct_qtd = round(100.0 * m["qtd"] / total_qtd_resp, 1) if total_qtd_resp else 0
            pct_valor = round(100.0 * m["valor"] / total_valor_resp, 1) if total_valor_resp else 0
            motivos_list.append({
                "motivo": m["motivo"],
                "qtd": m["qtd"],
                "valor": m["valor"],
                "pct_qtd": pct_qtd,
                "pct_valor": pct_valor,
            })
        resp_drill[resp_nome] = {"responsabilidade": resp_nome, "motivos": motivos_list}

    # heatmap: matriz DOW (0-6) × semanas
    weeks_label = weeks_sorted[-12:] if len(weeks_sorted) > 12 else weeks_sorted
    heatmap_matrix = []
    for dow in range(7):
        row_vals = []
        for wk in weeks_label:
            cnt = heatmap_week.get(wk, {}).get(dow, 0)
            row_vals.append(cnt)
        heatmap_matrix.append({"dow": DOW_LABELS[dow], "values": row_vals})

    # heatmap DOM × DOW (dia-do-mês como eixo Y, dia-da-semana como eixo X)
    heatmap_dom_dow = []
    for dom in range(1, 32):
        vals = [heatmap_dow_dom[dow].get(dom, 0) for dow in range(7)]
        if any(v > 0 for v in vals):
            heatmap_dom_dow.append({"dom": dom, "values": vals})

    # Faixa ao lado da evolução semanal: totais por dia da semana no período (todas as semanas ISO).
    dow_short_lbl = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    dow_tot_period = [0] * 7
    for _wk, dow_map in heatmap_week.items():
        for di in range(7):
            dow_tot_period[di] += int(dow_map.get(di, 0) or 0)
    max_cdow = max(dow_tot_period) if dow_tot_period else 0
    heatmap_dow_strip: list[dict[str, Any]] = []
    for i in range(7):
        cnt = int(dow_tot_period[i])
        heatmap_dow_strip.append({
            "dow": dow_short_lbl[i],
            "dow_full": DOW_LABELS[i],
            "count": cnt,
            "intensity": round((float(cnt) / float(max_cdow)) if max_cdow > 0 else 0.0, 4),
        })

    # acima_300 breakdown
    total_acima_300 = sum(1 for d in devs_pre_acima if (d.acima_300 or "NAO") == "SIM")
    pct_acima_300 = round(total_acima_300 / total_qtd * 100, 1) if total_qtd else 0.0

    # média por dia (dias com devolução)
    media_por_dia = round(total_qtd / len(days_sorted), 1) if days_sorted else 0.0
    media_valor_dia = round(total_valor / len(days_sorted), 2) if days_sorted else 0.0

    analise_destaque: list[str] = []
    if total_valor_kpi > 0 and top_motivos:
        m_fin = max(top_motivos, key=lambda x: float(x.get("valor") or 0))
        vm = float(m_fin.get("valor") or 0)
        if vm > 0:
            analise_destaque.append(
                f"O principal impacto financeiro veio do motivo {m_fin.get('motivo', '—')}, somando {_fmt_br_moeda(vm)}."
            )
    if total_valor_kpi > 0 and resp_breakdown:
        r_fin = max(resp_breakdown, key=lambda x: float(x.get("valor") or 0))
        rv = float(r_fin.get("valor") or 0)
        pct_r = round(100.0 * rv / total_valor_kpi, 1) if total_valor_kpi else 0.0
        if rv > 0:
            analise_destaque.append(
                f"A área {r_fin.get('responsabilidade', '—')} concentra {pct_r:.1f}% do valor devolvido no período."
            )
    crit_n = sum(1 for d in devs_pre_acima if float(d.valor or 0) >= 800.0)
    crit_v = sum(float(d.valor or 0) for d in devs_pre_acima if float(d.valor or 0) >= 800.0)
    if crit_n > 0 and total_valor_kpi > 0:
        pct_crit_v = round(100.0 * crit_v / total_valor_kpi, 1)
        if crit_n <= max(1, int(total_qtd_kpi * 0.25)) and pct_crit_v >= 30.0:
            analise_destaque.append(
                f"As devoluções acima de R$ 800 somam {crit_n} ocorrência(s), mas concentram cerca de {pct_crit_v:.0f}% do valor."
            )
    if desvio_pp is not None and desvio_pp > 0 and pct_devolucao_financeiro is not None:
        analise_destaque.append(
            f"Sobre valor base de rotas, o período está {desvio_pp:.2f} pontos percentuais acima da meta de 2%."
        )
    elif desvio_pp is not None and desvio_pp < 0 and pct_devolucao_financeiro is not None:
        analise_destaque.append(
            f"Sobre valor base de rotas, o período está {abs(desvio_pp):.2f} pontos percentuais abaixo da meta de 2%."
        )
    if n_done_r > 0 and desvio_rotas_pp > 0:
        analise_destaque.append(
            f"Sobre rotas concluídas (critério TV), o período está {desvio_rotas_pp:.1f} pontos percentuais acima da meta de 2%."
        )
    if excedente_sobre_meta is not None and excedente_sobre_meta > 0:
        analise_destaque.append(
            f"O excedente financeiro sobre a meta permitida é de {_fmt_br_moeda(excedente_sobre_meta)}."
        )
    if float(pct_devolucao_rotas or 0) >= 2.0 and pct_devolucao_financeiro is None:
        analise_destaque.append(
            f"Indicador operacional (paradas): {float(pct_devolucao_rotas):.1f}% de devoluções sobre rotas concluídas — "
            "sem valor base financeiro para calcular % sobre faturamento."
        )
    if total_evitadas_qtd > 0:
        _ev_msg = (
            f"Foram registradas {total_evitadas_qtd} devolução(ões) evitada(s) no período "
            f"({evitadas_clientes_distintos} cliente(s) distinto(s))"
        )
        if total_evitadas_valor_est and total_evitadas_valor_est > 0:
            _ev_msg += f", com valor estimado total de {_fmt_br_moeda(total_evitadas_valor_est)}."
        else:
            _ev_msg += "."
        analise_destaque.append(_ev_msg)
    if not analise_destaque:
        if total_qtd_kpi == 0:
            if total_evitadas_qtd > 0:
                analise_destaque.append(
                    "Sem devoluções realizadas no período com os filtros aplicados; há registros de devoluções evitadas."
                )
            else:
                analise_destaque.append("Sem devoluções no período com os filtros aplicados.")
        else:
            analise_destaque.append("Use os rankings e a lista para priorizar ações corretivas no período.")

    for m in top_motivos:
        m["pct_valor_total"] = round(100.0 * float(m["valor"]) / total_valor_kpi, 1) if total_valor_kpi else 0.0
    for r in resp_breakdown:
        r["pct_valor_total"] = round(100.0 * float(r["valor"]) / total_valor_kpi, 1) if total_valor_kpi else 0.0

    # Resumo semanal (accordion) e mini-cards do gráfico
    semana_resumo: Optional[str] = None
    if evolucao_semanal:
        nsem = len(evolucao_semanal)
        worst_w = max(evolucao_semanal, key=lambda x: float(x.get("valor") or 0))
        vw = float(worst_w.get("valor") or 0)
        if vw > 0:
            semana_resumo = (
                f"{nsem} semana(s) analisadas · pior semana: {worst_w.get('semana', '—')} · {_fmt_br_moeda(vw)}"
            )
        else:
            semana_resumo = f"{nsem} semana(s) analisadas"

    chart_insights: dict = {}
    if evolucao_diaria:
        wd = max(evolucao_diaria, key=lambda x: float(x.get("valor") or 0))
        if float(wd.get("valor") or 0) > 0:
            chart_insights["pior_dia"] = wd.get("data")
            chart_insights["pior_dia_valor"] = round(float(wd.get("valor") or 0), 2)
    if top_motivos:
        chart_insights["top_motivo"] = str(top_motivos[0].get("motivo") or "")
        chart_insights["top_motivo_valor"] = round(float(top_motivos[0].get("valor") or 0), 2)
        chart_insights["top_motivo_pct"] = float(top_motivos[0].get("pct_valor_total") or 0)
    if total_qtd:
        chart_insights["media_por_devolucao"] = round(float(total_valor) / float(total_qtd), 2)
    chart_insights["pct_ocorrencias_acima800"] = float(pct_acima_300)
    if total_valor_kpi and crit_v > 0:
        chart_insights["pct_valor_acima800"] = round(100.0 * float(crit_v) / float(total_valor_kpi), 1)

    filters_query = urlencode({
        k: str(v) for k, v in {
            "date_from": date_from or "",
            "date_to": date_to or "",
            "responsabilidade_id": responsabilidade_id or "",
            "motivo_id": motivo_id or "",
            "motorista_id": motorista_id or "",
            "vendedor_id": vendedor_id or "",
            "client_id": client_id or "",
            "client_scope": cfs if client_id else "",
            "valor_faixa": (valor_faixa or "").strip().lower() if valor_faixa else "",
            "criticas": "1" if somente_criticas else "",
            "acima_meta": "1" if somente_acima_meta else "",
        }.items() if v not in ("", None, False, 0)
    })

    return {
        "filters": {
            "date_from": date_from or date_i.strftime("%Y-%m-%d"),
            "date_to": date_to or date_f.strftime("%Y-%m-%d"),
            "responsabilidade_id": responsabilidade_id,
            "motivo_id": motivo_id,
            "motorista_id": motorista_id,
            "vendedor_id": vendedor_id,
            "client_id": client_id,
            "client_scope": cfs if client_id else "solo",
            "valor_faixa": (valor_faixa or "all").strip().lower(),
            "somente_criticas": somente_criticas,
            "somente_acima_meta": somente_acima_meta,
            "group_filter_note": group_filter_note_dev,
            "acima_meta_filter_note": acima_meta_filter_note,
        },
        "filters_query": filters_query,
        # KPIs
        "total_qtd": total_qtd,
        "total_valor": total_valor,
        "total_acima_300": total_acima_300,
        "pct_acima_300": pct_acima_300,
        "media_por_dia": media_por_dia,
        "media_valor_dia": media_valor_dia,
        "total_clientes_afetados": len(per_cliente),
        "total_motoristas_envolvidos": len(per_motorista),
        "valor_base_rotas": round(valor_base_rotas, 2),
        "valor_base_disponivel": bool(valor_base_rotas and valor_base_rotas > 0),
        "pct_devolucao_rotas": pct_devolucao_rotas,
        "pct_devolucao_financeiro": pct_devolucao_financeiro,
        "meta_pp": meta_pp,
        "valor_meta_permitido": valor_meta_permitido,
        "excedente_sobre_meta": excedente_sobre_meta,
        "desvio_pp": desvio_pp,
        "desvio_rotas_pp": desvio_rotas_pp,
        "rotas_kpi_devolucao": int(n_ret_r),
        "rotas_kpi_concluidas": int(n_done_r),
        "situacao_meta": situacao_meta,
        "faixa_alerta_meta": faixa_alerta_meta,
        "projecao_mes_fim": projecao_mes_fim,
        "meta_valor_dia_ref": meta_valor_dia_ref,
        "generated_at": datetime.now(tz).strftime("%d/%m/%Y %H:%M"),
        # evolução
        "evolucao_diaria": evolucao_diaria,
        "evolucao_semanal": evolucao_semanal,
        # breakdowns
        "resp_breakdown": resp_breakdown,
        "resp_drill": resp_drill,
        "top_motivos": top_motivos,
        "top_clientes": top_clientes,
        "top_motoristas": top_motoristas,
        "top_vendedores": top_vendedores,
        "top_clusters": top_clusters,
        # heatmap
        "heatmap_matrix": heatmap_matrix,
        "heatmap_weeks_labels": weeks_label,
        "heatmap_dom_dow": heatmap_dom_dow,
        "heatmap_dow_strip": heatmap_dow_strip,
        "dow_labels": DOW_LABELS,
        # tabela detalhada (últimos 500 da query)
        "rows_detail": rows_detail[:500],
        # select de filtros
        "motivos_filter": sorted(motivos_all, key=lambda m: m.nome),
        "resps_filter": sorted(resps_all, key=lambda r: r.nome),
        "drivers_filter": drivers_filter,
        "clients_filter": clients_filter,
        "vendedores_filter": vendedores_filter,
        # json para charts
        "evolucao_diaria_json": _json_for_inline_script(evolucao_diaria),
        "evolucao_semanal_json": _json_for_inline_script(evolucao_semanal),
        "top_motivos_json": _json_for_inline_script(top_motivos),
        "resp_breakdown_json": _json_for_inline_script(resp_breakdown),
        "resp_drill_json": _json_for_inline_script(resp_drill),
        "top_motoristas_json": _json_for_inline_script(top_motoristas),
        "top_clientes_json": _json_for_inline_script(top_clientes),
        "top_vendedores_json": _json_for_inline_script(top_vendedores),
        "vendedor_drill_json": _json_for_inline_script(vendedor_drill),
        "top_clusters_json": _json_for_inline_script(top_clusters),
        "heatmap_matrix_json": _json_for_inline_script(heatmap_matrix),
        "heatmap_weeks_labels_json": _json_for_inline_script(weeks_label),
        "heatmap_dom_dow_json": _json_for_inline_script(heatmap_dom_dow),
        "dow_labels_json": _json_for_inline_script(DOW_LABELS),
        "rows_detail_json": _json_for_inline_script(rows_detail[:500]),
        "chart_insights": chart_insights,
        "semana_resumo": semana_resumo,
        "analise_destaque": analise_destaque,
        # devoluções evitadas (manual)
        "total_evitadas_qtd": total_evitadas_qtd,
        "total_evitadas_valor_est": total_evitadas_valor_est,
        "evitadas_clientes_distintos": evitadas_clientes_distintos,
        "evitadas_lancamentos": evitadas_lancamentos,
        "top_tipos_evitadas": top_tipos_evitadas,
        "evitada_tipos_select": [{"id": k, "label": EVITADA_TIPO_LABELS[k]} for k in EVITADA_TIPOS_ORDENADOS],
        "evitada_default_date": datetime.now(tz).strftime("%Y-%m-%d"),
        "evitadas_lancamentos_json": _json_for_inline_script(evitadas_lancamentos),
        "top_tipos_evitadas_json": _json_for_inline_script(top_tipos_evitadas),
    }


@router.get("/bi/devolucoes", response_class=HTMLResponse)
async def bi_devolucoes_page(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    responsabilidade_id: Optional[str] = None,
    motivo_id: Optional[str] = None,
    motorista_id: Optional[str] = None,
    vendedor_id: Optional[str] = None,
    client_id: Optional[str] = None,
    client_scope: Optional[str] = None,
    valor_faixa: Optional[str] = None,
    criticas: Optional[str] = None,
    acima_meta: Optional[str] = None,
    session: Session = Depends(get_session),
):
    parsed_resp_id: Optional[int] = int(responsabilidade_id) if (responsabilidade_id or "").strip().isdigit() else None
    parsed_motivo_id: Optional[int] = int(motivo_id) if (motivo_id or "").strip().isdigit() else None
    parsed_motorista_id: Optional[int] = int(motorista_id) if (motorista_id or "").strip().isdigit() else None
    parsed_vendedor_id: Optional[int] = int(vendedor_id) if (vendedor_id or "").strip().isdigit() else None
    parsed_client_id: Optional[int] = int(client_id) if (client_id or "").strip().isdigit() else None
    somente_criticas = (criticas or "").strip().lower() in ("1", "true", "on", "yes", "sim")
    somente_acima_meta = (acima_meta or "").strip().lower() in ("1", "true", "on", "yes", "sim")
    dataset = _build_bi_devolucoes_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        responsabilidade_id=parsed_resp_id,
        motivo_id=parsed_motivo_id,
        motorista_id=parsed_motorista_id,
        vendedor_id=parsed_vendedor_id,
        client_id=parsed_client_id,
        client_filter_scope=(client_scope or "solo").strip().lower(),
        valor_faixa=valor_faixa,
        somente_criticas=somente_criticas,
        somente_acima_meta=somente_acima_meta,
    )
    return templates.TemplateResponse("bi_devolucoes_refactor.html", {"request": request, **dataset})


@router.get("/bi/devolucoes/export")
async def bi_devolucoes_export(
    format: str = "csv",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    responsabilidade_id: Optional[str] = None,
    motivo_id: Optional[str] = None,
    motorista_id: Optional[str] = None,
    vendedor_id: Optional[str] = None,
    client_id: Optional[str] = None,
    client_scope: Optional[str] = None,
    valor_faixa: Optional[str] = None,
    criticas: Optional[str] = None,
    acima_meta: Optional[str] = None,
    session: Session = Depends(get_session),
):
    parsed_resp_id: Optional[int] = int(responsabilidade_id) if (responsabilidade_id or "").strip().isdigit() else None
    parsed_motivo_id: Optional[int] = int(motivo_id) if (motivo_id or "").strip().isdigit() else None
    parsed_motorista_id: Optional[int] = int(motorista_id) if (motorista_id or "").strip().isdigit() else None
    parsed_vendedor_id: Optional[int] = int(vendedor_id) if (vendedor_id or "").strip().isdigit() else None
    parsed_client_id: Optional[int] = int(client_id) if (client_id or "").strip().isdigit() else None
    somente_criticas = (criticas or "").strip().lower() in ("1", "true", "on", "yes", "sim")
    somente_acima_meta = (acima_meta or "").strip().lower() in ("1", "true", "on", "yes", "sim")
    dataset = _build_bi_devolucoes_dataset(
        session=session,
        date_from=date_from,
        date_to=date_to,
        responsabilidade_id=parsed_resp_id,
        motivo_id=parsed_motivo_id,
        motorista_id=parsed_motorista_id,
        vendedor_id=parsed_vendedor_id,
        client_id=parsed_client_id,
        client_filter_scope=(client_scope or "solo").strip().lower(),
        valor_faixa=valor_faixa,
        somente_criticas=somente_criticas,
        somente_acima_meta=somente_acima_meta,
    )
    rows = dataset["rows_detail"]
    stamp = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y%m%d_%H%M")
    fmt = (format or "csv").strip().lower()

    headers_csv = [
        "data",
        "cliente",
        "nb",
        "vendedor",
        "motorista",
        "ajudante",
        "motivo",
        "responsabilidade",
        "cluster",
        "peso_kg",
        "valor",
        "registrado_em",
        "acima_300",
        "source",
    ]

    if fmt == "csv":
        out = io.StringIO()
        writer = csv.writer(out, delimiter=";")
        writer.writerow(headers_csv)
        for r in rows:
            writer.writerow([
                _fmt_br_data(r.get("data") or ""),
                r.get("cliente") or "",
                r.get("client_nb_fmt") or r.get("client_nb") or "",
                r.get("vendedor") or "",
                r.get("motorista") or "",
                r.get("ajudante") or "",
                r.get("motivo") or "",
                r.get("responsabilidade") or "",
                r.get("cluster") or "",
                r.get("peso_kg_fmt") or "—",
                _fmt_br_2(r.get("valor") or 0),
                r.get("registrado_em_br") or "—",
                r.get("acima_300") or "NAO",
                r.get("source") or "",
            ])
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename=bi_devolucoes_{stamp}.csv"},
        )

    if fmt == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws_sheet = wb.active
            ws_sheet.title = "Devoluções"
            ws_sheet.append(headers_csv)
            for r in rows:
                ws_sheet.append([
                    _fmt_br_data(r.get("data") or ""),
                    r.get("cliente") or "",
                    r.get("client_nb_fmt") or r.get("client_nb") or "",
                    r.get("vendedor") or "",
                    r.get("motorista") or "",
                    r.get("ajudante") or "",
                    r.get("motivo") or "",
                    r.get("responsabilidade") or "",
                    r.get("cluster") or "",
                    r.get("peso_kg_fmt") or "—",
                    float(r.get("valor") or 0),
                    r.get("registrado_em_br") or "—",
                    r.get("acima_300") or "NAO",
                    r.get("source") or "",
                ])
            xbuf = io.BytesIO()
            wb.save(xbuf)
            xbuf.seek(0)
            return StreamingResponse(
                iter([xbuf.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=bi_devolucoes_{stamp}.xlsx"},
            )
        except ImportError:
            pass

    return JSONResponse({"error": "Formato invalido. Use csv ou xlsx."}, status_code=400)
